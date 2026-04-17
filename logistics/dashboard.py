# dashboard.py
# Location: C:\Users\a.harayri\OneDrive - Al-Behar Group\Desktop\python\dashboard.py

import pandas as pd
import numpy as np
import json
import webbrowser
import os
from datetime import datetime
import warnings
import base64
import re
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

# Custom JSON encoder for numpy types and Python booleans
class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (np.integer, np.int64, np.int32)):
            return int(obj)
        elif isinstance(obj, (np.floating, np.float64, np.float32)):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, pd.Timestamp):
            return obj.strftime('%Y-%m-%d')
        elif isinstance(obj, bool):
            return obj
        elif isinstance(obj, (np.bool_)):
            return bool(obj)
        elif pd.isna(obj):
            return None
        return super().default(obj)

def get_logo_base64():
    """Convert logo image to base64 for embedding in HTML"""
    logo_path = r'C:\Users\a.harayri\OneDrive - Al-Behar Group\Desktop\python\img.jpg'
    
    if os.path.exists(logo_path):
        try:
            with open(logo_path, 'rb') as img_file:
                return base64.b64encode(img_file.read()).decode('utf-8')
        except Exception as e:
            print(f"Warning: Could not read logo file: {e}")
            return None
    else:
        print(f"Warning: Logo not found at: {logo_path}")
        return None

def extract_month_from_filename(filename):
    """Extract month number from S&OP filename (S&OP1 = January, etc.)"""
    if pd.isna(filename) or not isinstance(filename, str):
        return None
    
    match = re.search(r'S&OP(\d+)', filename, re.IGNORECASE)
    if match:
        month_num = int(match.group(1))
        if 1 <= month_num <= 12:
            return month_num
    return None

def get_month_name(month_num):
    """Convert month number to month name"""
    month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    if month_num and 1 <= month_num <= 12:
        return month_names[month_num - 1]
    return None

def get_month_abbreviation(month_num):
    """Convert month number to 3-letter abbreviation (Jan, Feb, etc.)"""
    month_abbr = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    if month_num and 1 <= month_num <= 12:
        return month_abbr[month_num - 1]
    return None

def clean_filename_for_key(filename):
    """Clean filename to create a valid dictionary key"""
    if not isinstance(filename, str):
        return str(filename)
    return re.sub(r'[^a-zA-Z0-9]', '_', filename)

def apply_tracker_updates(excel_path):
    """Apply tracker updates from CSV to the Data sheet before loading."""
    updates_path = os.path.join(os.path.dirname(excel_path), 'tracker_updates.csv')
    if not os.path.exists(updates_path):
        return

    print("\nApplying tracker updates from tracker_updates.csv ...")
    try:
        updates_df = pd.read_csv(updates_path, encoding='utf-8-sig')
        if updates_df.empty:
            print("   tracker_updates.csv is empty, nothing to apply")
            return

        wb = load_workbook(excel_path)
        if 'Data' not in wb.sheetnames:
            print("   Data sheet not found, tracker updates skipped")
            return

        ws = wb['Data']
        header_map = {}
        for cell in ws[1]:
            header_value = str(cell.value).strip() if cell.value is not None else ''
            if header_value:
                header_map[header_value] = cell.column

        required_headers = ['Status', 'Actual Month', 'Notes', 'Reason']
        for header in required_headers:
            if header not in header_map:
                print(f"   Missing '{header}' column in Data sheet, tracker updates skipped")
                return

        applied_count = 0
        for _, row in updates_df.iterrows():
            excel_row = pd.to_numeric(row.get('excel_row'), errors='coerce')
            if pd.isna(excel_row):
                continue
            excel_row = int(excel_row)
            if excel_row < 2:
                continue

            action = str(row.get('action', '')).strip().lower()
            new_status = str(row.get('new_status', '')).strip()
            new_month = str(row.get('new_month', '')).strip()
            new_notes = str(row.get('new_notes', '')).strip()
            new_reason = str(row.get('new_reason', '')).strip()

            if action == 'remove':
                ws.cell(row=excel_row, column=header_map['Status']).value = 'Removed'
            elif new_status:
                ws.cell(row=excel_row, column=header_map['Status']).value = new_status

            if new_month:
                ws.cell(row=excel_row, column=header_map['Actual Month']).value = new_month

            if new_notes:
                ws.cell(row=excel_row, column=header_map['Notes']).value = new_notes

            if new_reason:
                ws.cell(row=excel_row, column=header_map['Reason']).value = new_reason

            applied_count += 1

        wb.save(excel_path)
        archive_name = f"tracker_updates_applied_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        os.replace(updates_path, os.path.join(os.path.dirname(excel_path), archive_name))
        print(f"   Applied {applied_count} tracker updates to Data sheet")
        print(f"   Archived updates file as {archive_name}")
    except Exception as e:
        print(f"   Could not apply tracker updates: {e}")

# Fixed monthly targets from user
MONTHLY_TARGETS = {
    1: {'month': 'Jan', 'gp': 70955.6889, 'sales': 1340123.793, 'cogs': -1269168.104, 'commission': 3906.42, 'qty': 1371.74},
    2: {'month': 'Feb', 'gp': 109681.2675, 'sales': 1865415.873, 'cogs': -1755734.605, 'commission': 9053.4443, 'qty': 1730},
    3: {'month': 'Mar', 'gp': 89767.79265, 'sales': 2255162.847, 'cogs': -2165395.054, 'commission': 9520.0066, 'qty': 2084.71},
    4: {'month': 'Apr', 'gp': 122249.2298, 'sales': 2329650.58, 'cogs': -2207401.35, 'commission': 7330.4163, 'qty': 2198.465},
    5: {'month': 'May', 'gp': 139237.0366, 'sales': 2254543.159, 'cogs': -2115306.122, 'commission': 8535.2863, 'qty': 2143.86},
    6: {'month': 'Jun', 'gp': 169113.3709, 'sales': 3296326.514, 'cogs': -3127213.143, 'commission': 8717.4368, 'qty': 2982.04},
    7: {'month': 'Jul', 'gp': 147454.483, 'sales': 2324379.15, 'cogs': -2176924.667, 'commission': 15160.0265, 'qty': 2200.96},
    8: {'month': 'Aug', 'gp': 126528.9325, 'sales': 2418520.393, 'cogs': -2291991.46, 'commission': 7019.628, 'qty': 2309.1},
    9: {'month': 'Sep', 'gp': 150593.5497, 'sales': 2611603.899, 'cogs': -2461010.349, 'commission': 9231.846, 'qty': 2183.04},
    10: {'month': 'Oct', 'gp': 83173.66766, 'sales': 1925567.18, 'cogs': -1842393.512, 'commission': 9770.878, 'qty': 1783.21},
    11: {'month': 'Nov', 'gp': 100630.7955, 'sales': 1951227.219, 'cogs': -1850596.423, 'commission': 7409.019202, 'qty': 1923.19},
    12: {'month': 'Dec', 'gp': 103734.6638, 'sales': 1755026.813, 'cogs': -1651292.149, 'commission': 9347.25295, 'qty': 1601.27}
}

def get_status_for_month(selected_sop_month, current_month, dataset_type='actual'):
    """Determine status based on S&OP month selection"""
    if dataset_type == 'actual':
        if current_month <= selected_sop_month:
            return ['Invoiced']
        else:
            return ['Booked', 'Not Booked']
    elif dataset_type == 'forecast':
        return ['Booked', 'Not Booked']
    else:
        return ['Invoiced', 'Booked', 'Not Booked']

def calculate_commission(df_data, month_abbr, file_name):
    """Calculate commission from Broker section ONLY"""
    if 'Section' in df_data.columns and 'Actual Month' in df_data.columns and 'gp_jod' in df_data.columns:
        commission_df = df_data[(df_data['Section'] == 'Broker') & (df_data['Actual Month'] == month_abbr)]
        if len(commission_df) > 0:
            return float(commission_df['gp_jod'].sum())
    return 0.0

# Professional HTML Template
HTML_TEMPLATE_START = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Al-Behar Company - S&OP Dashboard</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.1/font/bootstrap-icons.css">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/flatpickr/dist/flatpickr.min.css">
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }

        body {
            font-family: 'Inter', sans-serif;
            background: #f4f7fc;
            color: #1e293b;
            font-size: 0.875rem;
            line-height: 1.5;
        }

        .dashboard-wrapper {
            max-width: 1600px;
            margin: 0 auto;
            padding: 10px;
        }

        .header {
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            color: white;
            padding: 12px 20px;
            border-radius: 12px;
            margin-bottom: 15px;
            box-shadow: 0 8px 20px rgba(0,0,0,0.1);
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .logo-container {
            display: flex;
            align-items: center;
            gap: 15px;
        }

        .company-logo {
            width: 60px;
            height: 60px;
            border-radius: 10px;
            background: white;
            padding: 6px;
            object-fit: contain;
            box-shadow: 0 3px 10px rgba(0,0,0,0.2);
            border: 2px solid rgba(255,255,255,0.3);
        }

        .header h1 {
            font-size: 24px;
            font-weight: 600;
            margin: 0;
        }

        .header h1 small {
            font-size: 14px;
            opacity: 0.8;
            display: block;
            font-weight: 400;
            margin-top: 2px;
        }

        .header .date-badge {
            background: rgba(255,255,255,0.1);
            padding: 6px 12px;
            border-radius: 25px;
            font-size: 12px;
            display: flex;
            align-items: center;
            gap: 6px;
        }

        .data-source {
            background: #10b981;
            color: white;
            padding: 3px 10px;
            border-radius: 25px;
            font-size: 10px;
            font-weight: 600;
            display: inline-flex;
            align-items: center;
            gap: 3px;
            margin-left: 8px;
        }

        .date-filter-panel {
            background: white;
            border-radius: 12px;
            padding: 15px;
            margin-bottom: 15px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.03);
            border: 1px solid rgba(0,0,0,0.03);
        }

        .date-filter-panel h4 {
            font-size: 14px;
            font-weight: 600;
            color: #1e3c72;
            margin-bottom: 10px;
            display: flex;
            align-items: center;
            gap: 6px;
        }

        .date-filter-grid {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 10px;
            align-items: end;
        }

        .date-input-group {
            display: flex;
            flex-direction: column;
            gap: 3px;
        }

        .date-input-group label {
            font-size: 12px;
            font-weight: 600;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.2px;
        }

        .date-input-group input {
            padding: 6px 10px;
            border: 1px solid #e2e8f0;
            border-radius: 6px;
            font-size: 14px;
            color: #1e293b;
            background: white;
        }

        .apply-btn {
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            color: white;
            border: none;
            padding: 8px 18px;
            border-radius: 25px;
            font-size: 13px;
            font-weight: 500;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            gap: 4px;
            height: 32px;
        }

        .achievement-card {
            background: white;
            border-radius: 12px;
            padding: 15px;
            margin-bottom: 15px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.03);
            border: 1px solid rgba(0,0,0,0.03);
        }

        .achievement-header {
            display: flex;
            align-items: center;
            gap: 8px;
            margin-bottom: 12px;
        }

        .achievement-header h3 {
            font-size: 18px;
            font-weight: 600;
            color: #1e3c72;
            margin: 0;
        }

        .achievement-grid {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 20px;
        }

        .achievement-section {
            background: #f8fafc;
            border-radius: 10px;
            padding: 12px;
        }

        .achievement-title {
            display: flex;
            align-items: center;
            gap: 6px;
            font-size: 13px;
            font-weight: 600;
            color: #1e3c72;
            margin-bottom: 12px;
        }

        .gp-title {
            border-left: 3px solid #059669;
            padding-left: 8px;
        }

        .sales-title {
            border-left: 3px solid #2563eb;
            padding-left: 8px;
        }

        .progress-container {
            margin-bottom: 15px;
        }

        .progress-header {
            display: flex;
            justify-content: space-between;
            margin-bottom: 3px;
            font-size: 11px;
        }

        .progress-bar-bg {
            width: 100%;
            height: 20px;
            background: #e2e8f0;
            border-radius: 10px;
            overflow: hidden;
            position: relative;
        }

        .progress-fill {
            height: 100%;
            background: linear-gradient(90deg, #059669, #10b981);
            border-radius: 10px;
            transition: width 0.5s ease;
            position: relative;
            display: flex;
            align-items: center;
            justify-content: flex-end;
            padding-right: 6px;
            color: white;
            font-size: 12px;
            font-weight: 600;
        }

        .progress-fill.sales {
            background: linear-gradient(90deg, #2563eb, #3b82f6);
        }

        .progress-fill.warning {
            background: linear-gradient(90deg, #d97706, #f59e0b);
        }

        .progress-fill.danger {
            background: linear-gradient(90deg, #dc2626, #ef4444);
        }

        .progress-fill.positive {
            background: linear-gradient(90deg, #059669, #10b981);
        }

        .target-details {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 10px;
            margin-top: 12px;
            padding-top: 12px;
            border-top: 1px dashed #e2e8f0;
        }

        .target-item {
            text-align: center;
        }

        .target-label {
            font-size: 11px;
            color: #64748b;
            margin-bottom: 2px;
        }

        .target-number {
            font-size: 16px;
            font-weight: 700;
            color: #1e3c72;
        }

        .target-sub {
            font-size: 9px;
            color: #94a3b8;
        }

        .dropdown-panel {
            background: white;
            border-radius: 12px;
            padding: 15px;
            margin-bottom: 15px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.03);
            border: 1px solid rgba(0,0,0,0.03);
        }

        .dropdown-panel h3 {
            font-size: 18px;
            font-weight: 600;
            color: #1e3c72;
            margin-bottom: 12px;
            display: flex;
            align-items: center;
            gap: 6px;
        }

        .dropdown-grid {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 10px;
            margin-bottom: 12px;
        }

        .dropdown-item {
            display: flex;
            flex-direction: column;
            gap: 3px;
        }

        .dropdown-item label {
            font-size: 12px;
            font-weight: 600;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.2px;
        }

        .dropdown-item select {
            padding: 6px 10px;
            border: 1px solid #e2e8f0;
            border-radius: 6px;
            font-size: 14px;
            font-weight: 500;
            color: #1e293b;
            background: white;
            cursor: pointer;
            appearance: none;
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24' fill='none' stroke='%2364748b' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='6 9 12 15 18 9'%3E%3C/polyline%3E%3C/svg%3E");
            background-repeat: no-repeat;
            background-position: right 8px center;
            background-size: 12px;
        }

        .dropdown-item select:hover {
            border-color: #2a5298;
        }

        .month-badge {
            background: #f8fafc;
            padding: 6px 12px;
            border-radius: 25px;
            font-size: 12px;
            font-weight: 500;
            color: #1e3c72;
            display: inline-flex;
            align-items: center;
            gap: 6px;
            border: 1px solid #e2e8f0;
        }

        .refresh-btn {
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            color: white;
            border: none;
            padding: 8px 18px;
            border-radius: 25px;
            font-size: 13px;
            font-weight: 500;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            gap: 4px;
        }

        .table-container {
            background: white;
            border-radius: 12px;
            padding: 15px;
            margin-bottom: 15px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.03);
            border: 1px solid rgba(0,0,0,0.03);
        }

        .table-container h3 {
            font-size: 18px;
            font-weight: 600;
            color: #1e3c72;
            margin-bottom: 12px;
            display: flex;
            align-items: center;
            gap: 6px;
        }

        .table-responsive {
            overflow-x: auto;
            border-radius: 8px;
            border: 1px solid #e2e8f0;
            max-height: 450px;
            overflow-y: auto;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            min-width: 1500px;
        }

        th {
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            padding: 8px 6px;
            font-weight: 700;
            color: white;
            text-align: left;
            white-space: nowrap;
            font-size: 12px;
            text-transform: uppercase;
            letter-spacing: 0.2px;
            position: sticky;
            top: 0;
            z-index: 10;
        }

        td {
            padding: 8px 6px;
            border-bottom: 1px solid #e2e8f0;
            color: #1e293b;
        }

        tr:hover {
            background: #eef2ff;
        }

        .text-right {
            text-align: right;
        }

        .text-center {
            text-align: center;
        }

        .font-bold {
            font-weight: 600;
        }

        .variance-positive {
            color: #059669;
            font-weight: 600;
        }

        .variance-negative {
            color: #dc2626;
            font-weight: 600;
        }

        .value-with-percent {
            display: flex;
            flex-direction: column;
            line-height: 1.3;
        }
        .value-with-percent .main-value {
            font-weight: 600;
        }
        .value-with-percent .percent-value {
            font-size: 9px;
            color: #64748b;
        }

        .tabs-nav {
            display: flex;
            gap: 2px;
            background: white;
            padding: 6px;
            border-radius: 35px;
            margin-bottom: 15px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.03);
            overflow-x: auto;
            flex-wrap: nowrap;
            border: 1px solid rgba(0,0,0,0.03);
        }

        .tab-btn {
            padding: 8px 14px;
            border: none;
            background: transparent;
            border-radius: 25px;
            font-size: 13px;
            font-weight: 500;
            color: #64748b;
            cursor: pointer;
            white-space: nowrap;
            display: flex;
            align-items: center;
            gap: 3px;
            transition: all 0.2s;
        }

        .tab-btn:hover {
            color: #1e3c72;
            background: #f8fafc;
        }

        .tab-btn.active {
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            color: white;
        }

        .tab-content {
            background: white;
            border-radius: 12px;
            padding: 15px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.03);
            margin-bottom: 15px;
            border: 1px solid rgba(0,0,0,0.03);
        }

        .tab-pane {
            display: none;
        }

        .tab-pane.active {
            display: block;
        }

        .analysis-grid {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 12px;
        }

        .analysis-grid-3col {
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 12px;
        }

        .chart-card {
            background: white;
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            padding: 12px;
            width: 100%;
        }

        .chart-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
        }

        .chart-header h4 {
            font-size: 15px;
            font-weight: 600;
            color: #1e3c72;
            margin: 0;
            display: flex;
            align-items: center;
            gap: 4px;
        }

        .chart-container {
            height: 220px;
            position: relative;
        }

        .inner-table {
            width: 100%;
            font-size: 13px;
            min-width: 0;
        }

        .inner-table th {
            background: #f8fafc;
            color: #1e3c72;
            padding: 8px 6px;
            font-size: 12px;
            position: static;
        }

        .inner-table td {
            padding: 8px 6px;
        }

        .footer {
            text-align: center;
            padding: 12px;
            color: #64748b;
            font-size: 10px;
        }

        .alert {
            padding: 8px 12px;
            border-radius: 8px;
            font-size: 11px;
            margin-bottom: 12px;
        }

        .status-badge {
            padding: 2px 5px;
            border-radius: 10px;
            font-size: 9px;
            font-weight: 600;
            display: inline-block;
            white-space: nowrap;
        }

        .status-achieved {
            background: #d1fae5;
            color: #065f46;
        }

        .status-on-track {
            background: #fef3c7;
            color: #92400e;
        }

        .status-at-risk {
            background: #fee2e2;
            color: #991b1b;
        }

        .no-data-message {
            text-align: center;
            padding: 40px;
            background: white;
            border-radius: 12px;
            margin: 15px 0;
        }

        .no-data-message i {
            font-size: 42px;
            color: #94a3b8;
            margin-bottom: 12px;
        }

        .no-data-message h3 {
            color: #1e293b;
            margin-bottom: 8px;
            font-size: 18px;
        }

        .no-data-message p {
            color: #64748b;
            font-size: 13px;
        }

        .comparison-table-container {
            background: white;
            border-radius: 12px;
            padding: 15px;
            margin-bottom: 20px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.03);
            border: 1px solid rgba(0,0,0,0.03);
        }

        .comparison-table-container h3 {
            font-size: 16px;
            font-weight: 600;
            color: #1e3c72;
            margin-bottom: 12px;
            display: flex;
            align-items: center;
            gap: 8px;
            border-left: 3px solid #2a5298;
            padding-left: 12px;
        }

        .comparison-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            min-width: 800px;
        }

        #salesPerClientTable {
            width: 100%;
            min-width: 0;
            table-layout: fixed;
        }

        #salesPerClientTable th,
        #salesPerClientTable td {
            text-align: center !important;
            vertical-align: middle !important;
            padding: 6px 4px;
            font-size: 12px;
        }

        #salesPerClientTable th:first-child,
        #salesPerClientTable td:first-child {
            width: 140px;
            white-space: normal;
            overflow-wrap: anywhere;claude
            line-height: 1.25;
        }

        #salesPerClientTable th:not(:first-child),
        #salesPerClientTable td:not(:first-child) {
            width: calc((100% - 140px) / 12);
        }

        .comparison-table th {
            background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
            padding: 10px 8px;
            font-weight: 700;
            color: white;
            text-align: center;
            font-size: 12px;
            text-transform: uppercase;
        }

        .comparison-table td {
            padding: 8px 6px;
            border-bottom: 1px solid #e2e8f0;
            text-align: right;
        }

        .comparison-table td:first-child {
            text-align: left;
            font-weight: 600;
            background: #f8fafc;
        }

        .comparison-table tr:hover td {
            background: #eff6ff;
        }

        .comparison-table tr:hover td:first-child {
            background: #e9ecef;
        }

        .variance-positive {
            color: #059669;
            font-weight: 600;
        }

        .variance-negative {
            color: #dc2626;
            font-weight: 600;
        }

        .grand-total-row {
            background: #f1f5f9;
            font-weight: 700;
            border-top: 2px solid #2a5298;
        }
        
        .grand-total-row td {
            font-weight: 700;
            background: #e9ecef;
        }

        .pnl-table {
            min-width: 1100px;
        }

        .pnl-table th:nth-child(2),
        .comparison-table th.actual-dataset-label {
            background: linear-gradient(135deg, #1d4ed8 0%, #3b82f6 100%);
        }

        .pnl-table th:nth-child(3),
        .comparison-table th.forecast-dataset-label,
        .comparison-table th.vs-forecast-label {
            background: linear-gradient(135deg, #dbeafe 0%, #eff6ff 100%);
            color: #1e40af;
        }

        .pnl-table th:nth-child(4),
        .comparison-table th.vs-target-label {
            background: linear-gradient(135deg, #fff7ed 0%, #ffedd5 100%);
            color: #b45309;
        }

        .pnl-table th:nth-child(5),
        .comparison-table th.prev-dataset-label,
        .comparison-table th.vs-prev-label {
            background: linear-gradient(135deg, #f9fafb 0%, #f3f4f6 100%);
            color: #4b5563;
        }

        .pnl-table td:nth-child(6),
        .comparison-table td:nth-child(6) {
            background: #eff6ff;
        }

        .pnl-table td:nth-child(7),
        .comparison-table td:nth-child(7) {
            background: #fff7ed;
        }

        .pnl-table td:nth-child(8),
        .comparison-table td:nth-child(8) {
            background: #f9fafb;
        }

        .pnl-table td:first-child {
            font-weight: 700;
            background: #f8fafc;
            color: #0f172a;
        }

        .pnl-table .subtotal-row td {
            background: #e2e8f0;
            font-weight: 700;
            border-top: 1px solid #94a3b8;
            border-bottom: 1px solid #94a3b8;
        }

        .pnl-table .percentage-row td:first-child {
            color: #334155;
            font-weight: 600;
        }

        .variance-positive {
            color: #059669;
            font-weight: 700;
        }

        .variance-negative {
            color: #dc2626;
            font-weight: 700;
        }

        .insight-card {
            background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
            border: 1px solid #dbeafe;
        }

        .insight-list {
            margin: 0;
            padding-left: 18px;
            font-size: 12px;
            line-height: 1.7;
            color: #334155;
        }

        .insight-list li + li {
            margin-top: 6px;
        }

        .analysis-trigger {
            margin-left: 8px;
            border: none;
            background: #e0e7ff;
            color: #1e3c72;
            width: 24px;
            height: 24px;
            border-radius: 999px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            font-size: 12px;
            cursor: pointer;
        }

        .analysis-trigger:hover {
            background: #c7d2fe;
            filter: brightness(0.95);
        }

        .combined-table-actions {
            display: flex;
            justify-content: flex-end;
            margin-bottom: 10px;
        }

        .toggle-btn {
            border: 1px solid #cbd5e1;
            background: #fff;
            color: #1e3c72;
            border-radius: 999px;
            padding: 6px 12px;
            font-size: 13px;
            font-weight: 600;
            cursor: pointer;
        }

        .toggle-btn:hover {
            filter: brightness(0.95);
        }

        .combined-table .group-row td {
            background: #dbeafe;
            color: #1e3a8a;
            font-weight: 800;
            border-top: 2px solid #93c5fd;
        }

        .combined-table .data-row td:first-child {
            font-weight: 700;
        }

        .combined-table .variance-row td:first-child {
            color: #475569;
        }

        .combined-table .full-year-col {
            background: #eef2ff !important;
            font-weight: 800;
        }

        .metric-sales {
            color: #2563eb;
        }

        .metric-gp {
            color: #0891b2;
        }

        .metric-qty {
            color: #7c3aed;
        }

        .metric-gm {
            color: #059669;
        }

        .metric-not-booked {
            color: #d97706;
        }

        .sales-block {
            background: #eff6ff;
        }

        .gm-block {
            background: #ecfdf5;
        }

        .vs-block {
            background: #f8fafc;
        }

        .sales-text {
            color: #2563eb;
            font-weight: 700;
        }

        .gm-text {
            color: #059669;
            font-weight: 700;
        }

        .vs-text {
            color: #475569;
            font-weight: 700;
        }

        .large-number {
            font-size: 14px;
            font-weight: 700;
        }

        .not-booked-note {
            display: block;
            margin-top: 2px;
            color: #dc2626;
            font-size: 11px;
            font-weight: 700;
        }

        .tracker-toolbar {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 12px;
            margin-bottom: 12px;
            flex-wrap: wrap;
        }

        .tracker-note {
            font-size: 12px;
            color: #475569;
        }

        .tracker-table-shell {
            min-height: 820px;
            max-height: 820px;
            overflow-y: auto;
            overflow-x: auto;
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            background: #fff;
        }

        .tracker-table,
        .tracker-table th,
        .tracker-table td,
        .tracker-table select,
        .tracker-table input,
        .tracker-table button {
            font-size: 14px;
            line-height: 1.4;
        }

        .tracker-table {
            table-layout: auto;
            min-width: 100%;
        }

        .tracker-table th,
        .tracker-table td {
            padding: 8px 6px;
            vertical-align: middle;
            white-space: normal;
            word-break: break-word;
            text-align: center;
        }

        .tracker-table th:nth-child(1),
        .tracker-table td:nth-child(1) {
            width: 90px;
            min-width: 90px;
            max-width: 100px;
        }

        .tracker-table th:nth-child(2),
        .tracker-table td:nth-child(2) {
            width: 90px;
            min-width: 90px;
            max-width: 100px;
        }

        .tracker-table th:nth-child(3),
        .tracker-table td:nth-child(3) {
            width: 120px;
            min-width: 120px;
            max-width: 140px;
        }

        .tracker-table tbody tr {
            min-height: 40px;
        }

        .tracker-table tbody tr:nth-child(even) {
            background: #f8fafc;
        }

        .tracker-table select,
        .tracker-table input {
            width: 100%;
            min-width: 110px;
            padding: 6px 8px;
            border: 1px solid #cbd5e1;
            border-radius: 6px;
            background: #fff;
        }

        .tracker-filter-dropdown {
            position: relative;
            min-width: 120px;
        }

        .tracker-filter-btn {
            width: 100%;
            min-width: 120px;
            padding: 6px 8px;
            border: 1px solid #cbd5e1;
            border-radius: 6px;
            background: #fff;
            color: #1e293b;
            text-align: left;
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 8px;
        }

        .tracker-filter-btn.active {
            border-color: #2563eb;
            box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.12);
        }

        .tracker-filter-menu {
            display: none;
            position: absolute;
            top: calc(100% + 4px);
            left: 0;
            min-width: 220px;
            max-height: 260px;
            overflow-y: auto;
            background: #fff;
            border: 1px solid #cbd5e1;
            border-radius: 8px;
            box-shadow: 0 10px 24px rgba(15, 23, 42, 0.12);
            padding: 8px;
            z-index: 30;
        }

        .tracker-filter-menu.open {
            display: block;
        }

        .tracker-filter-option {
            display: flex;
            align-items: center;
            gap: 8px;
            padding: 4px 2px;
            color: #334155;
        }

        .tracker-filter-option input {
            width: auto;
            min-width: 0;
            margin: 0;
        }

        .tracker-filter-empty {
            color: #64748b;
            padding: 4px 2px;
        }

        .tracker-filter-row th {
            background: #eef2ff;
            position: sticky;
            top: 38px;
            z-index: 9;
            padding: 6px 4px;
        }

        .tracker-status-booked {
            color: #1d4ed8;
            font-weight: 700;
            font-size: 15px;
        }

        .tracker-status-not-booked {
            color: #dc2626;
            font-weight: 700;
            font-size: 15px;
        }

        .tracker-table .status-badge {
            font-size: 15px;
            padding: 4px 8px;
        }

        .tracker-summary-grid {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 12px;
            margin-bottom: 12px;
        }

        .tracker-header-bar {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 12px;
            flex-wrap: wrap;
            margin-bottom: 10px;
        }

        .tracker-header-filter {
            min-width: 220px;
        }

        .combined-summary-grid {
            display: grid;
            grid-template-columns: repeat(4, minmax(0, 1fr));
            gap: 12px;
            margin-bottom: 14px;
        }

        .combined-summary-card {
            background: linear-gradient(135deg, #f8fafc 0%, #eef2ff 100%);
            border: 1px solid #dbeafe;
            border-radius: 12px;
            padding: 12px 14px;
        }

        .combined-summary-card .summary-label {
            font-size: 11px;
            font-weight: 700;
            text-transform: uppercase;
            color: #64748b;
            margin-bottom: 6px;
            letter-spacing: 0.3px;
        }

        .combined-summary-card .summary-value {
            font-size: 16px;
            font-weight: 800;
            color: #1e3a8a;
            margin-bottom: 4px;
        }

        .combined-summary-card .summary-sub {
            font-size: 12px;
            color: #475569;
        }

        .performance-badge {
            display: inline-block;
            padding: 2px 8px;
            border-radius: 999px;
            font-size: 9px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.3px;
        }

        .badge-top {
            background: #dcfce7;
            color: #166534;
        }

        .badge-worst {
            background: #fee2e2;
            color: #991b1b;
        }

        .pnl-table td:first-child {
            font-weight: 700;
            background: #f8fafc;
        }

        .pnl-table tr:nth-child(3),
        .pnl-table tr:nth-child(4),
        .pnl-table tr:nth-child(5) {
            background: #fbfdff;
        }

        .client-cards-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
            gap: 12px;
        }

        .client-card {
            background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
            border: 1px solid #dbeafe;
            border-radius: 12px;
            padding: 14px;
        }

        .client-card h4 {
            font-size: 13px;
            color: #1e3c72;
            margin-bottom: 8px;
        }

        .client-card-metrics {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 8px;
        }

        .client-card-metric {
            background: #fff;
            border-radius: 8px;
            padding: 8px;
            border: 1px solid #e2e8f0;
        }

        .client-card-label {
            font-size: 9px;
            color: #64748b;
            text-transform: uppercase;
        }

        .client-card-value {
            font-size: 13px;
            font-weight: 700;
            color: #1e293b;
        }

        @media (max-width: 1200px) {
            .dropdown-grid {
                grid-template-columns: repeat(2, 1fr);
            }
            .analysis-grid, .analysis-grid-3col {
                grid-template-columns: 1fr;
            }
            .achievement-grid {
                grid-template-columns: 1fr;
                gap: 12px;
            }
            .date-filter-grid {
                grid-template-columns: 1fr;
            }
        }

        @media (max-width: 768px) {
            .dropdown-grid {
                grid-template-columns: 1fr;
            }
            .header {
                flex-direction: column;
                gap: 8px;
                text-align: center;
            }
            table,
            .comparison-table,
            .inner-table {
                font-size: 12px;
            }
            th,
            .comparison-table th,
            .inner-table th {
                font-size: 11px;
            }
        }
    </style>
</head>
<body>
    <div class="dashboard-wrapper">
        <div class="header">
            <div class="logo-container">
                <img src="data:image/jpeg;base64,{{ logo_base64 }}" alt="Al-Behar Group Logo" class="company-logo">
                <div>
                    <h1>
                        Al-Behar Company
                        <small>Oils</small>
                        <small>S&OP Analytics Dashboard</small>
                    </h1>
                    <span id="dataSourceBadge" class="data-source">
                        <i class="bi bi-database"></i> Loading...
                    </span>
                </div>
            </div>
            <div class="date-badge">
                <i class="bi bi-calendar3"></i>
                <span id="dataDateRange">Loading...</span>
            </div>
        </div>

        <div id="noSummaryData" class="no-data-message" style="display: none;">
            <i class="bi bi-exclamation-triangle-fill"></i>
            <h3>No Summary Data Found</h3>
            <p>The dashboard requires a 'Summary' sheet with achievement data to display achievement bars.</p>
            <p>Please ensure your Excel file contains a 'Summary' sheet with the required columns.</p>
        </div>

                <div class="achievement-card" id="achievementCard">
            <div class="achievement-header">
                <i class="bi bi-bar-chart-fill" style="color: #1e3c72; font-size: 18px;"></i>
                <h3>Target Achievement</h3>
            </div>
            
            <div class="achievement-grid">
                <div class="achievement-section">
                    <div class="achievement-title sales-title">
                        <i class="bi bi-cart" style="color: #2563eb; font-size: 14px;"></i>
                        Sales Achievement YTD
                    </div>
                    <div class="progress-container">
                        <div class="progress-header">
                            <span class="metric-label">Sales Achievement</span>
                            <span class="metric-value" id="salesYtdPercent">0%</span>
                        </div>
                        <div class="progress-bar-bg">
                            <div class="progress-fill sales" id="salesYtdBar" style="width: 0%">0%</div>
                        </div>
                    </div>
                    <div class="target-details">
                        <div class="target-item">
                            <div class="target-label">Target</div>
                            <div class="target-number" id="salesYtdTarget">0 JOD</div>
                        </div>
                        <div class="target-item">
                            <div class="target-label">Actual</div>
                            <div class="target-number" id="salesYtdActual">0 JOD</div>
                        </div>
                        <div class="target-item">
                            <div class="target-label">Remaining</div>
                            <div class="target-number" id="salesYtdRemaining">0 JOD</div>
                        </div>
                    </div>
                </div>
                
                <div class="achievement-section">
                    <div class="achievement-title gp-title">
                        <i class="bi bi-piggy-bank" style="color: #059669; font-size: 14px;"></i>
                        Gross Profit Achievement YTD
                    </div>
                    <div class="progress-container">
                        <div class="progress-header">
                            <span class="metric-label">Trade GP Achievement</span>
                            <span class="metric-value" id="gpTradeYtdPercent">0%</span>
                        </div>
                        <div class="progress-bar-bg">
                            <div class="progress-fill" id="gpTradeYtdBar" style="width: 0%">0%</div>
                        </div>
                    </div>
                    <div class="progress-container">
                        <div class="progress-header">
                            <span class="metric-label">Broker GP Achievement</span>
                            <span class="metric-value" id="gpBrokerYtdPercent">0%</span>
                        </div>
                        <div class="progress-bar-bg">
                            <div class="progress-fill sales" id="gpBrokerYtdBar" style="width: 0%">0%</div>
                        </div>
                    </div>
                    <div class="target-details">
                        <div class="target-item">
                            <div class="target-label">Trade Target</div>
                            <div class="target-number" id="gpTradeYtdTarget">0 JOD</div>
                            <div class="target-sub" id="gpTradeYtdActual">Actual: 0 JOD</div>
                        </div>
                        <div class="target-item">
                            <div class="target-label">Broker Target</div>
                            <div class="target-number" id="gpBrokerYtdTarget">0 JOD</div>
                            <div class="target-sub" id="gpBrokerYtdActual">Actual: 0 JOD</div>
                        </div>
                        <div class="target-item">
                            <div class="target-label">Total Remaining</div>
                            <div class="target-number" id="gpTotalYtdRemaining">0 JOD</div>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <div class="date-filter-panel">
            <h4>
                <i class="bi bi-calendar-range"></i>
                Date Range Filter (Posting Date)
                <span id="activeFilterBadge" class="filter-badge" style="display: none; background: #1e3c72; color: white; padding: 4px 10px; border-radius: 20px; font-size: 11px; margin-left: 10px;">
                    <span id="filterDates"></span>
                    <i class="bi bi-x-circle" onclick="resetDateFilter()" style="cursor: pointer; margin-left: 5px;"></i>
                </span>
            </h4>
            <div class="date-filter-grid">
                <div class="date-input-group">
                    <label>From Date</label>
                    <input type="text" id="fromDate" class="datepicker" placeholder="Select start date">
                </div>
                <div class="date-input-group">
                    <label>To Date</label>
                    <input type="text" id="toDate" class="datepicker" placeholder="Select end date">
                </div>
                <div>
                    <button class="apply-btn" onclick="applyDateFilter()">
                        <i class="bi bi-funnel"></i> Apply Filter
                    </button>
                    <button class="apply-btn" onclick="resetDateFilter()" style="background: #64748b; margin-left: 5px;">
                        <i class="bi bi-arrow-counterclockwise"></i> Reset
                    </button>
                </div>
            </div>
        </div>

        <div class="tabs-nav">
            <button class="tab-btn active" onclick="showTab('sop')"><i class="bi bi-arrow-repeat"></i> S&OP Cycle</button>
            <button class="tab-btn" onclick="showTab('sales')"><i class="bi bi-graph-up"></i> Sales</button>
            <button class="tab-btn" onclick="showTab('profitability')"><i class="bi bi-cash-stack"></i> Profitability</button>
            <button class="tab-btn" onclick="showTab('customers')"><i class="bi bi-people"></i> Customers</button>
            <button class="tab-btn" onclick="showTab('products')"><i class="bi bi-box-seam"></i> Products</button>
            <button class="tab-btn" onclick="showTab('suppliers')"><i class="bi bi-truck"></i> Suppliers</button>
            <button class="tab-btn" onclick="showTab('tracker')"><i class="bi bi-clipboard2-check"></i> Tracker</button>
        </div>

        <!-- S&OP CYCLE TAB -->
        <div id="sopTab" class="tab-content tab-pane active">
            <div class="dropdown-panel">
                <h3>
                    <i class="bi bi-sliders2"></i>
                    S&OP Cycle Comparison
                </h3>
                
                <div class="dropdown-grid">
                    <div class="dropdown-item">
                        <label><i class="bi bi-check-circle"></i> Actual Dataset (Invoiced)</label>
                        <select id="actualDataset" class="dataset-select">
                            <option value="">Select Actual Dataset</option>
                        </select>
                    </div>
                    
                    <div class="dropdown-item">
                        <label><i class="bi bi-graph-up-arrow"></i> Forecast Dataset (Booked/Not Booked)</label>
                        <select id="forecastDataset" class="dataset-select">
                            <option value="">Select Forecast Dataset</option>
                        </select>
                    </div>
                    
                    <div class="dropdown-item">
                        <label><i class="bi bi-calendar2-week"></i> Previous Year Dataset</label>
                        <select id="prevYearDataset" class="dataset-select">
                            <option value="">Select Previous Year Dataset</option>
                        </select>
                    </div>
                </div>
                
                <div style="display: flex; justify-content: space-between; align-items: center; margin-top: 12px;">
                    <div class="month-badge" id="analysisMonth">
                        <i class="bi bi-calendar-month"></i>
                        Analysis Month: <span id="monthDisplay">Not Selected</span>
                    </div>
                    
                    <button class="refresh-btn" onclick="refreshComparison()">
                        <i class="bi bi-arrow-repeat"></i>
                        Update
                    </button>
                </div>
            </div>

            <div class="table-container">
                <h3>
                    <i class="bi bi-calendar-month"></i>
                    Month Comparison (<span id="monthName">Selected Month</span>)
                    <button class="analysis-trigger" onclick="openTableAnalysis('month')" title="Show analysis"><i class="bi bi-info-lg"></i></button>
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table pnl-table">
                        <thead>
                            <tr>
                                <th>Metric</th>
                                <th class="actual-dataset-label">Actual</th>
                                <th class="forecast-dataset-label">Forecast</th>
                                <th>Target</th>
                                <th class="prev-dataset-label">Prev Yr</th>
                                <th class="vs-forecast-label">Vs Forecast</th>
                                <th class="vs-target-label">Vs Target</th>
                                <th class="vs-prev-label">Vs Prev</th>
                            </tr>
                        </thead>
                        <tbody id="monthComparisonTableBody">
                            <tr>
                                <td colspan="12" class="text-center" style="padding: 25px;">
                                    <i class="bi bi-arrow-up-circle" style="font-size: 18px;"></i>
                                    <br>
                                    Select datasets above to view comparison
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <div class="table-container">
                <h3>
                    <i class="bi bi-calendar-range"></i>
                    YTD Comparison (Jan - <span id="ytdMonthName">Selected Month</span>)
                    <button class="analysis-trigger" onclick="openTableAnalysis('ytd')" title="Show analysis"><i class="bi bi-info-lg"></i></button>
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table pnl-table">
                        <thead>
                            <tr>
                                <th>Metric</th>
                                <th class="actual-dataset-label">Actual</th>
                                <th class="forecast-dataset-label">Forecast</th>
                                <th>Target</th>
                                <th class="prev-dataset-label">Prev Yr</th>
                                <th class="vs-forecast-label">Vs Forecast</th>
                                <th class="vs-target-label">Vs Target</th>
                                <th class="vs-prev-label">Vs Prev</th>
                            </tr>
                        </thead>
                        <tbody id="ytdComparisonTableBody">
                            <tr>
                                <td colspan="12" class="text-center" style="padding: 25px;">
                                    <i class="bi bi-arrow-up-circle" style="font-size: 18px;"></i>
                                    <br>
                                    Select datasets above to view YTD comparison
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <div class="table-container">
                <h3>
                    <i class="bi bi-calendar2-week"></i>
                    Full Year Comparison (All Months)
                    <button class="analysis-trigger" onclick="openTableAnalysis('fullYear')" title="Show analysis"><i class="bi bi-info-lg"></i></button>
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table pnl-table">
                        <thead>
                            <tr>
                                <th>Metric</th>
                                <th class="actual-dataset-label">Actual</th>
                                <th class="forecast-dataset-label">Forecast</th>
                                <th>Target</th>
                                <th class="prev-dataset-label">Prev Yr</th>
                                <th class="vs-forecast-label">Vs Forecast</th>
                                <th class="vs-target-label">Vs Target</th>
                                <th class="vs-prev-label">Vs Prev</th>
                            </tr>
                        </thead>
                        <tbody id="fullYearComparisonTableBody">
                            <tr>
                                <td colspan="12" class="text-center" style="padding: 25px;">
                                    <i class="bi bi-arrow-up-circle" style="font-size: 18px;"></i>
                                    <br>
                                    Select datasets above to view full year comparison
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <div class="analysis-grid">
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-pie-chart"></i> Sales by Status</h4>
                    </div>
                    <div class="chart-container">
                        <canvas id="salesStatusChart"></canvas>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-pie-chart"></i> Gross Profit by Status</h4>
                    </div>
                    <div class="chart-container">
                        <canvas id="gpStatusChart"></canvas>
                    </div>
                </div>
            </div>

            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-cart"></i>
                    Sales Monthly Comparison
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table combined-table" id="salesComparisonTable">
                        <thead>
                            <tr>
                                <th>Version</th>
                                <th>Jan</th>
                                <th>Feb</th>
                                <th>Mar</th>
                                <th>Apr</th>
                                <th>May</th>
                                <th>Jun</th>
                                <th>Jul</th>
                                <th>Aug</th>
                                <th>Sep</th>
                                <th>Oct</th>
                                <th>Nov</th>
                                <th>Dec</th>
                                <th>Full Year</th>
                            </tr>
                        </thead>
                        <tbody id="salesComparisonBody">
                            <tr>
                                <td colspan="14" class="text-center">Select datasets above to view comparison</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-piggy-bank"></i>
                    Gross Profit Monthly Comparison
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table combined-table" id="gpComparisonTable">
                        <thead>
                            <tr>
                                <th>Version</th>
                                <th>Jan</th>
                                <th>Feb</th>
                                <th>Mar</th>
                                <th>Apr</th>
                                <th>May</th>
                                <th>Jun</th>
                                <th>Jul</th>
                                <th>Aug</th>
                                <th>Sep</th>
                                <th>Oct</th>
                                <th>Nov</th>
                                <th>Dec</th>
                                <th>Full Year</th>
                            </tr>
                        </thead>
                        <tbody id="gpComparisonBody">
                            <tr>
                                <td colspan="14" class="text-center">Select datasets above to view comparison</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-box-seam"></i>
                    Quantity Monthly Comparison
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table combined-table" id="qtyComparisonTable">
                        <thead>
                            <tr>
                                <th>Version</th>
                                <th>Jan</th>
                                <th>Feb</th>
                                <th>Mar</th>
                                <th>Apr</th>
                                <th>May</th>
                                <th>Jun</th>
                                <th>Jul</th>
                                <th>Aug</th>
                                <th>Sep</th>
                                <th>Oct</th>
                                <th>Nov</th>
                                <th>Dec</th>
                                <th>Full Year</th>
                            </tr>
                        </thead>
                        <tbody id="qtyComparisonBody">
                            <tr>
                                <td colspan="14" class="text-center">Select datasets above to view comparison</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- NEW TABLE 1: Sales per Client -->
            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-people"></i>
                    Sales & Gross Profit per Client
                    <button class="analysis-trigger" onclick="openTableAnalysis('client')" title="Show analysis"><i class="bi bi-info-lg"></i></button>
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table" id="salesPerClientTable">
                        <thead>
                            <tr>
                                <th>Client</th>
                                <th class="actual-sales-label sales-block">Actual Sales</th>
                                <th class="sales-block">Contribution %</th>
                                <th class="sales-block">Not Booked Sales %</th>
                                <th class="actual-gm-label gm-block">Actual GM</th>
                                <th class="gm-block">GM%</th>
                                <th class="gm-block">Not Booked GM %</th>
                                <th class="sales-vs-prev-label vs-block">Sales vs Prev</th>
                                <th class="sales-vs-forecast-label vs-block">Sales vs Forecast</th>
                                <th class="sales-vs-target-label vs-block">Sales vs Target</th>
                                <th class="gm-vs-prev-label vs-block">GM vs Prev</th>
                                <th class="gm-vs-forecast-label vs-block">GM vs Forecast</th>
                                <th class="gm-vs-target-label vs-block">GM vs Target</th>
                            </tr>
                        </thead>
                        <tbody id="salesPerClientBody">
                            <tr>
                                <td colspan="13" class="text-center">Select datasets above to view client data</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- NEW TABLE 2: GM per Client -->
            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-piggy-bank"></i>
                    Gross Margin per Client
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table" id="gmPerClientTable" style="display:none;">
                        <thead>
                            <tr>
                                <th>Client</th>
                                <th>Actual (Full Year)</th>
                                <th>GM%</th>
                                <th>Contribution %</th>
                                <th>Not Booked % from Act</th>
                                <th>Act vs Prev</th>
                                <th>Act vs Forecast</th>
                                <th>Act vs Target</th>
                            </tr>
                        </thead>
                        <tbody id="gmPerClientBody">
                            <tr>
                                <td colspan="8" class="text-center">Select datasets above to view client GM data</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- NEW TABLE 3: Sales per Category -->
            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-tags"></i>
                    Sales & Gross Profit per Category
                    <button class="analysis-trigger" onclick="openTableAnalysis('category')" title="Show analysis"><i class="bi bi-info-lg"></i></button>
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table" id="salesPerCategoryTable">
                        <thead>
                            <tr>
                                <th>Category</th>
                                <th class="actual-sales-label sales-block">Actual Sales</th>
                                <th class="sales-block">Contribution %</th>
                                <th class="sales-block">Not Booked Sales %</th>
                                <th class="actual-gm-label gm-block">Actual GM</th>
                                <th class="gm-block">GM%</th>
                                <th class="gm-block">Not Booked GM %</th>
                                <th class="sales-vs-prev-label vs-block">Sales vs Prev</th>
                                <th class="sales-vs-forecast-label vs-block">Sales vs Forecast</th>
                                <th class="sales-vs-target-label vs-block">Sales vs Target</th>
                                <th class="gm-vs-prev-label vs-block">GM vs Prev</th>
                                <th class="gm-vs-forecast-label vs-block">GM vs Forecast</th>
                                <th class="gm-vs-target-label vs-block">GM vs Target</th>
                            </tr>
                        </thead>
                        <tbody id="salesPerCategoryBody">
                            <tr>
                                <td colspan="13" class="text-center">Select datasets above to view category data</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- NEW TABLE 4: GM per Category -->
            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-percent"></i>
                    Gross Margin per Category
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table" id="gmPerCategoryTable" style="display:none;">
                        <thead>
                            <tr>
                                <th>Category</th>
                                <th>Actual (Full Year)</th>
                                <th>GM%</th>
                                <th>Contribution %</th>
                                <th>Not Booked % from Act</th>
                                <th>Act vs Prev</th>
                                <th>Act vs Forecast</th>
                                <th>Act vs Target</th>
                            </tr>
                        </thead>
                        <tbody id="gmPerCategoryBody">
                            <tr>
                                <td colspan="8" class="text-center">Select datasets above to view category GM data</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- NEW TABLE 5: Suppliers Analysis -->
            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-truck"></i>
                    Suppliers Analysis
                    <button class="analysis-trigger" onclick="openTableAnalysis('supplier')" title="Show analysis"><i class="bi bi-info-lg"></i></button>
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table" id="suppliersTable">
                        <thead>
                            <tr>
                                <th>Supplier</th>
                                <th>Actual Sales (Full Year)</th>
                                <th>Sales Contribution %</th>
                                <th>GM</th>
                                <th>GM Contribution %</th>
                                <th>Broker GP</th>
                                <th>Broker Contribution %</th>
                            </tr>
                        </thead>
                        <tbody id="suppliersBody">
                            <tr>
                                <td colspan="7" class="text-center">Select datasets above to view supplier data</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- NEW TABLE 6: Broker per Client -->
            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-calculator"></i>
                    Broker per Client GM
                    <button class="analysis-trigger" onclick="openTableAnalysis('brokerClient')" title="Show analysis"><i class="bi bi-info-lg"></i></button>
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table" id="brokerPerClientTable">
                        <thead>
                            <tr>
                                <th>Client</th>
                                <th class="actual-gm-label">Actual GM</th>
                                <th>Contribution %</th>
                                <th>Not Booked % from Act</th>
                                <th class="broker-vs-prev-label">Act vs Prev</th>
                                <th class="broker-vs-forecast-label">Act vs Forecast</th>
                                <th class="broker-vs-target-label">Act vs Target</th>
                            </tr>
                        </thead>
                        <tbody id="brokerPerClientBody">
                            <tr>
                                <td colspan="7" class="text-center">Select datasets above to view broker per client data</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- NEW TABLE 7: Broker per Category -->
            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-calculator"></i>
                    Broker per Category GM
                    <button class="analysis-trigger" onclick="openTableAnalysis('brokerCategory')" title="Show analysis"><i class="bi bi-info-lg"></i></button>
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table" id="brokerPerCategoryTable">
                        <thead>
                            <tr>
                                <th>Category</th>
                                <th class="actual-gm-label">Actual GM</th>
                                <th>Contribution %</th>
                                <th>Not Booked % from Act</th>
                                <th class="broker-vs-prev-label">Act vs Prev</th>
                                <th class="broker-vs-forecast-label">Act vs Forecast</th>
                                <th class="broker-vs-target-label">Act vs Target</th>
                            </tr>
                        </thead>
                        <tbody id="brokerPerCategoryBody">
                            <tr>
                                <td colspan="7" class="text-center">Select datasets above to view broker per category data</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- NEW TABLE 8: Sales per Country -->
            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-globe2"></i>
                    Sales & Gross Profit per Country
                    <button class="analysis-trigger" onclick="openTableAnalysis('country')" title="Show analysis"><i class="bi bi-info-lg"></i></button>
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table" id="salesPerCountryTable">
                        <thead>
                            <tr>
                                <th>Country</th>
                                <th class="prev-sales-label">Prev Sales</th>
                                <th class="prev-gm-label">Prev GM</th>
                                <th class="actual-sales-label">Actual Sales</th>
                                <th class="actual-gm-label">Actual GM</th>
                                <th class="forecast-sales-label">Forecast Sales</th>
                                <th class="forecast-gm-label">Forecast GM</th>
                                <th>GM%</th>
                                <th>Contribution %</th>
                                <th>Target Sales</th>
                                <th>Target GM</th>
                            </tr>
                        </thead>
                        <tbody id="salesPerCountryBody">
                            <tr>
                                <td colspan="11" class="text-center">Select datasets above to view country data</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- NEW TABLE 9: GM per Country -->
            <div class="comparison-table-container">
                <h3>
                    <i class="bi bi-globe2"></i>
                    Gross Margin per Country
                </h3>
                <div class="table-responsive">
                    <table class="comparison-table" id="gmPerCountryTable" style="display:none;">
                        <thead>
                            <tr>
                                <th>Country</th>
                                <th>Prev Year</th>
                                <th>Actual (Full Year)</th>
                                <th>Forecast</th>
                                <th>Contribution %</th>
                                <th>Target</th>
                            </tr>
                        </thead>
                        <tbody id="gmPerCountryBody">
                            <tr>
                                <td colspan="6" class="text-center">Select datasets above to view country GM data</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- SALES ANALYSIS TAB -->
        <div id="salesTab" class="tab-content tab-pane">
            <div class="analysis-grid">
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-calendar-check"></i> Monthly Sales Trend (LCY)</h4>
                    </div>
                    <div class="chart-container">
                        <canvas id="trendChart"></canvas>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-graph-up-arrow"></i> Monthly Gross Profit Trend</h4>
                    </div>
                    <div class="chart-container">
                        <canvas id="monthlyGpChart"></canvas>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-pie-chart"></i> Sales by Category</h4>
                    </div>
                    <div class="chart-container">
                        <canvas id="categoryPieChart"></canvas>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-people"></i> Top Customers by Sales</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Customer</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="topCustomersSalesTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-bar-chart"></i> Sales by Country</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Country</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">%</th>
                                </tr>
                            </thead>
                            <tbody id="salesByCountryTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-check2-square"></i> Invoice Status Mix</h4>
                    </div>
                    <div class="chart-container">
                        <canvas id="statusMixChart"></canvas>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-check-circle"></i> Status Analysis</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Status</th>
                                    <th class="text-right">Count</th>
                                    <th class="text-right">Sales</th>
                                </tr>
                            </thead>
                            <tbody id="statusAnalysisTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card insight-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-lightbulb"></i> Sales Insights</h4>
                    </div>
                    <div id="salesInsightsBody"></div>
                </div>
            </div>
        </div>

        <!-- PROFITABILITY ANALYSIS TAB -->
        <div id="profitabilityTab" class="tab-content tab-pane">
            <div class="analysis-grid-3col">
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-box"></i> Profit by Product</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Product</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="profitByProductTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-people"></i> Profit by Customer</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Customer</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="profitByCustomerTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-tags"></i> Profit by Category</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Category</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="profitByCategoryTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-truck"></i> Profit by Supplier</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Supplier</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="profitBySupplierTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-calendar-month"></i> Profit by Quarter</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Quarter</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="profitByQuarterTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-exclamation-triangle"></i> Loss-Making Products</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Product</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="lossMakingProductsTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card insight-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-lightbulb"></i> Profitability Insights</h4>
                    </div>
                    <div id="profitabilityInsightsBody"></div>
                </div>
            </div>
        </div>

        <!-- CUSTOMER ANALYSIS TAB -->
        <div id="customersTab" class="tab-content tab-pane">
            <div class="analysis-grid">
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-cash-stack"></i> Client Profitability</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Customer</th>
                                    <th class="text-right">Sales (JOD)</th>
                                    <th class="text-right">GP (JOD)</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="clientProfitabilityTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-exclamation-triangle"></i> Loss & Low-Margin Clients</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Customer</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                    <th>Customer Avg GP%</th>
                                    <th>Status</th>
                                </tr>
                            </thead>
                            <tbody id="lossLowMarginTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-arrow-left-right"></i> Purchase to Sales Days</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Customer</th>
                                    <th class="text-right">Avg P-S Days</th>
                                    <th class="text-right">Max P-S Days</th>
                                    <th class="text-right">Transaction Count</th>
                                </tr>
                            </thead>
                            <tbody id="purchaseSalesDaysTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-bar-chart"></i> Top Customers by Sales</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Customer</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">Qty</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="customerVolumeTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card insight-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-lightbulb"></i> Customer Insights</h4>
                    </div>
                    <div id="customerInsightsBody"></div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-trophy"></i> Top & Worst Customers</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Type</th>
                                    <th>Customer</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                    <th class="text-right">Qty</th>
                                    <th class="text-right">Invoices</th>
                                    <th class="text-right">% GP</th>
                                </tr>
                            </thead>
                            <tbody id="topWorstCustomersTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-bullseye"></i> Customer Profitability Matrix</h4>
                    </div>
                    <div class="chart-container">
                        <canvas id="customerMatrixChart"></canvas>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-person-x"></i> Churn & Risk Watch</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Customer</th>
                                    <th>Status</th>
                                    <th class="text-right">Current Sales</th>
                                    <th class="text-right">Prev Sales</th>
                                    <th class="text-right">Open %</th>
                                    <th class="text-right">Avg Credit Days</th>
                                </tr>
                            </thead>
                            <tbody id="customerRiskWatchTable"></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>

        <!-- PRODUCT PERFORMANCE TAB -->
        <div id="productsTab" class="tab-content tab-pane">
            <div class="analysis-grid">
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-bar-chart"></i> Top Products by Sales & Profitability</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Product</th>
                                    <th class="text-right">Qty</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="bestSellingTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-tags"></i> Products by Category</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Category</th>
                                    <th class="text-right">Products</th>
                                    <th class="text-right">Sales</th>
                                </tr>
                            </thead>
                            <tbody id="productsByCategoryTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-pie-chart"></i> Category Contribution</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Category</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">% of Total</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="categoryContributionTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-receipt"></i> Invoice Count per Product</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Product</th>
                                    <th class="text-right">Invoice Count</th>
                                    <th class="text-right">Avg Invoice Value</th>
                                    <th class="text-right">Total Sales</th>
                                </tr>
                            </thead>
                            <tbody id="invoiceCountPerProductTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-trophy"></i> Best Vendor for Each Product</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Product</th>
                                    <th>Best Vendor</th>
                                    <th class="text-right">Total Purchases</th>
                                    <th class="text-right">Avg GP%</th>
                                </tr>
                            </thead>
                            <tbody id="bestVendorPerProductTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card insight-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-lightbulb"></i> Product Insights</h4>
                    </div>
                    <div id="productInsightsBody"></div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-trophy"></i> Top & Worst Products</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Type</th>
                                    <th>Product</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                    <th class="text-right">Qty</th>
                                    <th class="text-right">% GP</th>
                                </tr>
                            </thead>
                            <tbody id="topWorstProductsTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-bullseye"></i> Product Profitability Matrix</h4>
                    </div>
                    <div class="chart-container">
                        <canvas id="productMatrixChart"></canvas>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-hourglass-split"></i> Slow Moving & Obsolete</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Product</th>
                                    <th>Category</th>
                                    <th class="text-right">Last Sale</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">Qty</th>
                                </tr>
                            </thead>
                            <tbody id="slowMovingProductsTable"></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>

        <!-- SUPPLIER ANALYSIS TAB -->
        <div id="suppliersTab" class="tab-content tab-pane">
            <div class="analysis-grid">
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-cash-stack"></i> Vendor GP & GP%</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Supplier</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="vendorGpTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-arrow-left-right"></i> Purchase to Sales Days</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Supplier</th>
                                    <th class="text-right">Avg P-S Days</th>
                                    <th class="text-right">Max P-S Days</th>
                                    <th class="text-right">Transactions</th>
                                </tr>
                            </thead>
                            <tbody id="supplierPsDaysTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-exclamation-triangle"></i> Loss-Making Vendor Items</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Supplier</th>
                                    <th>Product</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="lossMakingVendorItemsTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-bar-chart"></i> Top Suppliers by Sales</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Supplier</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">Qty</th>
                                    <th class="text-right">GP%</th>
                                </tr>
                            </thead>
                            <tbody id="supplierVolumeTable"></tbody>
                        </table>
                    </div>
                </div>

                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-calendar-check"></i> Vendor Credit Days</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Supplier</th>
                                    <th class="text-right">Avg Credit Days</th>
                                    <th class="text-right">Max Credit Days</th>
                                    <th class="text-right">Transactions</th>
                                </tr>
                            </thead>
                            <tbody id="vendorCreditDaysTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card insight-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-lightbulb"></i> Supplier Insights</h4>
                    </div>
                    <div id="supplierInsightsBody"></div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-trophy"></i> Top & Worst Suppliers</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Type</th>
                                    <th>Supplier</th>
                                    <th class="text-right">Purchase</th>
                                    <th class="text-right">GP</th>
                                    <th class="text-right">GP%</th>
                                    <th class="text-right">% GP</th>
                                </tr>
                            </thead>
                            <tbody id="topWorstSuppliersTable"></tbody>
                        </table>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-bullseye"></i> Supplier Margin Matrix</h4>
                    </div>
                    <div class="chart-container">
                        <canvas id="supplierMatrixChart"></canvas>
                    </div>
                </div>
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-shield-exclamation"></i> Supplier Concentration & Risk</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Supplier</th>
                                    <th class="text-right">Purchase</th>
                                    <th class="text-right">% Purchase</th>
                                    <th class="text-right">Avg Credit</th>
                                    <th class="text-right">Avg P-S Days</th>
                                </tr>
                            </thead>
                            <tbody id="supplierRiskTable"></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>

        <div id="trackerTab" class="tab-content tab-pane">
            <div class="tracker-toolbar">
                <div class="tracker-note">
                    Review Booked and Not Booked pipeline by month and sales person. To update the Excel sheet, export the changes and save the file as `tracker_updates.csv` in this folder, then run `python dashboard.py`.
                </div>
                <button class="refresh-btn" onclick="exportTrackerUpdates()">
                    <i class="bi bi-download"></i>
                    Export Tracker Updates
                </button>
            </div>

            <div class="tracker-summary-grid">
                <div class="chart-card">
                    <div class="chart-header">
                        <h4><i class="bi bi-calendar-x"></i> Not Booked by Month</h4>
                    </div>
                    <div class="table-responsive">
                        <table class="inner-table">
                            <thead>
                                <tr>
                                    <th>Month</th>
                                    <th class="text-right">Rows</th>
                                    <th class="text-right">Sales</th>
                                    <th class="text-right">GP</th>
                                </tr>
                            </thead>
                            <tbody id="notBookedMonthTable"></tbody>
                        </table>
                    </div>
                </div>
            </div>

            <div class="comparison-table-container">
                <div class="tracker-header-bar">
                    <h3 style="margin-bottom:0;"><i class="bi bi-arrow-left-right"></i> Booked & Not Booked Tracker</h3>
                    <div class="tracker-header-filter">
                        <div id="trackerFileFilter"></div>
                    </div>
                </div>
                <div class="tracker-table-shell">
                    <table class="comparison-table tracker-table">
                        <thead>
                            <tr>
                                <th>Status</th>
                                <th>Month</th>
                                <th>Sales Person</th>
                                <th>Client</th>
                                <th>Product</th>
                                <th class="text-right">Sales</th>
                                <th class="text-right">GP</th>
                                <th>Action</th>
                                <th>New Status</th>
                                <th>New Month</th>
                                <th>Reason</th>
                            </tr>
                            <tr class="tracker-filter-row">
                                <th><div id="trackerStatusFilter"></div></th>
                                <th><div id="trackerMonthFilter"></div></th>
                                <th><div id="trackerSalesPersonFilter"></div></th>
                                <th><div id="trackerClientFilter"></div></th>
                                <th><div id="trackerProductFilter"></div></th>
                                <th></th>
                                <th></th>
                                <th colspan="4" style="text-align:right;">
                                    <button class="refresh-btn" type="button" onclick="resetTrackerFilters()">
                                        <i class="bi bi-arrow-counterclockwise"></i>
                                        Reset
                                    </button>
                                </th>
                            </tr>
                        </thead>
                        <tbody id="trackerCombinedBody"></tbody>
                    </table>
                </div>
            </div>
        </div>

        <div class="alert alert-info" id="dataInfo" style="display: none;">
            <i class="bi bi-info-circle-fill"></i>
            <span id="dataMessage"></span>
        </div>

        <div class="modal fade" id="tableAnalysisModal" tabindex="-1" aria-hidden="true">
            <div class="modal-dialog modal-lg modal-dialog-scrollable">
                <div class="modal-content">
                    <div class="modal-header">
                        <h5 class="modal-title" id="tableAnalysisModalTitle">Table Analysis</h5>
                        <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
                    </div>
                    <div class="modal-body" id="tableAnalysisModalBody"></div>
                    <div class="modal-footer">
                        <button type="button" class="btn btn-outline-primary btn-sm" onclick="copyTableAnalysis()">Copy to clipboard</button>
                        <button type="button" class="btn btn-secondary btn-sm" data-bs-dismiss="modal">Close</button>
                    </div>
                </div>
            </div>
        </div>

        <div class="footer">
            <p>
                <i class="bi bi-database"></i> S&OP Data: Data Sheet | Analysis Data: Demand Analysis Sheet | 
                <i class="bi bi-clock"></i> Last Updated: <span id="lastUpdated"></span> | 
                <i class="bi bi-person"></i> Created By ENG.Ashraf Harayri
            </p>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/flatpickr"></script>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script>
        const dashboardData = {{ chart_data }};
        Chart.defaults.font.size = 12;
        Chart.defaults.font.family = 'Inter, sans-serif';
        Chart.defaults.color = '#334155';
        Chart.defaults.plugins.legend.labels.boxWidth = 10;
        let datasets = [];
        let salesStatusChart = null;
        let gpStatusChart = null;
        let trendChart = null;
        let categoryPieChart = null;
        let monthlyGpChart = null;
        let statusMixChart = null;
        let customerMatrixChart = null;
        let productMatrixChart = null;
        let supplierMatrixChart = null;
        const chartRegistry = {};
        
        let currentFromDate = null;
        let currentToDate = null;
        let latestSopAnalysis = {};
        let tableAnalysisModal = null;
        let combinedVarianceRowsVisible = true;
        let trackerEdits = {};
        let trackerRecordLookup = {};
        let trackerRendered = false;
        let trackerFilters = { file: [], status: [], month: [], sales_person: [], client: [], product: [] };

        function showRuntimeError(message) {
            const badge = document.getElementById('dataSourceBadge');
            if (badge) {
                badge.innerHTML = '<i class="bi bi-exclamation-octagon"></i> Script Error';
                badge.style.background = '#dc2626';
            }
            const rangeEl = document.getElementById('dataDateRange');
            if (rangeEl) {
                rangeEl.innerText = message;
            }
            const info = document.getElementById('dataInfo');
            const infoMsg = document.getElementById('dataMessage');
            if (info && infoMsg) {
                info.style.display = 'block';
                info.className = 'alert alert-danger';
                infoMsg.textContent = message;
            }
        }

        window.onerror = function(message, source, lineno, colno) {
            showRuntimeError(`JS error: ${message} @ line ${lineno || 0}`);
            return false;
        };

        document.addEventListener('click', function(event) {
            if (!event.target.closest('.tracker-filter-dropdown')) {
                closeTrackerDropdowns();
            }
        });
        
        const monthlyTargets = {
            1: { month: 'Jan', gp: 70955.6889, sales: 1340123.793, qty: 1371.74, commission: 3906.42 },
            2: { month: 'Feb', gp: 109681.2675, sales: 1865415.873, qty: 1730, commission: 9053.4443 },
            3: { month: 'Mar', gp: 89767.79265, sales: 2255162.847, qty: 2084.71, commission: 9520.0066 },
            4: { month: 'Apr', gp: 122249.2298, sales: 2329650.58, qty: 2198.465, commission: 7330.4163 },
            5: { month: 'May', gp: 139237.0366, sales: 2254543.159, qty: 2143.86, commission: 8535.2863 },
            6: { month: 'Jun', gp: 169113.3709, sales: 3296326.514, qty: 2982.04, commission: 8717.4368 },
            7: { month: 'Jul', gp: 147454.483, sales: 2324379.15, qty: 2200.96, commission: 15160.0265 },
            8: { month: 'Aug', gp: 126528.9325, sales: 2418520.393, qty: 2309.1, commission: 7019.628 },
            9: { month: 'Sep', gp: 150593.5497, sales: 2611603.899, qty: 2183.04, commission: 9231.846 },
            10: { month: 'Oct', gp: 83173.66766, sales: 1925567.18, qty: 1783.21, commission: 9770.878 },
            11: { month: 'Nov', gp: 100630.7955, sales: 1951227.219, qty: 1923.19, commission: 7409.019202 },
            12: { month: 'Dec', gp: 103734.6638, sales: 1755026.813, qty: 1601.27, commission: 9347.25295 }
        };
        
        const monthNames = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'];
        
        function getStatusForMonth(selectedSopMonth, currentMonth, datasetType) {
            if (datasetType === 'actual') {
                if (currentMonth <= selectedSopMonth) {
                    return ['Invoiced'];
                } else {
                    return ['Booked', 'Not Booked'];
                }
            } else if (datasetType === 'forecast') {
                return ['Booked', 'Not Booked'];
            } else {
                return ['Invoiced', 'Booked', 'Not Booked'];
            }
        }
        
        function getCommissionForMonth(dataset, monthAbbr, selectedSopMonth) {
            if (!dataset) return 0;
            const cleanName = dataset.replace(/[^a-zA-Z0-9]/g, '_');
            const commissionKey = `${cleanName}_Broker_${monthAbbr}`;
            return dashboardData.commission_data && dashboardData.commission_data[commissionKey] ? dashboardData.commission_data[commissionKey] : 0;
        }
        
        function getFilteredDataForMonth(dataset, monthAbbr, selectedSopMonth, datasetType) {
            if (!dataset) return { sales: 0, qty: 0, gm: 0, commission: 0, gp: 0, gp_percent: 0, gm_percent: 0 };
            
            const monthNum = monthNames.indexOf(monthAbbr) + 1;
            const cleanName = dataset.replace(/[^a-zA-Z0-9]/g, '_');
            
            const commissionKey = `${cleanName}_Broker_${monthAbbr}`;
            const commission = dashboardData.commission_data && dashboardData.commission_data[commissionKey] ? dashboardData.commission_data[commissionKey] : 0;
            
            let tradeData = { sales: 0, qty: 0, gm: 0 };
            
            if (datasetType === 'prevYear') {
                const prevYearKey = `${cleanName}_Trade_All_Status_${monthAbbr}`;
                if (dashboardData.filtered_data && dashboardData.filtered_data[prevYearKey]) {
                    const data = dashboardData.filtered_data[prevYearKey];
                    tradeData.sales = data.sales || 0;
                    tradeData.qty = data.qty || 0;
                    tradeData.gm = data.gp || 0;
                }
            } else if (datasetType === 'forecast') {
                const forecastKey = `${cleanName}_Trade_All_Status_${monthAbbr}`;
                if (dashboardData.filtered_data && dashboardData.filtered_data[forecastKey]) {
                    const data = dashboardData.filtered_data[forecastKey];
                    tradeData.sales = data.sales || 0;
                    tradeData.qty = data.qty || 0;
                    tradeData.gm = data.gp || 0;
                } else {
                    const fallbackKey = `${cleanName}_Trade_Booked_Not Booked_${monthAbbr}`;
                    if (dashboardData.filtered_data && dashboardData.filtered_data[fallbackKey]) {
                        const data = dashboardData.filtered_data[fallbackKey];
                        tradeData.sales = data.sales || 0;
                        tradeData.qty = data.qty || 0;
                        tradeData.gm = data.gp || 0;
                    }
                }
            } else {
                const statuses = getStatusForMonth(selectedSopMonth, monthNum, datasetType);
                let filteredKey = `${cleanName}_Trade_${statuses.join('_')}_${monthAbbr}`;
                if (dashboardData.filtered_data && dashboardData.filtered_data[filteredKey]) {
                    const data = dashboardData.filtered_data[filteredKey];
                    tradeData.sales = data.sales || 0;
                    tradeData.qty = data.qty || 0;
                    tradeData.gm = data.gp || 0;
                } else {
                    const fallbackKey = `${cleanName}_Trade_${monthAbbr}`;
                    if (dashboardData.filtered_data && dashboardData.filtered_data[fallbackKey]) {
                        const data = dashboardData.filtered_data[fallbackKey];
                        tradeData.sales = data.sales || 0;
                        tradeData.qty = data.qty || 0;
                        tradeData.gm = data.gp || 0;
                    }
                }
            }
            
            const gp = tradeData.gm + commission;
            const gp_percent = tradeData.sales > 0 ? (gp / tradeData.sales) * 100 : 0;
            const gm_percent = tradeData.sales > 0 ? (tradeData.gm / tradeData.sales) * 100 : 0;
            
            return {
                sales: tradeData.sales,
                qty: tradeData.qty,
                gm: tradeData.gm,
                gm_percent: gm_percent,
                commission: commission,
                gp: gp,
                gp_percent: gp_percent
            };
        }
        
        function formatNumber(num) {
            if (num === undefined || num === null || isNaN(num)) return '0';
            return new Intl.NumberFormat('en-US', { minimumFractionDigits: 0, maximumFractionDigits: 0 }).format(num);
        }
        
        function formatNumberDecimal(num, decimals = 2) {
            if (num === undefined || num === null || isNaN(num)) return '0';
            return new Intl.NumberFormat('en-US', { minimumFractionDigits: decimals, maximumFractionDigits: decimals }).format(num);
        }

        function formatCompactNumber(num) {
            if (num === undefined || num === null || isNaN(num)) return '0';
            const abs = Math.abs(num);
            if (abs >= 1000000) return (num / 1000000).toFixed(1) + ' M';
            if (abs >= 1000) return (num / 1000).toFixed(1) + ' k';
            return formatNumber(num);
        }
        
        function formatPercent(num) {
            if (num === undefined || num === null || isNaN(num)) return '0%';
            return num.toFixed(1) + '%';
        }
        
        function getVarianceClass(val) {
            if (val > 0) return 'variance-positive';
            if (val < 0) return 'variance-negative';
            return '';
        }
        
        function getVarianceHtml(actual, compare, isPercentage = false) {
            const diff = actual - compare;
            let diffPct = 0;
            if (compare !== 0) {
                diffPct = (diff / Math.abs(compare)) * 100;
            } else if (actual !== 0) {
                diffPct = 100;
            }
            
            const diffClass = getVarianceClass(diff);
            const sign = diff >= 0 ? '+' : '';
            
            if (isPercentage) {
                return `<span class="${diffClass}">${sign}${diff.toFixed(1)}% (${diffPct.toFixed(1)}%)</span>`;
            } else {
                return `<span class="${diffClass}">${sign}${formatNumber(Math.abs(diff))}</span><br><small class="${diffClass}">(${diffPct.toFixed(1)}%)</small>`;
            }
        }

        function getVarianceCellHtml(diff, baseValue) {
            const diffClass = getVarianceClass(diff);
            const diffPct = baseValue !== 0 ? (diff / Math.abs(baseValue)) * 100 : (diff !== 0 ? 100 : 0);
            const sign = diff >= 0 ? '+' : '-';
            return `<span class="${diffClass}">${sign}${formatNumber(Math.abs(diff))}</span><br><small class="${diffClass}">(${diffPct.toFixed(1)}%)</small>`;
        }

        function getVariancePercentCellHtml(diff, baseValue) {
            const diffClass = getVarianceClass(diff);
            const diffPct = baseValue !== 0 ? (diff / Math.abs(baseValue)) * 100 : (diff !== 0 ? 100 : 0);
            const sign = diff >= 0 ? '+' : '-';
            return `<span class="${diffClass}">${sign}${Math.abs(diff).toFixed(1)}pp</span><br><small class="${diffClass}">(${diffPct.toFixed(1)}%)</small>`;
        }

        function getVarianceRateCellHtml(diff, baseValue) {
            const diffClass = getVarianceClass(diff);
            const diffPct = baseValue !== 0 ? (diff / Math.abs(baseValue)) * 100 : (diff !== 0 ? 100 : 0);
            const sign = diffPct >= 0 ? '+' : '-';
            return `<span class="${diffClass}">${sign}${Math.abs(diffPct).toFixed(1)}%</span>`;
        }

        function getVariancePctValue(diff, baseValue) {
            return baseValue !== 0 ? (diff / Math.abs(baseValue)) * 100 : (diff !== 0 ? 100 : 0);
        }

        function renderInsightsList(containerId, insights) {
            const el = document.getElementById(containerId);
            if (!el) return;
            if (!insights || !insights.length) {
                el.innerHTML = '<div class="text-center" style="padding: 18px;">No insights available</div>';
                return;
            }
            el.innerHTML = `<ul class="insight-list">${insights.map(item => `<li>${item}</li>`).join('')}</ul>`;
        }

        function getTopItemName(items, fallback = 'N/A') {
            return items && items.length ? items[0].name || fallback : fallback;
        }

        function buildAnalysisHtml(title, lines) {
            return `<div><div style="font-weight:700; color:#1e3c72; margin-bottom:10px;">${title}</div><ul class="insight-list">${lines.map(line => `<li>${line}</li>`).join('')}</ul></div>`;
        }

        function openTableAnalysis(type) {
            const modalEl = document.getElementById('tableAnalysisModal');
            if (!modalEl) return;
            if (!tableAnalysisModal) tableAnalysisModal = new bootstrap.Modal(modalEl);
            const ctx = latestSopAnalysis || {};
            let title = 'Table Analysis';
            let html = '<div class="text-center" style="padding:18px;">No analysis available</div>';
            if (type === 'month') {
                title = 'Month Analysis';
                html = buildAnalysisHtml(title, ctx.monthAnalysis || []);
            } else if (type === 'ytd') {
                title = 'YTD Performance';
                html = buildAnalysisHtml(title, ctx.ytdAnalysis || []);
            } else if (type === 'fullYear') {
                title = 'Full Year Snapshot';
                html = buildAnalysisHtml(title, ctx.fullYearAnalysis || []);
            } else if (type === 'combined') {
                title = 'Monthly Trend Analysis';
                html = buildAnalysisHtml(title, ctx.combinedAnalysis || []);
            } else if (type === 'client') {
                title = 'Client Profitability Analysis';
                html = buildAnalysisHtml(title, ctx.clientAnalysis || []);
            } else if (type === 'category') {
                title = 'Category Performance';
                html = buildAnalysisHtml(title, ctx.categoryAnalysis || []);
            } else if (type === 'supplier') {
                title = 'Supplier Health Check';
                html = buildAnalysisHtml(title, ctx.supplierAnalysis || []);
            } else if (type === 'brokerClient') {
                title = 'Broker Contribution by Client';
                html = buildAnalysisHtml(title, ctx.brokerClientAnalysis || []);
            } else if (type === 'brokerCategory') {
                title = 'Broker Performance by Category';
                html = buildAnalysisHtml(title, ctx.brokerCategoryAnalysis || []);
            } else if (type === 'country') {
                title = 'Country Performance';
                html = buildAnalysisHtml(title, ctx.countryAnalysis || []);
            }
            document.getElementById('tableAnalysisModalTitle').textContent = title;
            document.getElementById('tableAnalysisModalBody').innerHTML = html;
            tableAnalysisModal.show();
        }

        function copyTableAnalysis() {
            const body = document.getElementById('tableAnalysisModalBody');
            if (!body) return;
            const text = body.innerText || body.textContent || '';
            if (navigator.clipboard && text) {
                navigator.clipboard.writeText(text);
            }
        }

        function addGrandTotalRow(tableBodyId, totals, hasPercentageColumns = []) {
            const tableBody = document.getElementById(tableBodyId);
            if (!tableBody) return;
            
            const rows = tableBody.querySelectorAll('tr');
            if (rows.length === 0) return;
            
            const totalRow = document.createElement('tr');
            totalRow.className = 'grand-total-row';
            
            const firstCell = document.createElement('td');
            firstCell.textContent = 'GRAND TOTAL';
            firstCell.className = 'font-bold';
            totalRow.appendChild(firstCell);
            
            for (let i = 0; i < totals.length; i++) {
                const cell = document.createElement('td');
                cell.className = 'text-right font-bold';
                if (hasPercentageColumns.includes(i)) {
                    cell.textContent = formatPercent(totals[i]);
                } else {
                    cell.textContent = formatNumber(totals[i]);
                }
                totalRow.appendChild(cell);
            }
            
            tableBody.appendChild(totalRow);
        }
        
        function initDashboard() {
            try {
            console.log('Dashboard Data:', dashboardData);
            
            const hasSummaryData = dashboardData.achievement && dashboardData.achievement.length > 0;
            
            if (!hasSummaryData) {
                document.getElementById('noSummaryData').style.display = 'block';
                document.getElementById('achievementCard').style.display = 'none';
                document.getElementById('dataSourceBadge').innerHTML = '<i class="bi bi-exclamation-triangle"></i> Missing Summary Sheet';
                document.getElementById('dataSourceBadge').style.background = '#dc2626';
                return;
            }
            
            const sourceBadge = document.getElementById('dataSourceBadge');
            sourceBadge.innerHTML = '<i class="bi bi-database-check"></i> Live Excel Data';
            sourceBadge.style.background = '#10b981';
            
            flatpickr(".datepicker", {
                dateFormat: "m/d/Y",
                allowInput: true
            });
            
            if (dashboardData.file_names && dashboardData.file_names.length > 0) {
                datasets = dashboardData.file_names;
                populateDropdowns();
            }
            populateTrackerFilters();
            
            updateAchievementBars();
            try { updateDataInfo(); } catch (error) { console.error('updateDataInfo failed', error); }
            ['gmPerClientTable', 'gmPerCategoryTable', 'gmPerCountryTable'].forEach(id => {
                const table = document.getElementById(id);
                if (table) {
                    const container = table.closest('.comparison-table-container');
                    if (container) container.style.display = 'none';
                }
            });
            
            try { renderSalesAnalysis(); } catch (error) { console.error('renderSalesAnalysis failed', error); }
            try { renderProfitabilityAnalysis(); } catch (error) { console.error('renderProfitabilityAnalysis failed', error); }
            try { renderCustomerAnalysis(); } catch (error) { console.error('renderCustomerAnalysis failed', error); }
            try { renderProductAnalysis(); } catch (error) { console.error('renderProductAnalysis failed', error); }
            try { renderSupplierAnalysis(); } catch (error) { console.error('renderSupplierAnalysis failed', error); }
            const currentDate = new Date();
            const currentMonth = currentDate.getMonth() + 1;
            const currentSOP = datasets.find(d => d === `S&OP${currentMonth}`);
            
            if (currentSOP) {
                document.getElementById('actualDataset').value = currentSOP;
                
                const prevMonth = currentMonth - 1;
                if (prevMonth >= 1) {
                    const prevSOP = datasets.find(d => d === `S&OP${prevMonth}`);
                    if (prevSOP) {
                        document.getElementById('forecastDataset').value = prevSOP;
                    }
                }
                
                const prevYearDataset = datasets.find(d => d === '2024' || d === '2023' || d === '2022' || d.includes('2024'));
                if (prevYearDataset) {
                    document.getElementById('prevYearDataset').value = prevYearDataset;
                }
                
                updateMonthDisplay(currentSOP);
                setTimeout(() => refreshComparison(), 100);
            }
            } catch (error) {
                console.error('initDashboard failed', error);
                showRuntimeError(`Dashboard init failed: ${error && error.message ? error.message : error}`);
            }
        }
        
        function applyDateFilter() {
            const fromDate = document.getElementById('fromDate').value;
            const toDate = document.getElementById('toDate').value;
            
            if (!fromDate || !toDate) {
                alert('Please select both From and To dates');
                return;
            }
            
            currentFromDate = fromDate;
            currentToDate = toDate;
            
            document.getElementById('filterDates').innerHTML = fromDate + ' - ' + toDate;
            document.getElementById('activeFilterBadge').style.display = 'inline-flex';
            
            renderSalesAnalysis();
            renderProfitabilityAnalysis();
            renderCustomerAnalysis();
            renderProductAnalysis();
            renderSupplierAnalysis();
        }
        
        function resetDateFilter() {
            document.getElementById('fromDate').value = '';
            document.getElementById('toDate').value = '';
            currentFromDate = null;
            currentToDate = null;
            document.getElementById('activeFilterBadge').style.display = 'none';
            
            renderSalesAnalysis();
            renderProfitabilityAnalysis();
            renderCustomerAnalysis();
            renderProductAnalysis();
            renderSupplierAnalysis();
        }
        
        function filterDataByDate(dataArray, dateKey) {
            if (!currentFromDate || !currentToDate || !dataArray) {
                return dataArray;
            }
            
            const fromDateObj = new Date(currentFromDate);
            const toDateObj = new Date(currentToDate);
            
            return dataArray.filter(item => {
                const itemDate = item[dateKey] ? new Date(item[dateKey]) : null;
                return itemDate && itemDate >= fromDateObj && itemDate <= toDateObj;
            });
        }
        
        function filterTrendDataByDate(data) {
            if (!currentFromDate || !currentToDate || !data || !data.labels || !data.values) {
                return data;
            }
            
            const fromDateObj = new Date(currentFromDate);
            const toDateObj = new Date(currentToDate);
            
            const filteredLabels = [];
            const filteredValues = [];
            
            for (let i = 0; i < data.labels.length; i++) {
                const labelDate = new Date(data.labels[i]);
                if (!isNaN(labelDate) && labelDate >= fromDateObj && labelDate <= toDateObj) {
                    filteredLabels.push(data.labels[i]);
                    filteredValues.push(data.values[i]);
                }
            }
            
            return {
                labels: filteredLabels,
                values: filteredValues
            };
        }

        function getDemandRecords() {
            return dashboardData.demand_records || [];
        }

        function getFilteredDemandRecords() {
            const records = getDemandRecords();
            if (!currentFromDate || !currentToDate) return records;
            const fromDateObj = new Date(currentFromDate);
            const toDateObj = new Date(currentToDate);
            toDateObj.setHours(23, 59, 59, 999);
            return records.filter(record => {
                if (!record.posting_date) return false;
                const postingDate = new Date(record.posting_date);
                return !isNaN(postingDate) && postingDate >= fromDateObj && postingDate <= toDateObj;
            });
        }

        function aggregateDemandRecords(records, groupField) {
            const grouped = {};
            records.forEach(record => {
                const key = record[groupField] || 'N/A';
                if (!grouped[key]) {
                    grouped[key] = { sales: 0, gp: 0, qty: 0, count: 0 };
                }
                grouped[key].sales += Number(record.sales_jod || 0);
                grouped[key].gp += Number(record.gp_jod || 0);
                grouped[key].qty += Number(record.qty_mt || 0);
                grouped[key].count += 1;
            });
            return Object.entries(grouped).map(([name, values]) => ({
                name,
                sales: values.sales,
                gp: values.gp,
                qty: values.qty,
                count: values.count,
                gm: values.sales > 0 ? (values.gp / values.sales) * 100 : 0
            }));
        }

        function buildMonthlySeries(records, valueField) {
            const grouped = {};
            records.forEach(record => {
                if (!record.posting_date) return;
                const date = new Date(record.posting_date);
                if (isNaN(date)) return;
                const key = `${date.getFullYear()}-${String(date.getMonth() + 1).padStart(2, '0')}`;
                grouped[key] = (grouped[key] || 0) + Number(record[valueField] || 0);
            });
            const labels = Object.keys(grouped).sort();
            return {
                labels,
                values: labels.map(label => grouped[label])
            };
        }

        function buildEntityStats(records, groupField) {
            const grouped = {};
            records.forEach(record => {
                const key = record[groupField] || 'N/A';
                if (!grouped[key]) {
                    grouped[key] = {
                        name: key,
                        sales: 0,
                        gp: 0,
                        qty: 0,
                        count: 0,
                        invoices: new Set(),
                        openCount: 0,
                        lastDate: null,
                        categorySet: new Set(),
                        productSet: new Set(),
                        customerSet: new Set(),
                        supplierSet: new Set(),
                        purchase: 0,
                        creditDays: [],
                        psDays: []
                    };
                }
                const item = grouped[key];
                item.sales += Number(record.sales_jod || 0);
                item.gp += Number(record.gp_jod || 0);
                item.qty += Number(record.qty_mt || 0);
                item.count += 1;
                item.purchase += Number(record.purchase_amount_lcy || record.purchase_amount || 0);
                if (record.invoice_no) item.invoices.add(record.invoice_no);
                if (String(record.status || '').toLowerCase().includes('open')) item.openCount += 1;
                if (record.category) item.categorySet.add(record.category);
                if (record.product_name) item.productSet.add(record.product_name);
                if (record.customer_name) item.customerSet.add(record.customer_name);
                if (record.supplier_name) item.supplierSet.add(record.supplier_name);
                if (record.posting_date) {
                    const postingDate = new Date(record.posting_date);
                    if (!isNaN(postingDate) && (!item.lastDate || postingDate > item.lastDate)) item.lastDate = postingDate;
                    const dueDate = record.due_date ? new Date(record.due_date) : null;
                    const purchDate = record.purch_posting_date ? new Date(record.purch_posting_date) : null;
                    if (dueDate && !isNaN(dueDate) && postingDate && !isNaN(postingDate)) {
                        const creditDays = (dueDate - postingDate) / (1000 * 60 * 60 * 24);
                        if (creditDays >= 0) item.creditDays.push(creditDays);
                    }
                    if (purchDate && !isNaN(purchDate) && postingDate && !isNaN(postingDate)) {
                        const psDays = (postingDate - purchDate) / (1000 * 60 * 60 * 24);
                        if (psDays >= 0) item.psDays.push(psDays);
                    }
                }
            });
            return Object.values(grouped).map(item => ({
                ...item,
                invoices: item.invoices.size,
                gm: item.sales > 0 ? (item.gp / item.sales) * 100 : 0,
                gpShare: 0,
                openPct: item.count > 0 ? (item.openCount / item.count) * 100 : 0,
                avgCreditDays: item.creditDays.length ? item.creditDays.reduce((a, b) => a + b, 0) / item.creditDays.length : 0,
                avgPsDays: item.psDays.length ? item.psDays.reduce((a, b) => a + b, 0) / item.psDays.length : 0,
                lastDateText: item.lastDate ? item.lastDate.toISOString().slice(0, 10) : 'N/A',
                categories: Array.from(item.categorySet),
                products: Array.from(item.productSet),
                customers: Array.from(item.customerSet),
                suppliers: Array.from(item.supplierSet)
            }));
        }

        function attachGpShare(items) {
            const totalGp = items.reduce((sum, item) => sum + item.gp, 0);
            return items.map(item => ({ ...item, gpShare: totalGp !== 0 ? (item.gp / totalGp) * 100 : 0 }));
        }

        function buildTopWorstTableRows(items, valueKey, formatter, extraCellsBuilder) {
            const top = [];
            const worst = [];
            const topKeys = new Set();
            [...items].sort((a, b) => b[valueKey] - a[valueKey]).forEach(item => {
                const key = String(item.name || '').trim();
                if (top.length < 5 && !topKeys.has(key)) {
                    top.push({ bucket: 'Top', item });
                    topKeys.add(key);
                }
            });
            [...items].sort((a, b) => a[valueKey] - b[valueKey]).forEach(item => {
                const key = String(item.name || '').trim();
                if (worst.length < 5 && !topKeys.has(key) && !worst.some(entry => String(entry.item.name || '').trim() === key)) {
                    worst.push({ bucket: 'Worst', item });
                }
            });
            return [...top, ...worst].map(({ bucket, item }) => {
                const badgeClass = bucket === 'Top' ? 'badge-top' : 'badge-worst';
                return `<tr><td><span class="performance-badge ${badgeClass}">${bucket}</span></td>${extraCellsBuilder(item, formatter)}</tr>`;
            }).join('');
        }

        function renderBubbleMatrix(chartRefName, canvasId, items, xKey, yKey, rKey, labelKey, colorTop = 'rgba(37,99,235,0.65)', colorWorst = 'rgba(220,38,38,0.65)') {
            const canvas = document.getElementById(canvasId);
            if (!canvas) return null;
            if (chartRegistry[chartRefName]) chartRegistry[chartRefName].destroy();
            const topItems = [...items].sort((a, b) => b.gp - a.gp).slice(0, 5);
            const worstItems = [...items].sort((a, b) => a.gp - b.gp).slice(0, 5);
            chartRegistry[chartRefName] = new Chart(canvas.getContext('2d'), {
                type: 'bubble',
                data: {
                    datasets: [
                        {
                            label: 'Top 5',
                            data: topItems.map(item => ({ x: item[xKey], y: item[yKey], r: Math.max(6, Math.sqrt(Math.abs(item[rKey] || 0)) / 4), label: item[labelKey] || item.name })),
                            backgroundColor: colorTop
                        },
                        {
                            label: 'Worst 5',
                            data: worstItems.map(item => ({ x: item[xKey], y: item[yKey], r: Math.max(6, Math.sqrt(Math.abs(item[rKey] || 0)) / 4), label: item[labelKey] || item.name })),
                            backgroundColor: colorWorst
                        }
                    ]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        tooltip: {
                            callbacks: {
                                label: function(context) {
                                    const point = context.raw || {};
                                    return `${point.label || ''}: X=${formatNumber(point.x)}, Y=${formatPercent(point.y || 0)}`;
                                }
                            }
                        }
                    },
                    scales: {
                        x: { ticks: { callback: value => formatNumber(value) } },
                        y: { ticks: { callback: value => `${value.toFixed ? value.toFixed(1) : value}%` } }
                    }
                }
            });
            return chartRegistry[chartRefName];
        }

        function updateAchievementBars() {
            const achievement = dashboardData.achievement || [];
            const summaryRow = achievement.length > 0 ? achievement[0] : {};

            const salesTarget = Number(summaryRow['Total total sales Target YTD'] || 0);
            const salesActual = Number(summaryRow['Total Sales YTD'] || 0);
            const salesRemaining = Number(summaryRow['sales remaning YTD'] || 0);
            const salesPct = salesTarget > 0 ? (salesActual / salesTarget) * 100 : 0;

            const gpTradeTarget = Number(summaryRow['Total Gross profit Target YTD(trade)'] || 0);
            const gpBrokerTarget = Number(summaryRow['Total Gross profit Target YTD(broker)'] || 0);
            const gpTradeActual = Number(summaryRow['Total gross profit YTD(trade)'] || 0);
            const gpBrokerActual = Number(summaryRow['Total gross profit YTD(broker)'] || 0);
            const gpTotalRemaining = Number(summaryRow['total gross profit remaning YTD'] || 0);
            const gpTradePct = gpTradeTarget > 0 ? (gpTradeActual / gpTradeTarget) * 100 : 0;
            const gpBrokerPct = gpBrokerTarget > 0 ? (gpBrokerActual / gpBrokerTarget) * 100 : 0;

            document.getElementById('salesYtdPercent').textContent = formatPercent(salesPct);
            document.getElementById('salesYtdBar').style.width = Math.min(salesPct, 100) + '%';
            document.getElementById('salesYtdBar').textContent = formatPercent(salesPct);
            document.getElementById('salesYtdTarget').textContent = formatNumber(salesTarget) + ' JOD';
            document.getElementById('salesYtdActual').textContent = formatNumber(salesActual) + ' JOD';
            document.getElementById('salesYtdRemaining').textContent = formatNumber(salesRemaining) + ' JOD';

            document.getElementById('gpTradeYtdPercent').textContent = formatPercent(gpTradePct);
            document.getElementById('gpTradeYtdBar').style.width = Math.min(gpTradePct, 100) + '%';
            document.getElementById('gpTradeYtdBar').textContent = formatPercent(gpTradePct);
            document.getElementById('gpTradeYtdTarget').textContent = formatNumber(gpTradeTarget) + ' JOD';
            document.getElementById('gpTradeYtdActual').textContent = 'Actual: ' + formatNumber(gpTradeActual) + ' JOD';

            document.getElementById('gpBrokerYtdPercent').textContent = formatPercent(gpBrokerPct);
            document.getElementById('gpBrokerYtdBar').style.width = Math.min(gpBrokerPct, 100) + '%';
            document.getElementById('gpBrokerYtdBar').textContent = formatPercent(gpBrokerPct);
            document.getElementById('gpBrokerYtdTarget').textContent = formatNumber(gpBrokerTarget) + ' JOD';
            document.getElementById('gpBrokerYtdActual').textContent = 'Actual: ' + formatNumber(gpBrokerActual) + ' JOD';
            document.getElementById('gpTotalYtdRemaining').textContent = formatNumber(gpTotalRemaining) + ' JOD';

            [['salesYtdBar', salesPct], ['gpTradeYtdBar', gpTradePct], ['gpBrokerYtdBar', gpBrokerPct]].forEach(([id, pct]) => {
                const el = document.getElementById(id);
                el.classList.remove('positive', 'warning', 'danger');
                if (pct >= 100) el.classList.add('positive');
                else if (pct >= 80) el.classList.add('warning');
                else el.classList.add('danger');
            });
        }

        function populateDropdowns() {
            const selects = [
                document.getElementById('actualDataset'),
                document.getElementById('forecastDataset'),
                document.getElementById('prevYearDataset')
            ];
            
            selects.forEach(select => {
                while (select.options.length > 1) {
                    select.remove(1);
                }
                
                datasets.forEach(dataset => {
                    const option = document.createElement('option');
                    option.value = dataset;
                    option.textContent = dataset;
                    select.appendChild(option);
                });
            });
        }
        
        function updateSopDatasetLabels(actual, forecast, prevYear) {
            document.querySelectorAll('.actual-dataset-label').forEach(el => el.textContent = actual || 'Actual');
            document.querySelectorAll('.forecast-dataset-label').forEach(el => el.textContent = forecast || 'Forecast');
            document.querySelectorAll('.prev-dataset-label').forEach(el => el.textContent = prevYear || 'Prev Year');
            const labelMap = {
                '.actual-sales-label': `${actual || 'Actual'} Sales`,
                '.actual-gm-label': `${actual || 'Actual'} GM`,
                '.forecast-sales-label': `${forecast || 'Forecast'} Sales`,
                '.forecast-gm-label': `${forecast || 'Forecast'} GM`,
                '.prev-sales-label': `${prevYear || 'Prev Year'} Sales`,
                '.prev-gm-label': `${prevYear || 'Prev Year'} GM`,
                '.sales-vs-prev-label': `Sales vs ${prevYear || 'Prev Year'}`,
                '.gm-vs-prev-label': `GM vs ${prevYear || 'Prev Year'}`,
                '.sales-vs-forecast-label': `Sales vs ${forecast || 'Forecast'}`,
                '.gm-vs-forecast-label': `GM vs ${forecast || 'Forecast'}`,
                '.sales-vs-target-label': 'Sales vs Target',
                '.gm-vs-target-label': 'GM vs Target',
                '.broker-vs-prev-label': `Broker GM vs ${prevYear || 'Prev Year'}`,
                '.broker-vs-forecast-label': `Broker GM vs ${forecast || 'Forecast'}`,
                '.broker-vs-target-label': 'Broker GM vs Target',
                '.vs-forecast-label': `Vs ${forecast || 'Forecast'}`,
                '.vs-target-label': 'Vs Target',
                '.vs-prev-label': `Vs ${prevYear || 'Prev Year'}`
            };
            Object.entries(labelMap).forEach(([selector, text]) => {
                document.querySelectorAll(selector).forEach(el => el.textContent = text);
            });
        }

        function updateMonthDisplay(dataset) {
            const monthDisplay = document.getElementById('monthDisplay');
            const monthName = document.getElementById('monthName');
            const ytdMonthName = document.getElementById('ytdMonthName');
            
            if (!dataset) {
                monthDisplay.textContent = 'Not Selected';
                monthName.textContent = 'Selected Month';
                ytdMonthName.textContent = 'Selected Month';
                return;
            }
            
            const match = dataset.match(/S&OP(\\d+)/i);
            if (match) {
                const monthNum = parseInt(match[1]);
                const monthNamesLong = ['January', 'February', 'March', 'April', 'May', 'June',
                                   'July', 'August', 'September', 'October', 'November', 'December'];
                monthDisplay.textContent = monthNamesLong[monthNum - 1] + ' (Month ' + monthNum + ')';
                monthName.textContent = monthNamesLong[monthNum - 1];
                ytdMonthName.textContent = monthNamesLong[monthNum - 1];
            } else {
                monthDisplay.textContent = dataset;
                monthName.textContent = dataset;
                ytdMonthName.textContent = dataset;
            }
        }
        
        document.getElementById('actualDataset').addEventListener('change', function() {
            updateMonthDisplay(this.value);
        });
        
        function refreshComparison() {
            const actual = document.getElementById('actualDataset').value;
            const forecast = document.getElementById('forecastDataset').value;
            const prevYear = document.getElementById('prevYearDataset').value;
            
            if (!actual) {
                alert('Please select an Actual dataset first');
                return;
            }
            
            updateSopDatasetLabels(actual, forecast, prevYear);
            const loadingHtml = '<tr><td colspan="14" class="text-center"><div class="spinner-border text-primary" style="width: 1rem; height: 1rem;"></div> Loading...</td></tr>';
            document.getElementById('monthComparisonTableBody').innerHTML = loadingHtml;
            document.getElementById('ytdComparisonTableBody').innerHTML = loadingHtml;
            document.getElementById('fullYearComparisonTableBody').innerHTML = loadingHtml;
            document.getElementById('salesComparisonBody').innerHTML = loadingHtml;
            document.getElementById('gpComparisonBody').innerHTML = loadingHtml;
            document.getElementById('qtyComparisonBody').innerHTML = loadingHtml;
            document.getElementById('salesPerClientBody').innerHTML = loadingHtml;
            document.getElementById('salesPerCategoryBody').innerHTML = loadingHtml;
            document.getElementById('suppliersBody').innerHTML = loadingHtml;
            document.getElementById('brokerPerClientBody').innerHTML = loadingHtml;
            document.getElementById('brokerPerCategoryBody').innerHTML = loadingHtml;
            document.getElementById('salesPerCountryBody').innerHTML = loadingHtml;
            
            generateAllComparisons(actual, forecast, prevYear);
        }

        function generateAllComparisons(actual, forecast, prevYear) {
            const match = actual.match(/S&OP(\\d+)/i);
            if (!match) {
                document.getElementById('monthComparisonTableBody').innerHTML = '<tr><td colspan="12" class="text-center">Selected dataset is not an S&OP file</td></tr>';
                return;
            }
            
            const selectedSopMonth = parseInt(match[1]);
            const monthAbbr = monthNames[selectedSopMonth - 1];
            
            const monthlyActualSales = [];
            const monthlyActualGP = [];
            const monthlyActualQty = [];
            const monthlyActualCommission = [];
            const monthlyActualGM = [];
            const monthlyActualGMPercent = [];
            const monthlyActualGPPercent = [];
            const monthlyActualNotBookedSalesPercent = [];
            const monthlyActualNotBookedGMPercent = [];
            const monthlyActualNotBookedSales = [];
            const monthlyActualNotBookedGP = [];
            const monthlyActualNotBookedQty = [];
            const monthlyForecastSales = [];
            const monthlyForecastGP = [];
            const monthlyForecastQty = [];
            const monthlyForecastCommission = [];
            const monthlyForecastGM = [];
            const monthlyForecastGMPercent = [];
            const monthlyForecastGPPercent = [];
            const monthlyForecastNotBookedSalesPercent = [];
            const monthlyForecastNotBookedGMPercent = [];
            const monthlyForecastNotBookedSales = [];
            const monthlyForecastNotBookedGP = [];
            const monthlyForecastNotBookedQty = [];
            const monthlyPrevYearSales = [];
            const monthlyPrevYearGP = [];
            const monthlyPrevYearQty = [];
            const monthlyPrevYearCommission = [];
            const monthlyPrevYearGM = [];
            const monthlyPrevYearGMPercent = [];
            const monthlyPrevYearGPPercent = [];
            const monthlyPrevYearNotBookedSalesPercent = [];
            const monthlyPrevYearNotBookedGMPercent = [];
            const monthlyTargetSales = [];
            const monthlyTargetGP = [];
            const monthlyTargetQty = [];
            const monthlyTargetCommission = [];
            const monthlyTargetGM = [];
            const monthlyTargetGMPercent = [];
            const monthlyTargetGPPercent = [];
            const monthlyTargetNotBookedSalesPercent = [];
            const monthlyTargetNotBookedGMPercent = [];

            function getNotBookedPercentages(dataset, monthAbbrLoop) {
                const sopRecords = dashboardData.sop_records || [];
                const fileRecords = sopRecords.filter(record =>
                    String(record.file || '').trim() === String(dataset || '').trim() &&
                    String(record.section || '').trim() === 'Trade' &&
                    String(record.actual_month || '').trim() === monthAbbrLoop
                );
                const allSales = fileRecords.reduce((sum, record) => sum + Number(record.sales_jod || 0), 0);
                const allGm = fileRecords.reduce((sum, record) => sum + Number(record.gp_jod || 0), 0);
                const notBookedRecords = fileRecords.filter(record => String(record.status || '').trim() === 'Not Booked');
                const notBookedSales = notBookedRecords.reduce((sum, record) => sum + Number(record.sales_jod || 0), 0);
                const notBookedGm = notBookedRecords.reduce((sum, record) => sum + Number(record.gp_jod || 0), 0);
                return {
                    salesPct: allSales > 0 ? (notBookedSales / allSales) * 100 : 0,
                    gmPct: allGm > 0 ? (notBookedGm / allGm) * 100 : 0
                };
            }

            function getNotBookedValues(dataset, monthAbbrLoop) {
                const sopRecords = dashboardData.sop_records || [];
                const fileRecords = sopRecords.filter(record =>
                    String(record.file || '').trim() === String(dataset || '').trim() &&
                    String(record.actual_month || '').trim() === monthAbbrLoop &&
                    String(record.status || '').trim() === 'Not Booked'
                );
                return {
                    sales: fileRecords.filter(record => String(record.section || '').trim() === 'Trade').reduce((sum, record) => sum + Number(record.sales_jod || 0), 0),
                    qty: fileRecords.filter(record => String(record.section || '').trim() === 'Trade').reduce((sum, record) => sum + Number(record.qty_mt || 0), 0),
                    gp: fileRecords.reduce((sum, record) => sum + Number(record.gp_jod || 0), 0)
                };
            }
            
            for (let i = 1; i <= 12; i++) {
                const monthAbbrLoop = monthNames[i - 1];
                
                const actualData = getFilteredDataForMonth(actual, monthAbbrLoop, selectedSopMonth, 'actual');
                monthlyActualSales.push(actualData.sales);
                monthlyActualGP.push(actualData.gp);
                monthlyActualQty.push(actualData.qty);
                monthlyActualCommission.push(actualData.commission);
                monthlyActualGM.push(actualData.gm);
                monthlyActualGMPercent.push(actualData.gm_percent);
                monthlyActualGPPercent.push(actualData.gp_percent);
                const actualNotBooked = getNotBookedPercentages(actual, monthAbbrLoop);
                monthlyActualNotBookedSalesPercent.push(actualNotBooked.salesPct);
                monthlyActualNotBookedGMPercent.push(actualNotBooked.gmPct);
                const actualNotBookedValues = getNotBookedValues(actual, monthAbbrLoop);
                monthlyActualNotBookedSales.push(actualNotBookedValues.sales);
                monthlyActualNotBookedGP.push(actualNotBookedValues.gp);
                monthlyActualNotBookedQty.push(actualNotBookedValues.qty);
                
                let forecastData = { sales: 0, qty: 0, gp: 0, commission: 0, gm: 0, gm_percent: 0, gp_percent: 0 };
                if (forecast) {
                    forecastData = getFilteredDataForMonth(forecast, monthAbbrLoop, selectedSopMonth, 'forecast');
                }
                monthlyForecastSales.push(forecastData.sales);
                monthlyForecastGP.push(forecastData.gp);
                monthlyForecastQty.push(forecastData.qty);
                monthlyForecastCommission.push(forecastData.commission);
                monthlyForecastGM.push(forecastData.gm);
                monthlyForecastGMPercent.push(forecastData.gm_percent);
                monthlyForecastGPPercent.push(forecastData.gp_percent);
                const forecastNotBooked = getNotBookedPercentages(forecast, monthAbbrLoop);
                monthlyForecastNotBookedSalesPercent.push(forecastNotBooked.salesPct);
                monthlyForecastNotBookedGMPercent.push(forecastNotBooked.gmPct);
                const forecastNotBookedValues = getNotBookedValues(forecast, monthAbbrLoop);
                monthlyForecastNotBookedSales.push(forecastNotBookedValues.sales);
                monthlyForecastNotBookedGP.push(forecastNotBookedValues.gp);
                monthlyForecastNotBookedQty.push(forecastNotBookedValues.qty);
                
                let prevYearData = { sales: 0, qty: 0, gp: 0, commission: 0, gm: 0, gm_percent: 0, gp_percent: 0 };
                if (prevYear) {
                    prevYearData = getFilteredDataForMonth(prevYear, monthAbbrLoop, selectedSopMonth, 'prevYear');
                }
                monthlyPrevYearSales.push(prevYearData.sales);
                monthlyPrevYearGP.push(prevYearData.gp);
                monthlyPrevYearQty.push(prevYearData.qty);
                monthlyPrevYearCommission.push(prevYearData.commission);
                monthlyPrevYearGM.push(prevYearData.gm);
                monthlyPrevYearGMPercent.push(prevYearData.gm_percent);
                monthlyPrevYearGPPercent.push(prevYearData.gp_percent);
                const prevNotBooked = getNotBookedPercentages(prevYear, monthAbbrLoop);
                monthlyPrevYearNotBookedSalesPercent.push(prevNotBooked.salesPct);
                monthlyPrevYearNotBookedGMPercent.push(prevNotBooked.gmPct);
                
                const targetCommission = monthlyTargets[i]?.commission || 0;
                const targetSales = monthlyTargets[i]?.sales || 0;
                const targetGM = monthlyTargets[i]?.gp || 0;
                const targetGMPercent = targetSales > 0 ? (targetGM / targetSales) * 100 : 0;
                const targetGP = targetGM + targetCommission;
                const targetGPPercent = targetSales > 0 ? (targetGP / targetSales) * 100 : 0;
                
                monthlyTargetSales.push(targetSales);
                monthlyTargetGP.push(targetGP);
                monthlyTargetQty.push(monthlyTargets[i]?.qty || 0);
                monthlyTargetCommission.push(targetCommission);
                monthlyTargetGM.push(targetGM);
                monthlyTargetGMPercent.push(targetGMPercent);
                monthlyTargetGPPercent.push(targetGPPercent);
                monthlyTargetNotBookedSalesPercent.push(0);
                monthlyTargetNotBookedGMPercent.push(0);
            }
            
            const fullYearActualSales = monthlyActualSales.reduce((a, b) => a + b, 0);
            const fullYearActualGP = monthlyActualGP.reduce((a, b) => a + b, 0);
            const fullYearActualQty = monthlyActualQty.reduce((a, b) => a + b, 0);
            const fullYearActualCommission = monthlyActualCommission.reduce((a, b) => a + b, 0);
            const fullYearActualGM = monthlyActualGM.reduce((a, b) => a + b, 0);
            const fullYearForecastSales = monthlyForecastSales.reduce((a, b) => a + b, 0);
            const fullYearForecastGP = monthlyForecastGP.reduce((a, b) => a + b, 0);
            const fullYearForecastQty = monthlyForecastQty.reduce((a, b) => a + b, 0);
            const fullYearForecastCommission = monthlyForecastCommission.reduce((a, b) => a + b, 0);
            const fullYearForecastGM = monthlyForecastGM.reduce((a, b) => a + b, 0);
            const fullYearPrevYearSales = monthlyPrevYearSales.reduce((a, b) => a + b, 0);
            const fullYearPrevYearGP = monthlyPrevYearGP.reduce((a, b) => a + b, 0);
            const fullYearPrevYearQty = monthlyPrevYearQty.reduce((a, b) => a + b, 0);
            const fullYearPrevYearCommission = monthlyPrevYearCommission.reduce((a, b) => a + b, 0);
            const fullYearPrevYearGM = monthlyPrevYearGM.reduce((a, b) => a + b, 0);
            const fullYearTargetSales = monthlyTargetSales.reduce((a, b) => a + b, 0);
            const fullYearTargetGP = monthlyTargetGP.reduce((a, b) => a + b, 0);
            const fullYearTargetQty = monthlyTargetQty.reduce((a, b) => a + b, 0);
            const fullYearTargetCommission = monthlyTargetCommission.reduce((a, b) => a + b, 0);
            const fullYearTargetGM = monthlyTargetGM.reduce((a, b) => a + b, 0);
            
            const fullYearActualGMPercent = fullYearActualSales > 0 ? (fullYearActualGM / fullYearActualSales) * 100 : 0;
            const fullYearActualGPPercent = fullYearActualSales > 0 ? (fullYearActualGP / fullYearActualSales) * 100 : 0;
            const fullYearForecastGMPercent = fullYearForecastSales > 0 ? (fullYearForecastGM / fullYearForecastSales) * 100 : 0;
            const fullYearForecastGPPercent = fullYearForecastSales > 0 ? (fullYearForecastGP / fullYearForecastSales) * 100 : 0;
            const fullYearPrevYearGMPercent = fullYearPrevYearSales > 0 ? (fullYearPrevYearGM / fullYearPrevYearSales) * 100 : 0;
            const fullYearPrevYearGPPercent = fullYearPrevYearSales > 0 ? (fullYearPrevYearGP / fullYearPrevYearSales) * 100 : 0;
            const fullYearTargetGMPercent = fullYearTargetSales > 0 ? (fullYearTargetGM / fullYearTargetSales) * 100 : 0;
            const fullYearTargetGPPercent = fullYearTargetSales > 0 ? (fullYearTargetGP / fullYearTargetSales) * 100 : 0;
            
            const monthActualSales = monthlyActualSales[selectedSopMonth - 1];
            const monthActualGP = monthlyActualGP[selectedSopMonth - 1];
            const monthActualQty = monthlyActualQty[selectedSopMonth - 1];
            const monthActualCommission = monthlyActualCommission[selectedSopMonth - 1];
            const monthActualGM = monthlyActualGM[selectedSopMonth - 1];
            const monthActualGMPercent = monthlyActualGMPercent[selectedSopMonth - 1];
            const monthActualGPPercent = monthlyActualGPPercent[selectedSopMonth - 1];
            
            const monthForecastSales = monthlyForecastSales[selectedSopMonth - 1];
            const monthForecastGP = monthlyForecastGP[selectedSopMonth - 1];
            const monthForecastQty = monthlyForecastQty[selectedSopMonth - 1];
            const monthForecastCommission = monthlyForecastCommission[selectedSopMonth - 1];
            const monthForecastGM = monthlyForecastGM[selectedSopMonth - 1];
            const monthForecastGMPercent = monthlyForecastGMPercent[selectedSopMonth - 1];
            const monthForecastGPPercent = monthlyForecastGPPercent[selectedSopMonth - 1];
            
            const monthPrevYearSales = monthlyPrevYearSales[selectedSopMonth - 1];
            const monthPrevYearGP = monthlyPrevYearGP[selectedSopMonth - 1];
            const monthPrevYearQty = monthlyPrevYearQty[selectedSopMonth - 1];
            const monthPrevYearCommission = monthlyPrevYearCommission[selectedSopMonth - 1];
            const monthPrevYearGM = monthlyPrevYearGM[selectedSopMonth - 1];
            const monthPrevYearGMPercent = monthlyPrevYearGMPercent[selectedSopMonth - 1];
            const monthPrevYearGPPercent = monthlyPrevYearGPPercent[selectedSopMonth - 1];
            
            const monthTargetSales = monthlyTargetSales[selectedSopMonth - 1];
            const monthTargetGP = monthlyTargetGP[selectedSopMonth - 1];
            const monthTargetQty = monthlyTargetQty[selectedSopMonth - 1];
            const monthTargetCommission = monthlyTargetCommission[selectedSopMonth - 1];
            const monthTargetGM = monthlyTargetGM[selectedSopMonth - 1];
            const monthTargetGMPercent = monthlyTargetGMPercent[selectedSopMonth - 1];
            const monthTargetGPPercent = monthlyTargetGPPercent[selectedSopMonth - 1];
            
            let ytdActualSales = 0, ytdActualGP = 0, ytdActualQty = 0, ytdActualCommission = 0, ytdActualGM = 0;
            let ytdForecastSales = 0, ytdForecastGP = 0, ytdForecastQty = 0, ytdForecastCommission = 0, ytdForecastGM = 0;
            let ytdPrevYearSales = 0, ytdPrevYearGP = 0, ytdPrevYearQty = 0, ytdPrevYearCommission = 0, ytdPrevYearGM = 0;
            let ytdTargetSales = 0, ytdTargetGP = 0, ytdTargetQty = 0, ytdTargetCommission = 0, ytdTargetGM = 0;
            
            for (let i = 1; i <= selectedSopMonth; i++) {
                ytdActualSales += monthlyActualSales[i-1];
                ytdActualGP += monthlyActualGP[i-1];
                ytdActualQty += monthlyActualQty[i-1];
                ytdActualCommission += monthlyActualCommission[i-1];
                ytdActualGM += monthlyActualGM[i-1];
                
                ytdForecastSales += monthlyForecastSales[i-1];
                ytdForecastGP += monthlyForecastGP[i-1];
                ytdForecastQty += monthlyForecastQty[i-1];
                ytdForecastCommission += monthlyForecastCommission[i-1];
                ytdForecastGM += monthlyForecastGM[i-1];
                
                ytdPrevYearSales += monthlyPrevYearSales[i-1];
                ytdPrevYearGP += monthlyPrevYearGP[i-1];
                ytdPrevYearQty += monthlyPrevYearQty[i-1];
                ytdPrevYearCommission += monthlyPrevYearCommission[i-1];
                ytdPrevYearGM += monthlyPrevYearGM[i-1];
                
                ytdTargetSales += monthlyTargetSales[i-1];
                ytdTargetGP += monthlyTargetGP[i-1];
                ytdTargetQty += monthlyTargetQty[i-1];
                ytdTargetCommission += monthlyTargetCommission[i-1];
                ytdTargetGM += monthlyTargetGM[i-1];
            }
            
            const ytdActualGMPercent = ytdActualSales > 0 ? (ytdActualGM / ytdActualSales) * 100 : 0;
            const ytdActualGPPercent = ytdActualSales > 0 ? (ytdActualGP / ytdActualSales) * 100 : 0;
            const ytdForecastGMPercent = ytdForecastSales > 0 ? (ytdForecastGM / ytdForecastSales) * 100 : 0;
            const ytdForecastGPPercent = ytdForecastSales > 0 ? (ytdForecastGP / ytdForecastSales) * 100 : 0;
            const ytdPrevYearGMPercent = ytdPrevYearSales > 0 ? (ytdPrevYearGM / ytdPrevYearSales) * 100 : 0;
            const ytdPrevYearGPPercent = ytdPrevYearSales > 0 ? (ytdPrevYearGP / ytdPrevYearSales) * 100 : 0;
            const ytdTargetGMPercent = ytdTargetSales > 0 ? (ytdTargetGM / ytdTargetSales) * 100 : 0;
            const ytdTargetGPPercent = ytdTargetSales > 0 ? (ytdTargetGP / ytdTargetSales) * 100 : 0;
            
            generateExistingComparisonTables(
                actual, forecast, prevYear, selectedSopMonth, monthAbbr,
                monthActualSales, monthActualQty, monthActualGM, monthActualGMPercent, monthActualCommission, monthActualGP, monthActualGPPercent,
                monthForecastSales, monthForecastQty, monthForecastGM, monthForecastGMPercent, monthForecastCommission, monthForecastGP, monthForecastGPPercent,
                monthPrevYearSales, monthPrevYearQty, monthPrevYearGM, monthPrevYearGMPercent, monthPrevYearCommission, monthPrevYearGP, monthPrevYearGPPercent,
                monthTargetSales, monthTargetQty, monthTargetGM, monthTargetGMPercent, monthTargetCommission, monthTargetGP, monthTargetGPPercent,
                ytdActualSales, ytdActualQty, ytdActualGM, ytdActualGMPercent, ytdActualCommission, ytdActualGP, ytdActualGPPercent,
                ytdForecastSales, ytdForecastQty, ytdForecastGM, ytdForecastGMPercent, ytdForecastCommission, ytdForecastGP, ytdForecastGPPercent,
                ytdPrevYearSales, ytdPrevYearQty, ytdPrevYearGM, ytdPrevYearGMPercent, ytdPrevYearCommission, ytdPrevYearGP, ytdPrevYearGPPercent,
                ytdTargetSales, ytdTargetQty, ytdTargetGM, ytdTargetGMPercent, ytdTargetCommission, ytdTargetGP, ytdTargetGPPercent,
                fullYearActualSales, fullYearActualQty, fullYearActualGM, fullYearActualGMPercent, fullYearActualCommission, fullYearActualGP, fullYearActualGPPercent,
                fullYearForecastSales, fullYearForecastQty, fullYearForecastGM, fullYearForecastGMPercent, fullYearForecastCommission, fullYearForecastGP, fullYearForecastGPPercent,
                fullYearPrevYearSales, fullYearPrevYearQty, fullYearPrevYearGM, fullYearPrevYearGMPercent, fullYearPrevYearCommission, fullYearPrevYearGP, fullYearPrevYearGPPercent,
                fullYearTargetSales, fullYearTargetQty, fullYearTargetGM, fullYearTargetGMPercent, fullYearTargetCommission, fullYearTargetGP, fullYearTargetGPPercent
            );
            
            generateNewComparisonTables({
                sales: {
                    actual: monthlyActualSales,
                    forecast: monthlyForecastSales,
                    target: monthlyTargetSales,
                    prevYear: monthlyPrevYearSales,
                    fullYearActual: fullYearActualSales,
                    fullYearForecast: fullYearForecastSales,
                    fullYearTarget: fullYearTargetSales,
                    fullYearPrevYear: fullYearPrevYearSales
                },
                gp: {
                    actual: monthlyActualGP,
                    forecast: monthlyForecastGP,
                    target: monthlyTargetGP,
                    prevYear: monthlyPrevYearGP,
                    fullYearActual: fullYearActualGP,
                    fullYearForecast: fullYearForecastGP,
                    fullYearTarget: fullYearTargetGP,
                    fullYearPrevYear: fullYearPrevYearGP
                },
                gm: {
                    actual: monthlyActualGMPercent,
                    forecast: monthlyForecastGMPercent,
                    target: monthlyTargetGMPercent,
                    prevYear: monthlyPrevYearGMPercent,
                    fullYearActual: fullYearActualGMPercent,
                    fullYearForecast: fullYearForecastGMPercent,
                    fullYearTarget: fullYearTargetGMPercent,
                    fullYearPrevYear: fullYearPrevYearGMPercent
                },
                gpPercent: {
                    actual: monthlyActualGPPercent,
                    forecast: monthlyForecastGPPercent,
                    target: monthlyTargetGPPercent,
                    prevYear: monthlyPrevYearGPPercent,
                    fullYearActual: fullYearActualGPPercent,
                    fullYearForecast: fullYearForecastGPPercent,
                    fullYearTarget: fullYearTargetGPPercent,
                    fullYearPrevYear: fullYearPrevYearGPPercent
                },
                notBookedSalesPct: {
                    actual: monthlyActualNotBookedSalesPercent,
                    forecast: monthlyForecastNotBookedSalesPercent,
                    target: monthlyTargetNotBookedSalesPercent,
                    prevYear: monthlyPrevYearNotBookedSalesPercent,
                    fullYearActual: fullYearActualSales > 0 ? ((monthlyActualSales.reduce((s, v, idx) => s + (v * monthlyActualNotBookedSalesPercent[idx] / 100), 0)) / fullYearActualSales) * 100 : 0,
                    fullYearForecast: fullYearForecastSales > 0 ? ((monthlyForecastSales.reduce((s, v, idx) => s + (v * monthlyForecastNotBookedSalesPercent[idx] / 100), 0)) / fullYearForecastSales) * 100 : 0,
                    fullYearTarget: 0,
                    fullYearPrevYear: fullYearPrevYearSales > 0 ? ((monthlyPrevYearSales.reduce((s, v, idx) => s + (v * monthlyPrevYearNotBookedSalesPercent[idx] / 100), 0)) / fullYearPrevYearSales) * 100 : 0
                },
                notBookedGmPct: {
                    actual: monthlyActualNotBookedGMPercent,
                    forecast: monthlyForecastNotBookedGMPercent,
                    target: monthlyTargetNotBookedGMPercent,
                    prevYear: monthlyPrevYearNotBookedGMPercent,
                    fullYearActual: fullYearActualGM > 0 ? ((monthlyActualGM.reduce((s, v, idx) => s + (v * monthlyActualNotBookedGMPercent[idx] / 100), 0)) / fullYearActualGM) * 100 : 0,
                    fullYearForecast: fullYearForecastGM > 0 ? ((monthlyForecastGM.reduce((s, v, idx) => s + (v * monthlyForecastNotBookedGMPercent[idx] / 100), 0)) / fullYearForecastGM) * 100 : 0,
                    fullYearTarget: 0,
                    fullYearPrevYear: fullYearPrevYearGM > 0 ? ((monthlyPrevYearGM.reduce((s, v, idx) => s + (v * monthlyPrevYearNotBookedGMPercent[idx] / 100), 0)) / fullYearPrevYearGM) * 100 : 0
                },
                qty: {
                    actual: monthlyActualQty,
                    forecast: monthlyForecastQty,
                    target: monthlyTargetQty,
                    prevYear: monthlyPrevYearQty,
                    fullYearActual: fullYearActualQty,
                    fullYearForecast: fullYearForecastQty,
                    fullYearTarget: fullYearTargetQty,
                    fullYearPrevYear: fullYearPrevYearQty
                },
                notBookedAbsolute: {
                    sales: monthlyActualNotBookedSales,
                    gp: monthlyActualNotBookedGP,
                    qty: monthlyActualNotBookedQty
                },
                forecastNotBookedAbsolute: {
                    sales: monthlyForecastNotBookedSales,
                    gp: monthlyForecastNotBookedGP,
                    qty: monthlyForecastNotBookedQty
                }
            }, actual, forecast, prevYear, selectedSopMonth);
            
            generateStatusCharts(actual, forecast, prevYear, selectedSopMonth);
            generateNewDetailedTables(actual, forecast, prevYear, selectedSopMonth, monthlyActualSales, monthlyActualGM, monthlyActualCommission, monthlyForecastSales, monthlyPrevYearSales, monthlyTargetSales);
            latestSopAnalysis = {
                actual,
                forecast,
                prevYear,
                selectedSopMonth,
                monthAbbr,
                monthActualSales,
                monthForecastSales,
                monthPrevYearSales,
                monthTargetSales,
                monthActualGP,
                monthForecastGP,
                monthPrevYearGP,
                monthTargetGP,
                monthActualGMPercent,
                monthForecastGMPercent,
                monthPrevYearGMPercent,
                monthTargetGMPercent,
                ytdActualSales,
                ytdForecastSales,
                ytdPrevYearSales,
                ytdTargetSales,
                ytdActualGP,
                ytdForecastGP,
                ytdPrevYearGP,
                ytdTargetGP,
                ytdActualGMPercent,
                ytdPrevYearGMPercent,
                ytdTargetGMPercent,
                fullYearActualSales,
                fullYearForecastSales,
                fullYearPrevYearSales,
                fullYearTargetSales,
                fullYearActualGP,
                fullYearForecastGP,
                fullYearPrevYearGP,
                fullYearTargetGP,
                fullYearActualGMPercent,
                fullYearPrevYearGMPercent,
                fullYearTargetGMPercent,
                monthlyActualSales,
                monthlyForecastSales,
                monthlyPrevYearSales,
                monthlyTargetSales,
                monthlyActualGP,
                monthlyForecastGP,
                monthlyPrevYearGP,
                monthlyTargetGP,
                monthlyActualGMPercent,
                monthlyForecastGMPercent,
                monthlyPrevYearGMPercent,
                monthlyTargetGMPercent
            };
            buildSopTableAnalyses();
        }
        
        function generateNewDetailedTables(actual, forecast, prevYear, selectedSopMonth, monthlyActualSales, monthlyActualGM, monthlyActualCommission, monthlyForecastSales, monthlyPrevYearSales, monthlyTargetSales) {
            const sopRecords = dashboardData.sop_records || [];
            const fullYearTargetSales = monthlyTargetSales.reduce((a, b) => a + b, 0);
            const fullYearTargetGM = Object.values(monthlyTargets).reduce((sum, item) => sum + (item.gp || 0), 0);
            const fullYearTargetCommission = Object.values(monthlyTargets).reduce((sum, item) => sum + (item.commission || 0), 0);
            function textValue(value) { return value === undefined || value === null ? '' : String(value).trim(); }
            function escapeHtml(value) { return textValue(value).replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;').replace(/"/g, '&quot;').replace(/'/g, '&#39;'); }
            function clientDisplayName(value) {
                const words = textValue(value).split(/\\s+/).filter(Boolean);
                if (words.length <= 3) return words.join(' ');
                return `${words[0]} ${words[1]} ${words[words.length - 1]}`;
            }
            function getDatasetRecords(dataset, section, allowedStatuses = null) {
                return sopRecords.filter(record => textValue(record.file) === textValue(dataset) && (!section || textValue(record.section) === section) && (!allowedStatuses || allowedStatuses.includes(textValue(record.status))));
            }
            function aggregateRecords(records, groupField) {
                const aggregated = {};
                records.forEach(record => {
                    const key = textValue(record[groupField]) || 'Unknown';
                    if (!aggregated[key]) aggregated[key] = { sales: 0, gp: 0, qty: 0 };
                    aggregated[key].sales += Number(record.sales_jod || 0);
                    aggregated[key].gp += Number(record.gp_jod || 0);
                    aggregated[key].qty += Number(record.qty_mt || 0);
                });
                return aggregated;
            }
            function unionKeys(...maps) {
                const keys = new Set();
                maps.forEach(map => Object.keys(map || {}).forEach(key => keys.add(key)));
                return Array.from(keys);
            }
            function renderNoData(bodyId, colspan, message) {
                const body = document.getElementById(bodyId);
                if (body) body.innerHTML = `<tr><td colspan="${colspan}" class="text-center">${message}</td></tr>`;
            }
            function buildCombinedEntityTable(groupField, bodyId) {
                const actualMap = aggregateRecords(getDatasetRecords(actual, 'Trade'), groupField);
                const forecastMap = aggregateRecords(getDatasetRecords(forecast, 'Trade'), groupField);
                const prevMap = aggregateRecords(getDatasetRecords(prevYear, 'Trade'), groupField);
                const actualNotBookedMap = aggregateRecords(getDatasetRecords(actual, 'Trade', ['Not Booked']), groupField);
                const keys = unionKeys(actualMap, forecastMap, prevMap, actualNotBookedMap);
                if (!keys.length) return renderNoData(bodyId, 13, 'No data available');
                const totalActualSales = Object.values(actualMap).reduce((s, item) => s + item.sales, 0);
                let html = '';
                let totals = { sales: 0, gp: 0, notBookedSales: 0, notBookedGp: 0, prevSales: 0, prevGp: 0, forecastSales: 0, forecastGp: 0, targetSales: 0, targetGp: 0 };
                keys.sort((a, b) => (actualMap[b]?.sales || 0) - (actualMap[a]?.sales || 0)).forEach(key => {
                    const actualSales = actualMap[key]?.sales || 0;
                    const actualGm = actualMap[key]?.gp || 0;
                    const prevSales = prevMap[key]?.sales || 0;
                    const prevGm = prevMap[key]?.gp || 0;
                    const forecastSales = forecastMap[key]?.sales || 0;
                    const forecastGm = forecastMap[key]?.gp || 0;
                    const contribution = totalActualSales > 0 ? (actualSales / totalActualSales) * 100 : 0;
                    const targetSales = (contribution / 100) * fullYearTargetSales;
                    const targetGm = (contribution / 100) * fullYearTargetGM;
                    const notBookedSales = actualNotBookedMap[key]?.sales || 0;
                    const notBookedGm = actualNotBookedMap[key]?.gp || 0;
                    totals.sales += actualSales; totals.gp += actualGm; totals.notBookedSales += notBookedSales; totals.notBookedGp += notBookedGm; totals.prevSales += prevSales; totals.prevGp += prevGm; totals.forecastSales += forecastSales; totals.forecastGp += forecastGm; totals.targetSales += targetSales; totals.targetGp += targetGm;
                    const displayName = groupField === 'client' ? clientDisplayName(key) : key;
                    html += `<tr><td class="text-left">${escapeHtml(displayName)}</td><td class="text-right">${formatNumber(actualSales)}</td><td class="text-right">${formatNumber(actualGm)}</td><td class="text-right">${formatPercent(actualSales > 0 ? (actualGm / actualSales) * 100 : 0)}</td><td class="text-right">${formatPercent(contribution)}</td><td class="text-right">${formatPercent(actualSales > 0 ? (notBookedSales / actualSales) * 100 : 0)}</td><td class="text-right">${formatPercent(actualGm > 0 ? (notBookedGm / actualGm) * 100 : 0)}</td><td class="text-right">${getVarianceRateCellHtml(actualSales - prevSales, prevSales)}</td><td class="text-right">${getVarianceRateCellHtml(actualGm - prevGm, prevGm)}</td><td class="text-right">${getVarianceRateCellHtml(actualSales - forecastSales, forecastSales)}</td><td class="text-right">${getVarianceRateCellHtml(actualGm - forecastGm, forecastGm)}</td><td class="text-right">${getVarianceRateCellHtml(actualSales - targetSales, targetSales)}</td><td class="text-right">${getVarianceRateCellHtml(actualGm - targetGm, targetGm)}</td></tr>`;
                });
                document.getElementById(bodyId).innerHTML = html;
                addGrandTotalRow(bodyId, [totals.sales, totals.gp, totals.sales > 0 ? (totals.gp / totals.sales) * 100 : 0, 100, totals.sales > 0 ? (totals.notBookedSales / totals.sales) * 100 : 0, totals.gp > 0 ? (totals.notBookedGp / totals.gp) * 100 : 0, getVariancePctValue(totals.sales - totals.prevSales, totals.prevSales), getVariancePctValue(totals.gp - totals.prevGp, totals.prevGp), getVariancePctValue(totals.sales - totals.forecastSales, totals.forecastSales), getVariancePctValue(totals.gp - totals.forecastGp, totals.forecastGp), getVariancePctValue(totals.sales - totals.targetSales, totals.targetSales), getVariancePctValue(totals.gp - totals.targetGp, totals.targetGp)], [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]);
            }
            function buildCountryCombinedTable() {
                const actualMap = aggregateRecords(getDatasetRecords(actual, 'Trade'), 'country');
                const forecastMap = aggregateRecords(getDatasetRecords(forecast, 'Trade'), 'country');
                const prevMap = aggregateRecords(getDatasetRecords(prevYear, 'Trade'), 'country');
                const keys = unionKeys(actualMap, forecastMap, prevMap);
                if (!keys.length) return renderNoData('salesPerCountryBody', 11, 'No country data available');
                const totalActualSales = Object.values(actualMap).reduce((s, item) => s + item.sales, 0);
                let html = '';
                let totals = { prevSales: 0, prevGp: 0, sales: 0, gp: 0, forecastSales: 0, forecastGp: 0, targetSales: 0, targetGp: 0 };
                keys.sort((a, b) => (actualMap[b]?.sales || 0) - (actualMap[a]?.sales || 0)).forEach(key => {
                    const prevSales = prevMap[key]?.sales || 0;
                    const prevGp = prevMap[key]?.gp || 0;
                    const actualSales = actualMap[key]?.sales || 0;
                    const actualGp = actualMap[key]?.gp || 0;
                    const forecastSales = forecastMap[key]?.sales || 0;
                    const forecastGp = forecastMap[key]?.gp || 0;
                    const contribution = totalActualSales > 0 ? (actualSales / totalActualSales) * 100 : 0;
                    const targetSales = (contribution / 100) * fullYearTargetSales;
                    const targetGp = (contribution / 100) * fullYearTargetGM;
                    totals.prevSales += prevSales; totals.prevGp += prevGp; totals.sales += actualSales; totals.gp += actualGp; totals.forecastSales += forecastSales; totals.forecastGp += forecastGp; totals.targetSales += targetSales; totals.targetGp += targetGp;
                    html += `<tr><td class="text-left">${escapeHtml(key)}</td><td class="text-right">${formatNumber(prevSales)}</td><td class="text-right">${formatNumber(prevGp)}</td><td class="text-right">${formatNumber(actualSales)}</td><td class="text-right">${formatNumber(actualGp)}</td><td class="text-right">${formatNumber(forecastSales)}</td><td class="text-right">${formatNumber(forecastGp)}</td><td class="text-right">${formatPercent(actualSales > 0 ? (actualGp / actualSales) * 100 : 0)}</td><td class="text-right">${formatPercent(contribution)}</td><td class="text-right">${formatNumber(targetSales)}</td><td class="text-right">${formatNumber(targetGp)}</td></tr>`;
                });
                document.getElementById('salesPerCountryBody').innerHTML = html;
                addGrandTotalRow('salesPerCountryBody', [totals.prevSales, totals.prevGp, totals.sales, totals.gp, totals.forecastSales, totals.forecastGp, totals.sales > 0 ? (totals.gp / totals.sales) * 100 : 0, 100, totals.targetSales, totals.targetGp], [6, 7]);
            }
            function buildSuppliersTable() {
                const actualTrade = aggregateRecords(getDatasetRecords(actual, 'Trade'), 'supplier');
                const actualBroker = aggregateRecords(getDatasetRecords(actual, 'Broker'), 'supplier');
                const keys = unionKeys(actualTrade, actualBroker);
                if (!keys.length) return renderNoData('suppliersBody', 7, 'No supplier data available');
                const totalTradeSales = Object.values(actualTrade).reduce((sum, item) => sum + item.sales, 0);
                const totalTradeGm = Object.values(actualTrade).reduce((sum, item) => sum + item.gp, 0);
                const totalBrokerGp = Object.values(actualBroker).reduce((sum, item) => sum + item.gp, 0);
                let html = '';
                keys.sort((a, b) => (actualTrade[b]?.sales || 0) - (actualTrade[a]?.sales || 0)).forEach(key => {
                    const sales = actualTrade[key]?.sales || 0;
                    const gm = actualTrade[key]?.gp || 0;
                    const brokerGp = actualBroker[key]?.gp || 0;
                    html += `<tr><td class="text-left">${escapeHtml(key)}</td><td class="text-right">${formatNumber(sales)}</td><td class="text-right">${formatPercent(totalTradeSales > 0 ? (sales / totalTradeSales) * 100 : 0)}</td><td class="text-right">${formatNumber(gm)}</td><td class="text-right">${formatPercent(totalTradeGm > 0 ? (gm / totalTradeGm) * 100 : 0)}</td><td class="text-right">${formatNumber(brokerGp)}</td><td class="text-right">${formatPercent(totalBrokerGp > 0 ? (brokerGp / totalBrokerGp) * 100 : 0)}</td></tr>`;
                });
                document.getElementById('suppliersBody').innerHTML = html;
                addGrandTotalRow('suppliersBody', [totalTradeSales, 100, totalTradeGm, 100, totalBrokerGp, 100], [1, 3, 5]);
            }
            function buildBrokerTable(groupField, bodyId) {
                const actualMap = aggregateRecords(getDatasetRecords(actual, 'Broker'), groupField);
                const forecastMap = aggregateRecords(getDatasetRecords(forecast, 'Broker'), groupField);
                const prevMap = aggregateRecords(getDatasetRecords(prevYear, 'Broker'), groupField);
                const notBookedMap = aggregateRecords(getDatasetRecords(actual, 'Broker', ['Not Booked']), groupField);
                const keys = unionKeys(actualMap, forecastMap, prevMap, notBookedMap);
                if (!keys.length) return renderNoData(bodyId, 7, 'No broker data available');
                const totalActualBroker = Object.values(actualMap).reduce((sum, item) => sum + item.gp, 0);
                let html = '';
                let totals = { actual: 0, notBooked: 0, prev: 0, forecast: 0, target: 0 };
                keys.sort((a, b) => (actualMap[b]?.gp || 0) - (actualMap[a]?.gp || 0)).forEach(key => {
                    const actualBroker = actualMap[key]?.gp || 0;
                    const forecastBroker = forecastMap[key]?.gp || 0;
                    const prevBroker = prevMap[key]?.gp || 0;
                    const notBookedBroker = notBookedMap[key]?.gp || 0;
                    const contribution = totalActualBroker > 0 ? (actualBroker / totalActualBroker) * 100 : 0;
                    const targetBroker = (contribution / 100) * fullYearTargetCommission;
                    totals.actual += actualBroker; totals.notBooked += notBookedBroker; totals.prev += prevBroker; totals.forecast += forecastBroker; totals.target += targetBroker;
                    html += `<tr><td class="text-left">${escapeHtml(key)}</td><td class="text-right">${formatNumber(actualBroker)}</td><td class="text-right">${formatPercent(contribution)}</td><td class="text-right">${formatPercent(actualBroker > 0 ? (notBookedBroker / actualBroker) * 100 : 0)}</td><td class="text-right">${getVarianceRateCellHtml(actualBroker - prevBroker, prevBroker)}</td><td class="text-right">${getVarianceRateCellHtml(actualBroker - forecastBroker, forecastBroker)}</td><td class="text-right">${getVarianceRateCellHtml(actualBroker - targetBroker, targetBroker)}</td></tr>`;
                });
                document.getElementById(bodyId).innerHTML = html;
                addGrandTotalRow(bodyId, [totals.actual, 100, totals.actual > 0 ? (totals.notBooked / totals.actual) * 100 : 0, getVariancePctValue(totals.actual - totals.prev, totals.prev), getVariancePctValue(totals.actual - totals.forecast, totals.forecast), getVariancePctValue(totals.actual - totals.target, totals.target)], [1, 2, 3, 4, 5]);
            }
            buildCombinedEntityTable('client', 'salesPerClientBody');
            buildCombinedEntityTable('category', 'salesPerCategoryBody');
            buildCountryCombinedTable();
            buildSuppliersTable();
            buildBrokerTable('client', 'brokerPerClientBody');
            buildBrokerTable('category', 'brokerPerCategoryBody');
            const gmBodies = ['gmPerClientBody', 'gmPerCategoryBody', 'gmPerCountryBody'];
            gmBodies.forEach(id => { const el = document.getElementById(id); if (el) el.innerHTML = ''; });
        }

        function generateExistingComparisonTables(
            actual, forecast, prevYear, selectedSopMonth, monthAbbr,
            monthActualSales, monthActualQty, monthActualGM, monthActualGMPercent, monthActualCommission, monthActualGP, monthActualGPPercent,
            monthForecastSales, monthForecastQty, monthForecastGM, monthForecastGMPercent, monthForecastCommission, monthForecastGP, monthForecastGPPercent,
            monthPrevYearSales, monthPrevYearQty, monthPrevYearGM, monthPrevYearGMPercent, monthPrevYearCommission, monthPrevYearGP, monthPrevYearGPPercent,
            monthTargetSales, monthTargetQty, monthTargetGM, monthTargetGMPercent, monthTargetCommission, monthTargetGP, monthTargetGPPercent,
            ytdActualSales, ytdActualQty, ytdActualGM, ytdActualGMPercent, ytdActualCommission, ytdActualGP, ytdActualGPPercent,
            ytdForecastSales, ytdForecastQty, ytdForecastGM, ytdForecastGMPercent, ytdForecastCommission, ytdForecastGP, ytdForecastGPPercent,
            ytdPrevYearSales, ytdPrevYearQty, ytdPrevYearGM, ytdPrevYearGMPercent, ytdPrevYearCommission, ytdPrevYearGP, ytdPrevYearGPPercent,
            ytdTargetSales, ytdTargetQty, ytdTargetGM, ytdTargetGMPercent, ytdTargetCommission, ytdTargetGP, ytdTargetGPPercent,
            fullYearActualSales, fullYearActualQty, fullYearActualGM, fullYearActualGMPercent, fullYearActualCommission, fullYearActualGP, fullYearActualGPPercent,
            fullYearForecastSales, fullYearForecastQty, fullYearForecastGM, fullYearForecastGMPercent, fullYearForecastCommission, fullYearForecastGP, fullYearForecastGPPercent,
            fullYearPrevYearSales, fullYearPrevYearQty, fullYearPrevYearGM, fullYearPrevYearGMPercent, fullYearPrevYearCommission, fullYearPrevYearGP, fullYearPrevYearGPPercent,
            fullYearTargetSales, fullYearTargetQty, fullYearTargetGM, fullYearTargetGMPercent, fullYearTargetCommission, fullYearTargetGP, fullYearTargetGPPercent
        ) {
            const monthActualData = { sales: monthActualSales, gm: monthActualGM, commission: monthActualCommission, gp: monthActualGP };
            const monthForecastData = { sales: monthForecastSales, gm: monthForecastGM, commission: monthForecastCommission, gp: monthForecastGP };
            const monthPrevYearData = { sales: monthPrevYearSales, gm: monthPrevYearGM, commission: monthPrevYearCommission, gp: monthPrevYearGP };
            const monthTargetData = { sales: monthTargetSales, gm: monthTargetGM, commission: monthTargetCommission, gp: monthTargetGP };

            const ytdActualData = { sales: ytdActualSales, gm: ytdActualGM, commission: ytdActualCommission, gp: ytdActualGP };
            const ytdForecastData = { sales: ytdForecastSales, gm: ytdForecastGM, commission: ytdForecastCommission, gp: ytdForecastGP };
            const ytdPrevYearData = { sales: ytdPrevYearSales, gm: ytdPrevYearGM, commission: ytdPrevYearCommission, gp: ytdPrevYearGP };
            const ytdTargetData = { sales: ytdTargetSales, gm: ytdTargetGM, commission: ytdTargetCommission, gp: ytdTargetGP };

            const fullYearActualData = { sales: fullYearActualSales, gm: fullYearActualGM, commission: fullYearActualCommission, gp: fullYearActualGP };
            const fullYearForecastData = { sales: fullYearForecastSales, gm: fullYearForecastGM, commission: fullYearForecastCommission, gp: fullYearForecastGP };
            const fullYearPrevYearData = { sales: fullYearPrevYearSales, gm: fullYearPrevYearGM, commission: fullYearPrevYearCommission, gp: fullYearPrevYearGP };
            const fullYearTargetData = { sales: fullYearTargetSales, gm: fullYearTargetGM, commission: fullYearTargetCommission, gp: fullYearTargetGP };

            generateComparisonTable('monthComparisonTableBody', monthActualData, monthForecastData, monthTargetData, monthPrevYearData);
            generateComparisonTable('ytdComparisonTableBody', ytdActualData, ytdForecastData, ytdTargetData, ytdPrevYearData);
            generateComparisonTable('fullYearComparisonTableBody', fullYearActualData, fullYearForecastData, fullYearTargetData, fullYearPrevYearData);
        }
        
        function generateNewComparisonTables(data, actualLabel, forecastLabel, prevLabel, selectedSopMonth) {
            const salesBody = document.getElementById('salesComparisonBody');
            const gpBody = document.getElementById('gpComparisonBody');
            const qtyBody = document.getElementById('qtyComparisonBody');
            if (!salesBody || !gpBody || !qtyBody) return;

            function renderMonthlyCells(values, notBookedValues = null) {
                return values.slice(0, 12).map((value, index) => {
                    const showNotBooked = notBookedValues && index + 1 > selectedSopMonth && Number(notBookedValues[index] || 0) !== 0;
                    const notBookedHtml = showNotBooked ? `<span class="not-booked-note">Not Booked (${formatCompactNumber(notBookedValues[index])})</span>` : '';
                    return `<td class="text-right">${formatNumber(value)}${notBookedHtml}</td>`;
                }).join('');
            }

            function renderVarianceCells(actualValues, compareValues) {
                return actualValues.slice(0, 12).map((value, index) => `<td class="text-right">${getVarianceRateCellHtml(value - compareValues[index], compareValues[index])}</td>`).join('');
            }

            function renderFullYearValue(value) {
                return `<td class="text-right full-year-col">${formatNumber(value)}</td>`;
            }

            function renderFullYearVariance(actualValue, compareValue) {
                return `<td class="text-right full-year-col">${getVarianceRateCellHtml(actualValue - compareValue, compareValue)}</td>`;
            }

            function buildMetricRows(metricData, cssClass, actualNotBookedValues = null, forecastNotBookedValues = null) {
                const prevDisplay = `${prevLabel || '2024'} Actual`;
                const forecastDisplay = forecastLabel || 'S&OP8';
                const actualDisplay = actualLabel || 'S&OP9';
                const targetDisplay = 'Target';
                return `
                    <tr class="${cssClass}">
                        <td class="font-bold">${prevDisplay}</td>
                        ${renderMonthlyCells(metricData.prevYear)}
                        ${renderFullYearValue(metricData.fullYearPrevYear)}
                    </tr>
                    <tr class="${cssClass}">
                        <td class="font-bold">${forecastDisplay}</td>
                        ${renderMonthlyCells(metricData.forecast, forecastNotBookedValues)}
                        ${renderFullYearValue(metricData.fullYearForecast)}
                    </tr>
                    <tr class="${cssClass}">
                        <td class="font-bold">${actualDisplay}</td>
                        ${renderMonthlyCells(metricData.actual, actualNotBookedValues)}
                        ${renderFullYearValue(metricData.fullYearActual)}
                    </tr>
                    <tr class="${cssClass}">
                        <td class="font-bold">${targetDisplay}</td>
                        ${renderMonthlyCells(metricData.target)}
                        ${renderFullYearValue(metricData.fullYearTarget)}
                    </tr>
                    <tr class="variance-row">
                        <td class="font-bold">${actualDisplay} vs ${prevDisplay}</td>
                        ${renderVarianceCells(metricData.actual, metricData.prevYear)}
                        ${renderFullYearVariance(metricData.fullYearActual, metricData.fullYearPrevYear)}
                    </tr>
                    <tr class="variance-row">
                        <td class="font-bold">${actualDisplay} vs ${forecastDisplay}</td>
                        ${renderVarianceCells(metricData.actual, metricData.forecast)}
                        ${renderFullYearVariance(metricData.fullYearActual, metricData.fullYearForecast)}
                    </tr>
                    <tr class="variance-row">
                        <td class="font-bold">${actualDisplay} vs ${targetDisplay}</td>
                        ${renderVarianceCells(metricData.actual, metricData.target)}
                        ${renderFullYearVariance(metricData.fullYearActual, metricData.fullYearTarget)}
                    </tr>
                `;
            }

            salesBody.innerHTML = buildMetricRows(data.sales, 'metric-sales', data.notBookedAbsolute?.sales || null, null);
            gpBody.innerHTML = buildMetricRows(data.gp, 'metric-gp', data.notBookedAbsolute?.gp || null, null);
            qtyBody.innerHTML = buildMetricRows(data.qty, 'metric-qty', data.notBookedAbsolute?.qty || null, null);
        }
        
        function generateComparisonTable(tableBodyId, actualData, forecastData, targetData, prevYearData) {
            const rows = [
                { name: 'Sales', actual: actualData.sales, forecast: forecastData.sales, target: targetData.sales, prev: prevYearData.sales, rowClass: 'subtotal-row' },
                { name: 'COGS', actual: actualData.sales - actualData.gm, forecast: forecastData.sales - forecastData.gm, target: targetData.sales - targetData.gm, prev: prevYearData.sales - prevYearData.gm },
                { name: 'Trade Gross Margin', actual: actualData.gm, forecast: forecastData.gm, target: targetData.gm, prev: prevYearData.gm, rowClass: 'subtotal-row' },
                { name: 'GM%', actual: actualData.sales > 0 ? (actualData.gm / actualData.sales) * 100 : 0, forecast: forecastData.sales > 0 ? (forecastData.gm / forecastData.sales) * 100 : 0, target: targetData.sales > 0 ? (targetData.gm / targetData.sales) * 100 : 0, prev: prevYearData.sales > 0 ? (prevYearData.gm / prevYearData.sales) * 100 : 0, isPercent: true, rowClass: 'percentage-row' },
                { name: 'Broker GP', actual: actualData.commission, forecast: forecastData.commission, target: targetData.commission, prev: prevYearData.commission },
                { name: 'Total Gross Profit', actual: actualData.gp, forecast: forecastData.gp, target: targetData.gp, prev: prevYearData.gp, rowClass: 'subtotal-row' },
                { name: 'GP%', actual: actualData.sales > 0 ? (actualData.gp / actualData.sales) * 100 : 0, forecast: forecastData.sales > 0 ? (forecastData.gp / forecastData.sales) * 100 : 0, target: targetData.sales > 0 ? (targetData.gp / targetData.sales) * 100 : 0, prev: prevYearData.sales > 0 ? (prevYearData.gp / prevYearData.sales) * 100 : 0, isPercent: true, rowClass: 'percentage-row' }
            ];
            let html = '';
            rows.forEach(row => {
                html += `<tr class="${row.rowClass || ''}"><td class="font-bold">${row.name}</td><td class="text-right">${row.isPercent ? formatPercent(row.actual) : formatNumber(row.actual)}</td><td class="text-right">${row.isPercent ? formatPercent(row.forecast) : formatNumber(row.forecast)}</td><td class="text-right">${row.isPercent ? formatPercent(row.target) : formatNumber(row.target)}</td><td class="text-right">${row.isPercent ? formatPercent(row.prev) : formatNumber(row.prev)}</td><td class="text-right">${row.isPercent ? getVariancePercentCellHtml(row.actual - row.forecast, row.forecast) : getVarianceRateCellHtml(row.actual - row.forecast, row.forecast)}</td><td class="text-right">${row.isPercent ? getVariancePercentCellHtml(row.actual - row.target, row.target) : getVarianceRateCellHtml(row.actual - row.target, row.target)}</td><td class="text-right">${row.isPercent ? getVariancePercentCellHtml(row.actual - row.prev, row.prev) : getVarianceRateCellHtml(row.actual - row.prev, row.prev)}</td></tr>`;
            });
            document.getElementById(tableBodyId).innerHTML = html;
        }

        function buildSopTableAnalyses() {
            const ctx = latestSopAnalysis || {};
            if (!ctx.actual) return;
            const peakIndex = ctx.monthlyActualSales.indexOf(Math.max(...ctx.monthlyActualSales));
            const secondPeakIndex = [...ctx.monthlyActualSales].map((v, i) => ({ v, i })).sort((a, b) => b.v - a.v)[1]?.i ?? peakIndex;
            const lowIndex = ctx.monthlyActualSales.indexOf(Math.min(...ctx.monthlyActualSales));
            const topClient = document.querySelector('#salesPerClientBody tr td')?.textContent || 'N/A';
            const topCategory = document.querySelector('#salesPerCategoryBody tr td')?.textContent || 'N/A';
            const topSupplier = document.querySelector('#suppliersBody tr td')?.textContent || 'N/A';
            const topBrokerClient = document.querySelector('#brokerPerClientBody tr td')?.textContent || 'N/A';
            const topBrokerCategory = document.querySelector('#brokerPerCategoryBody tr td')?.textContent || 'N/A';

            latestSopAnalysis.monthAnalysis = [
                `Sales reached ${formatNumber(ctx.monthActualSales)} against target ${formatNumber(ctx.monthTargetSales)} which is ${formatPercent(ctx.monthTargetSales > 0 ? (ctx.monthActualSales / ctx.monthTargetSales) * 100 : 0)} of target.`,
                `Gross profit closed at ${formatNumber(ctx.monthActualGP)} with GP% ${formatPercent(ctx.monthActualSales > 0 ? (ctx.monthActualGP / ctx.monthActualSales) * 100 : 0)} and variance vs ${ctx.forecast || 'forecast'} of ${formatPercent(getVariancePctValue(ctx.monthActualGP - ctx.monthForecastGP, ctx.monthForecastGP))}.`,
                `Year over year sales moved ${formatPercent(getVariancePctValue(ctx.monthActualSales - ctx.monthPrevYearSales, ctx.monthPrevYearSales))} while GM% changed by ${(ctx.monthActualGMPercent - ctx.monthPrevYearGMPercent).toFixed(1)}pp.`,
                `The main visible drivers are top category ${topCategory} and top client ${topClient}, so those rows should anchor the monthly review.`
            ];
            latestSopAnalysis.ytdAnalysis = [
                `YTD sales are ${formatNumber(ctx.ytdActualSales)} which is ${formatPercent(ctx.ytdTargetSales > 0 ? (ctx.ytdActualSales / ctx.ytdTargetSales) * 100 : 0)} of target and ${formatPercent(getVariancePctValue(ctx.ytdActualSales - ctx.ytdPrevYearSales, ctx.ytdPrevYearSales))} versus ${ctx.prevYear || 'previous year'}.`,
                `YTD gross profit is ${formatNumber(ctx.ytdActualGP)} with GP% ${formatPercent(ctx.ytdActualSales > 0 ? (ctx.ytdActualGP / ctx.ytdActualSales) * 100 : 0)}.`,
                `Margin is ${(ctx.ytdActualGMPercent - ctx.ytdTargetGMPercent).toFixed(1)}pp versus target and ${(ctx.ytdActualGMPercent - ctx.ytdPrevYearGMPercent).toFixed(1)}pp versus ${ctx.prevYear || 'previous year'}.`,
                `Use the detailed client and category tables to close the remaining YTD gap in the red rows first.`
            ];
            latestSopAnalysis.fullYearAnalysis = [
                `Full-year sales achievement stands at ${formatPercent(ctx.fullYearTargetSales > 0 ? (ctx.fullYearActualSales / ctx.fullYearTargetSales) * 100 : 0)} with a sales variance vs target of ${formatPercent(getVariancePctValue(ctx.fullYearActualSales - ctx.fullYearTargetSales, ctx.fullYearTargetSales))}.`,
                `Total gross profit is ${formatNumber(ctx.fullYearActualGP)} and GP% is ${formatPercent(ctx.fullYearActualSales > 0 ? (ctx.fullYearActualGP / ctx.fullYearActualSales) * 100 : 0)}.`,
                `Gross profit moved ${formatPercent(getVariancePctValue(ctx.fullYearActualGP - ctx.fullYearPrevYearGP, ctx.fullYearPrevYearGP))} versus ${ctx.prevYear || 'previous year'} and ${formatPercent(getVariancePctValue(ctx.fullYearActualGP - ctx.fullYearForecastGP, ctx.fullYearForecastGP))} versus ${ctx.forecast || 'forecast'}.`,
                `Broker contribution should be reviewed through the broker tables, where the leading broker client is ${topBrokerClient}.`
            ];
            latestSopAnalysis.combinedAnalysis = [
                `Highest sales month is ${monthNames[peakIndex]} at ${formatNumber(ctx.monthlyActualSales[peakIndex])}; second highest is ${monthNames[secondPeakIndex]} at ${formatNumber(ctx.monthlyActualSales[secondPeakIndex])}.`,
                `Lowest sales month is ${monthNames[lowIndex]} at ${formatNumber(ctx.monthlyActualSales[lowIndex])}, showing the seasonal trough.`,
                `GM% ranges from ${formatPercent(Math.min(...ctx.monthlyActualGMPercent))} to ${formatPercent(Math.max(...ctx.monthlyActualGMPercent))} through the year.`,
                `The variance rows in this table now show only percentages, so green cells highlight months beating the benchmark and red cells show the biggest misses.`
            ];
            latestSopAnalysis.clientAnalysis = [
                `Top client by current actual sales is ${topClient}.`,
                `Clients with high Not Booked Sales % or Not Booked GM % are the clearest short-term risk to closing the cycle.`,
                `The red versus columns show which clients are behind ${ctx.forecast || 'forecast'}, ${ctx.prevYear || 'previous year'}, or target in percentage terms.`,
                `Prioritize commercial follow-up on the red client rows before pushing new volume into already healthy green rows.`
            ];
            latestSopAnalysis.categoryAnalysis = [
                `Top category by current actual sales is ${topCategory}.`,
                `Not Booked Sales % and Not Booked GM % identify categories where pipeline risk is highest.`,
                `The versus columns now show percentage change only, making it easier to compare category performance regardless of size.`,
                `Focus price and mix actions on categories with red GM variances and weak sales conversion.`
            ];
            latestSopAnalysis.supplierAnalysis = [
                `The leading supplier in the current selection is ${topSupplier}.`,
                `Sales contribution and GM contribution should be read together to spot concentration risk and supplier quality.`,
                `A supplier with high sales contribution but weak GM contribution is diluting profitability.`,
                `Broker GP contribution helps identify where supplier economics rely more on brokerage than trade margin.`
            ];
            latestSopAnalysis.brokerClientAnalysis = [
                `Top broker client is ${topBrokerClient}.`,
                `High not-booked percentages point to broker pipeline that has not yet converted into realized GM.`,
                `The versus columns show broker GM performance as percentages against ${ctx.forecast || 'forecast'}, ${ctx.prevYear || 'previous year'}, and target.`,
                `Use this table to challenge broker focus on low-conversion clients and reinforce the high-conversion accounts.`
            ];
            latestSopAnalysis.brokerCategoryAnalysis = [
                `Top broker category is ${topBrokerCategory}.`,
                `Not-booked percentages should be reviewed with brokers first because they indicate risk in promised category business.`,
                `The percentage versus columns highlight whether broker category performance is improving or deteriorating across cycles.`,
                `Move broker effort toward categories that are green on conversion and gross margin delivery.`
            ];
            latestSopAnalysis.countryAnalysis = [
                `This table compares actual, forecast, previous year, and target by country using the selected files.`,
                `Countries with high GM% but lower sales scale may be margin opportunities worth expanding.`,
                `Countries that are red versus both ${ctx.forecast || 'forecast'} and ${ctx.prevYear || 'previous year'} need immediate recovery focus.`,
                `Use the target columns to check whether geographic mix is drifting away from the annual plan.`
            ];
        }

        function generateStatusCharts(actual, forecast, prevYear, selectedSopMonth) {
            if (!actual) return;
            
            let invoicedSales = 0, invoicedGp = 0, bookedSales = 0, bookedGp = 0, notBookedSales = 0, notBookedGp = 0;
            
            for (let month = 1; month <= 12; month++) {
                const monthAbbr = monthNames[month - 1];
                const cleanName = actual.replace(/[^a-zA-Z0-9]/g, '_');
                
                if (month <= selectedSopMonth) {
                    const invoicedKey = `${cleanName}_Trade_Invoiced_${monthAbbr}`;
                    if (dashboardData.filtered_data && dashboardData.filtered_data[invoicedKey]) {
                        const data = dashboardData.filtered_data[invoicedKey];
                        const commission = getCommissionForMonth(actual, monthAbbr, selectedSopMonth);
                        invoicedSales += data.sales || 0;
                        invoicedGp += (data.gp || 0) + commission;
                    }
                } else {
                    const bookedKey = `${cleanName}_Trade_Booked_Not Booked_${monthAbbr}`;
                    if (dashboardData.filtered_data && dashboardData.filtered_data[bookedKey]) {
                        const data = dashboardData.filtered_data[bookedKey];
                        const commission = getCommissionForMonth(actual, monthAbbr, selectedSopMonth);
                        bookedSales += (data.sales || 0) * 0.7;
                        bookedGp += ((data.gp || 0) + commission) * 0.7;
                        notBookedSales += (data.sales || 0) * 0.3;
                        notBookedGp += ((data.gp || 0) + commission) * 0.3;
                    }
                }
            }
            
            if (salesStatusChart) salesStatusChart.destroy();
            if (gpStatusChart) gpStatusChart.destroy();
            
            const salesCtx = document.getElementById('salesStatusChart').getContext('2d');
            salesStatusChart = new Chart(salesCtx, {
                type: 'doughnut',
                data: { labels: ['Invoiced', 'Booked', 'Not Booked'], datasets: [{ data: [invoicedSales, bookedSales, notBookedSales], backgroundColor: ['#10b981', '#3b82f6', '#f59e0b'], borderWidth: 0 }] },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'bottom', labels: { boxWidth: 8, font: { size: 9 } } }, tooltip: { callbacks: { label: function(context) { const label = context.label || ''; const value = context.raw || 0; const total = context.dataset.data.reduce((a, b) => a + b, 0); const percentage = total > 0 ? ((value / total) * 100).toFixed(1) : '0'; return `${label}: ${formatNumber(value)} JOD (${percentage}%)`; } } } } }
            });
            
            const gpCtx = document.getElementById('gpStatusChart').getContext('2d');
            gpStatusChart = new Chart(gpCtx, {
                type: 'doughnut',
                data: { labels: ['Invoiced', 'Booked', 'Not Booked'], datasets: [{ data: [invoicedGp, bookedGp, notBookedGp], backgroundColor: ['#10b981', '#3b82f6', '#f59e0b'], borderWidth: 0 }] },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'bottom', labels: { boxWidth: 8, font: { size: 9 } } }, tooltip: { callbacks: { label: function(context) { const label = context.label || ''; const value = context.raw || 0; const total = context.dataset.data.reduce((a, b) => a + b, 0); const percentage = total > 0 ? ((value / total) * 100).toFixed(1) : '0'; return `${label}: ${formatNumber(value)} JOD (${percentage}%)`; } } } } }
            });
        }
        
        function updateDataInfo() {
            const info = dashboardData.data_info || {};
            document.getElementById('dataDateRange').innerText = info.date_range || 'N/A';
            document.getElementById('lastUpdated').innerText = info.last_updated || new Date().toLocaleString();
        }
        
        function showTab(tabName, evt = null) {
            document.querySelectorAll('.tab-pane').forEach(tab => tab.classList.remove('active'));
            document.querySelectorAll('.tab-btn').forEach(btn => btn.classList.remove('active'));
            document.getElementById(tabName + 'Tab').classList.add('active');
            if (tabName === 'tracker' && !trackerRendered) {
                try {
                    renderTrackerAnalysis();
                    trackerRendered = true;
                } catch (error) {
                    console.error('Tracker render failed:', error);
                    const trackerBody = document.getElementById('trackerCombinedBody');
                    if (trackerBody) trackerBody.innerHTML = '<tr><td colspan="13" class="text-center">Tracker data could not be loaded</td></tr>';
                }
            }
            const clickEvent = evt || window.event;
            if (clickEvent && clickEvent.target) {
                clickEvent.target.classList.add('active');
            }
        }
        
        function renderSalesAnalysis() {
            const records = getFilteredDemandRecords();
            const monthlySales = buildMonthlySeries(records, 'sales_jod');
            const monthlyGp = buildMonthlySeries(records, 'gp_jod');
            const categories = aggregateDemandRecords(records, 'category').sort((a, b) => b.sales - a.sales);
            const customers = aggregateDemandRecords(records, 'customer_name').sort((a, b) => b.sales - a.sales);
            const statuses = aggregateDemandRecords(records, 'status').sort((a, b) => b.sales - a.sales);
            const totalSales = records.reduce((sum, record) => sum + Number(record.sales_jod || 0), 0);
            const totalGp = records.reduce((sum, record) => sum + Number(record.gp_jod || 0), 0);

            if (trendChart) trendChart.destroy();
            const trendCtx = document.getElementById('trendChart').getContext('2d');
            trendChart = new Chart(trendCtx, {
                type: 'bar',
                data: { labels: monthlySales.labels, datasets: [{ label: 'Sales (LCY)', data: monthlySales.values, backgroundColor: '#2a5298' }] },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } } }
            });

            if (monthlyGpChart) monthlyGpChart.destroy();
            const gpCtx = document.getElementById('monthlyGpChart').getContext('2d');
            monthlyGpChart = new Chart(gpCtx, {
                type: 'line',
                data: { labels: monthlyGp.labels, datasets: [{ label: 'Gross Profit', data: monthlyGp.values, borderColor: '#059669', backgroundColor: 'rgba(5,150,105,0.15)', fill: true, tension: 0.35 }] },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { display: false } } }
            });

            if (categoryPieChart) categoryPieChart.destroy();
            const categoryCtx = document.getElementById('categoryPieChart').getContext('2d');
            categoryPieChart = new Chart(categoryCtx, {
                type: 'doughnut',
                data: {
                    labels: categories.slice(0, 6).map(item => item.name),
                    datasets: [{ data: categories.slice(0, 6).map(item => item.sales), backgroundColor: ['#1e3c72', '#2563eb', '#059669', '#f59e0b', '#ef4444', '#64748b'] }]
                },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'bottom', labels: { boxWidth: 8, font: { size: 9 } } } } }
            });

            if (statusMixChart) statusMixChart.destroy();
            const statusCtx = document.getElementById('statusMixChart').getContext('2d');
            statusMixChart = new Chart(statusCtx, {
                type: 'pie',
                data: {
                    labels: statuses.map(item => item.name),
                    datasets: [{ data: statuses.map(item => item.count), backgroundColor: ['#2563eb', '#10b981', '#f59e0b', '#ef4444', '#64748b'] }]
                },
                options: { responsive: true, maintainAspectRatio: false, plugins: { legend: { position: 'bottom', labels: { boxWidth: 8, font: { size: 9 } } } } }
            });

            let customerHtml = '';
            customers.slice(0, 10).forEach(item => {
                customerHtml += `<tr><td class="text-left">${item.name || 'N/A'}</td><td class="text-right">${formatNumber(item.sales)}</td><td class="text-right">${item.gm.toFixed(1)}%</td></tr>`;
            });
            document.querySelector('#topCustomersSalesTable').innerHTML = customerHtml || '<tr><td colspan="3" class="text-center">No data</td></tr>';

            let countryHtml = '<tr><td colspan="3" class="text-center">Country is not available in Demand Analysis sheet</td></tr>';
            document.querySelector('#salesByCountryTable').innerHTML = countryHtml;

            let statusHtml = '';
            statuses.forEach(item => {
                statusHtml += `<tr><td class="text-left">${item.name || 'N/A'}</td><td class="text-right">${formatNumber(item.count || 0)}</td><td class="text-right">${formatNumber(item.sales)}</td></tr>`;
            });
            document.querySelector('#statusAnalysisTable').innerHTML = statusHtml || '<tr><td colspan="3" class="text-center">No data</td></tr>';
            renderInsightsList('salesInsightsBody', [
                `Filtered sales are ${formatNumber(totalSales)} with gross profit of ${formatNumber(totalGp)} and a margin of ${formatPercent(totalSales > 0 ? (totalGp / totalSales) * 100 : 0)}.`,
                `Top sales category is ${categories[0]?.name || 'N/A'} at ${formatNumber(categories[0]?.sales || 0)}.`,
                `Largest customer in the filtered period is ${customers[0]?.name || 'N/A'} with ${formatNumber(customers[0]?.sales || 0)} sales.`,
                `Status with the highest sales value is ${statuses[0]?.name || 'N/A'} at ${formatNumber(statuses[0]?.sales || 0)}.`
            ]);
        }

        function renderProfitabilityAnalysis() {
            const records = getFilteredDemandRecords();
            const products = aggregateDemandRecords(records, 'product_name').sort((a, b) => b.gp - a.gp);
            const customers = aggregateDemandRecords(records, 'customer_name').sort((a, b) => b.gp - a.gp);
            const categories = aggregateDemandRecords(records, 'category').sort((a, b) => b.gp - a.gp);
            const suppliers = aggregateDemandRecords(records, 'supplier_name').sort((a, b) => b.gp - a.gp);
            const quarterMap = {};
            records.forEach(record => {
                if (!record.posting_date) return;
                const date = new Date(record.posting_date);
                if (isNaN(date)) return;
                const quarter = `Q${Math.floor(date.getMonth() / 3) + 1} ${date.getFullYear()}`;
                if (!quarterMap[quarter]) quarterMap[quarter] = { sales: 0, gp: 0 };
                quarterMap[quarter].sales += Number(record.sales_jod || 0);
                quarterMap[quarter].gp += Number(record.gp_jod || 0);
            });
            const quarters = Object.entries(quarterMap).map(([name, values]) => ({ name, sales: values.sales, gp: values.gp, gm: values.sales > 0 ? (values.gp / values.sales) * 100 : 0 }));

            function tableHtml(items, mapper, colspan, empty='No data') {
                const rows = items.map(mapper).join('');
                return rows || `<tr><td colspan="${colspan}" class="text-center">${empty}</td></tr>`;
            }

            document.querySelector('#profitByProductTable').innerHTML = tableHtml(products.slice(0, 10), p => `<tr><td class="text-left">${(p.name || 'N/A').substring(0, 20)}</td><td class="text-right">${formatNumber(p.sales)}</td><td class="text-right">${formatNumber(p.gp)}</td><td class="text-right">${p.gm.toFixed(1)}%</td></tr>`, 4);
            document.querySelector('#profitByCustomerTable').innerHTML = tableHtml(customers.slice(0, 10), c => `<tr><td class="text-left">${(c.name || 'N/A').substring(0, 20)}</td><td class="text-right">${formatNumber(c.sales)}</td><td class="text-right">${formatNumber(c.gp)}</td><td class="text-right">${c.gm.toFixed(1)}%</td></tr>`, 4);
            document.querySelector('#profitByCategoryTable').innerHTML = tableHtml(categories.slice(0, 10), c => `<tr><td class="text-left">${c.name || 'N/A'}</td><td class="text-right">${formatNumber(c.sales)}</td><td class="text-right">${formatNumber(c.gp)}</td><td class="text-right">${c.gm.toFixed(1)}%</td></tr>`, 4);
            document.querySelector('#profitBySupplierTable').innerHTML = tableHtml(suppliers.slice(0, 10), s => `<tr><td class="text-left">${(s.name || 'N/A').substring(0, 20)}</td><td class="text-right">${formatNumber(s.sales)}</td><td class="text-right">${formatNumber(s.gp)}</td><td class="text-right">${s.gm.toFixed(1)}%</td></tr>`, 4);
            document.querySelector('#profitByQuarterTable').innerHTML = tableHtml(quarters, q => `<tr><td class="text-left">${q.name}</td><td class="text-right">${formatNumber(q.sales)}</td><td class="text-right">${formatNumber(q.gp)}</td><td class="text-right">${q.gm.toFixed(1)}%</td></tr>`, 4);
            document.querySelector('#lossMakingProductsTable').innerHTML = tableHtml(products.filter(p => p.gp < 0).slice(0, 10), p => `<tr><td class="text-left">${(p.name || 'N/A').substring(0, 20)}</td><td class="text-right">${formatNumber(p.sales)}</td><td class="text-right text-danger">${formatNumber(p.gp)}</td><td class="text-right text-danger">${p.gm.toFixed(1)}%</td></tr>`, 4, 'No loss-making products');
            renderInsightsList('profitabilityInsightsBody', [
                `Highest profit product is ${products[0]?.name || 'N/A'} with ${formatNumber(products[0]?.gp || 0)} gross profit.`,
                `Highest profit customer is ${customers[0]?.name || 'N/A'} with ${formatNumber(customers[0]?.gp || 0)} gross profit.`,
                `Strongest category margin in the filtered data is ${categories.sort((a, b) => b.gm - a.gm)[0]?.name || 'N/A'} at ${formatPercent(categories.sort((a, b) => b.gm - a.gm)[0]?.gm || 0)}.`,
                `${products.filter(p => p.gp < 0).length} products are currently loss-making in the filtered period.`
            ]);
        }

        function renderCustomerAnalysis() {
            const records = getFilteredDemandRecords();
            const customers = attachGpShare(buildEntityStats(records, 'customer_name')).sort((a, b) => b.sales - a.sales);
            const profitabilityHtml = customers.slice(0, 20).map(c => `<tr><td class="text-left">${(c.name || 'N/A').substring(0, 25)}</td><td class="text-right">${formatNumber(c.sales)}</td><td class="text-right">${formatNumber(c.gp)}</td><td class="text-right">${c.gm.toFixed(1)}%</td></tr>`).join('');
            document.querySelector('#clientProfitabilityTable').innerHTML = profitabilityHtml || '<tr><td colspan="4" class="text-center">No data</td></tr>';

            const lowMargin = customers.filter(c => c.gp < 0 || c.gm < 5).sort((a, b) => a.gm - b.gm);
            const lowMarginHtml = lowMargin.slice(0, 20).map(c => {
                const status = c.gp < 0 ? '<span class="status-badge status-at-risk">LOSS</span>' : '<span class="status-badge status-on-track">Low Margin</span>';
                return `<tr><td class="text-left">${(c.name || 'N/A').substring(0, 22)}</td><td class="text-right">${formatNumber(c.sales)}</td><td class="text-right ${c.gp < 0 ? 'text-danger' : ''}">${formatNumber(c.gp)}</td><td class="text-right">${c.gm.toFixed(1)}%</td><td class="text-right">5.0%</td><td>${status}</td></tr>`;
            }).join('');
            document.querySelector('#lossLowMarginTable').innerHTML = lowMarginHtml || '<tr><td colspan="6" class="text-center">No loss or low-margin customers</td></tr>';

            const dayMap = {};
            records.forEach(record => {
                const customer = record.customer_name || 'N/A';
                const posting = record.posting_date ? new Date(record.posting_date) : null;
                const purch = record.purch_posting_date ? new Date(record.purch_posting_date) : null;
                if (!posting || !purch || isNaN(posting) || isNaN(purch)) return;
                const diff = (posting - purch) / (1000 * 60 * 60 * 24);
                if (diff < 0) return;
                if (!dayMap[customer]) dayMap[customer] = [];
                dayMap[customer].push(diff);
            });
            const daysRows = Object.entries(dayMap).map(([customer, values]) => ({ customer, avg: values.reduce((a, b) => a + b, 0) / values.length, max: Math.max(...values), count: values.length })).sort((a, b) => b.avg - a.avg);
            document.querySelector('#purchaseSalesDaysTable').innerHTML = daysRows.slice(0, 20).map(c => `<tr><td class="text-left">${c.customer.substring(0, 22)}</td><td class="text-right">${c.avg.toFixed(1)} days</td><td class="text-right">${c.max.toFixed(0)} days</td><td class="text-right">${c.count}</td></tr>`).join('') || '<tr><td colspan="4" class="text-center">No date-linked data</td></tr>';

            document.querySelector('#customerVolumeTable').innerHTML = customers.slice(0, 20).map(c => `<tr><td class="text-left">${(c.name || 'N/A').substring(0, 22)}</td><td class="text-right">${formatNumber(c.sales)}</td><td class="text-right">${formatNumber(c.qty)}</td><td class="text-right">${c.gm.toFixed(1)}%</td></tr>`).join('') || '<tr><td colspan="4" class="text-center">No data</td></tr>';
            document.querySelector('#topWorstCustomersTable').innerHTML = buildTopWorstTableRows(customers, 'gp', formatNumber, (item) => `<td class="text-left">${(item.name || 'N/A').substring(0, 22)}</td><td class="text-right">${formatNumber(item.sales)}</td><td class="text-right">${formatNumber(item.gp)}</td><td class="text-right">${item.gm.toFixed(1)}%</td><td class="text-right">${formatNumber(item.qty)}</td><td class="text-right">${item.invoices}</td><td class="text-right">${item.gpShare.toFixed(1)}%</td>`) || '<tr><td colspan="8" class="text-center">No data</td></tr>';
            renderBubbleMatrix('customerMatrixChart', 'customerMatrixChart', customers, 'sales', 'gm', 'gp', 'name');

            const latestDate = records.reduce((maxDate, record) => {
                const current = record.posting_date ? new Date(record.posting_date) : null;
                return current && !isNaN(current) && (!maxDate || current > maxDate) ? current : maxDate;
            }, null);
            const customerPrevMap = {};
            records.forEach(record => {
                if (!record.posting_date || !latestDate) return;
                const current = new Date(record.posting_date);
                if (isNaN(current)) return;
                const prevYearDate = new Date(latestDate);
                prevYearDate.setFullYear(prevYearDate.getFullYear() - 1);
                if (current <= prevYearDate) {
                    const key = record.customer_name || 'N/A';
                    customerPrevMap[key] = (customerPrevMap[key] || 0) + Number(record.sales_jod || 0);
                }
            });
            const riskRows = customers.map(customer => ({
                name: customer.name,
                currentSales: customer.sales,
                prevSales: customerPrevMap[customer.name] || 0,
                openPct: customer.openPct,
                avgCreditDays: customer.avgCreditDays,
                status: customer.sales === 0 && (customerPrevMap[customer.name] || 0) > 0 ? 'Churned' : (customer.openPct > 30 ? 'Open Risk' : 'Active')
            })).sort((a, b) => {
                if (a.status === 'Churned' && b.status !== 'Churned') return -1;
                if (a.status !== 'Churned' && b.status === 'Churned') return 1;
                return (b.openPct + b.avgCreditDays) - (a.openPct + a.avgCreditDays);
            });
            document.querySelector('#customerRiskWatchTable').innerHTML = riskRows.slice(0, 10).map(r => `<tr><td class="text-left">${(r.name || 'N/A').substring(0, 22)}</td><td>${r.status}</td><td class="text-right">${formatNumber(r.currentSales)}</td><td class="text-right">${formatNumber(r.prevSales)}</td><td class="text-right">${r.openPct.toFixed(1)}%</td><td class="text-right">${r.avgCreditDays.toFixed(1)} days</td></tr>`).join('') || '<tr><td colspan="6" class="text-center">No data</td></tr>';
            renderInsightsList('customerInsightsBody', [
                `Top customer by sales is ${customers[0]?.name || 'N/A'} with ${formatNumber(customers[0]?.sales || 0)}.`,
                `${lowMargin.length} customers are below the minimum margin threshold or loss-making.`,
                `Longest average purchase-to-sales cycle belongs to ${daysRows[0]?.customer || 'N/A'} at ${(daysRows[0]?.avg || 0).toFixed(1)} days.`,
                `Customer cards above compare actual sales and GP against summary-sheet targets for each client.`,
                `Top 5 and Worst 5 customers are ranked by gross profit and include invoice count plus share of total GP.`
            ]);
        }

        function renderProductAnalysis() {
            const records = getFilteredDemandRecords();
            const products = attachGpShare(buildEntityStats(records, 'product_name')).sort((a, b) => b.sales - a.sales);
            const categories = aggregateDemandRecords(records, 'category').sort((a, b) => b.sales - a.sales);
            const invoiceMap = {};
            const vendorMap = {};
            records.forEach(record => {
                const product = record.product_name || 'N/A';
                const invoice = record.invoice_no || `ROW-${Math.random()}`;
                if (!invoiceMap[product]) invoiceMap[product] = { invoices: new Set(), sales: 0 };
                invoiceMap[product].invoices.add(invoice);
                invoiceMap[product].sales += Number(record.sales_jod || 0);
                const supplier = record.supplier_name || 'N/A';
                const key = `${product}|||${supplier}`;
                if (!vendorMap[key]) vendorMap[key] = { product, supplier, sales: 0, gp: 0, count: 0 };
                vendorMap[key].sales += Number(record.sales_jod || 0);
                vendorMap[key].gp += Number(record.gp_jod || 0);
                vendorMap[key].count += 1;
            });
            const invoiceRows = Object.entries(invoiceMap).map(([product, values]) => ({ product, invoice_count: values.invoices.size, avg_invoice_value: values.invoices.size ? values.sales / values.invoices.size : 0, total_sales: values.sales })).sort((a, b) => b.invoice_count - a.invoice_count);
            const bestVendorRows = Object.values(vendorMap).map(v => ({ product: v.product, best_vendor: v.supplier, total_purchases: v.count, avg_gp_percent: v.sales > 0 ? (v.gp / v.sales) * 100 : 0, sales: v.sales })).sort((a, b) => b.sales - a.sales);

            document.querySelector('#bestSellingTable').innerHTML = products.slice(0, 10).map(p => `<tr><td class="text-left">${(p.name || 'N/A').substring(0, 20)}</td><td class="text-right">${formatNumber(p.qty)}</td><td class="text-right">${formatNumber(p.sales)}</td><td class="text-right">${formatNumber(p.gp)}</td><td class="text-right">${p.gm.toFixed(1)}%</td></tr>`).join('') || '<tr><td colspan="5" class="text-center">No data</td></tr>';
            document.querySelector('#productsByCategoryTable').innerHTML = categories.slice(0, 10).map(c => `<tr><td class="text-left">${c.name || 'N/A'}</td><td class="text-right">${formatNumber(records.filter(r => (r.category || 'N/A') === c.name).length)}</td><td class="text-right">${formatNumber(c.sales)}</td></tr>`).join('') || '<tr><td colspan="3" class="text-center">No data</td></tr>';
            document.querySelector('#categoryContributionTable').innerHTML = categories.slice(0, 10).map(c => `<tr><td class="text-left">${c.name || 'N/A'}</td><td class="text-right">${formatNumber(c.sales)}</td><td class="text-right">${((c.sales / Math.max(1, products.reduce((s, p) => s + p.sales, 0))) * 100).toFixed(1)}%</td><td class="text-right">${formatNumber(c.gp)}</td><td class="text-right">${c.gm.toFixed(1)}%</td></tr>`).join('') || '<tr><td colspan="5" class="text-center">No data</td></tr>';
            document.querySelector('#invoiceCountPerProductTable').innerHTML = invoiceRows.slice(0, 10).map(p => `<tr><td class="text-left">${p.product.substring(0, 20)}</td><td class="text-right">${p.invoice_count}</td><td class="text-right">${formatNumber(p.avg_invoice_value)}</td><td class="text-right">${formatNumber(p.total_sales)}</td></tr>`).join('') || '<tr><td colspan="4" class="text-center">No data</td></tr>';
            document.querySelector('#bestVendorPerProductTable').innerHTML = bestVendorRows.slice(0, 10).map(p => `<tr><td class="text-left">${p.product.substring(0, 20)}</td><td class="text-left">${p.best_vendor}</td><td class="text-right">${formatNumber(p.total_purchases)}</td><td class="text-right">${p.avg_gp_percent.toFixed(1)}%</td></tr>`).join('') || '<tr><td colspan="4" class="text-center">No data</td></tr>';
            document.querySelector('#topWorstProductsTable').innerHTML = buildTopWorstTableRows(products, 'gp', formatNumber, (item) => `<td class="text-left">${(item.name || 'N/A').substring(0, 20)}</td><td class="text-right">${formatNumber(item.sales)}</td><td class="text-right">${formatNumber(item.gp)}</td><td class="text-right">${item.gm.toFixed(1)}%</td><td class="text-right">${formatNumber(item.qty)}</td><td class="text-right">${item.gpShare.toFixed(1)}%</td>`) || '<tr><td colspan="7" class="text-center">No data</td></tr>';
            renderBubbleMatrix('productMatrixChart', 'productMatrixChart', products, 'qty', 'gm', 'gp', 'name', 'rgba(5,150,105,0.65)', 'rgba(220,38,38,0.65)');
            const latestProductDate = records.reduce((maxDate, record) => {
                const current = record.posting_date ? new Date(record.posting_date) : null;
                return current && !isNaN(current) && (!maxDate || current > maxDate) ? current : maxDate;
            }, null);
            const slowMoving = products.filter(p => {
                if (!latestProductDate || p.lastDateText === 'N/A') return false;
                const lastDate = new Date(p.lastDateText);
                return !isNaN(lastDate) && ((latestProductDate - lastDate) / (1000 * 60 * 60 * 24)) >= 180;
            }).sort((a, b) => a.sales - b.sales);
            document.querySelector('#slowMovingProductsTable').innerHTML = slowMoving.slice(0, 10).map(p => `<tr><td class="text-left">${(p.name || 'N/A').substring(0, 20)}</td><td class="text-left">${(p.categories[0] || 'N/A').substring(0, 16)}</td><td class="text-right">${p.lastDateText}</td><td class="text-right">${formatNumber(p.sales)}</td><td class="text-right">${formatNumber(p.gp)}</td><td class="text-right">${formatNumber(p.qty)}</td></tr>`).join('') || '<tr><td colspan="6" class="text-center">No slow-moving products in current filter</td></tr>';
            renderInsightsList('productInsightsBody', [
                `Best-selling product is ${products[0]?.name || 'N/A'} with ${formatNumber(products[0]?.sales || 0)} sales.`,
                `Best product by gross profit is ${[...products].sort((a, b) => b.gp - a.gp)[0]?.name || 'N/A'} at ${formatNumber([...products].sort((a, b) => b.gp - a.gp)[0]?.gp || 0)}.`,
                `Largest category is ${categories[0]?.name || 'N/A'} with ${formatNumber(categories[0]?.sales || 0)} sales.`,
                `Most frequently invoiced product is ${invoiceRows[0]?.product || 'N/A'} with ${invoiceRows[0]?.invoice_count || 0} invoices.`,
                `${slowMoving.length} products qualify as slow-moving based on the last 180 days check in the current dataset.`
            ]);
        }

        function renderSupplierAnalysis() {
            const records = getFilteredDemandRecords();
            const suppliers = attachGpShare(buildEntityStats(records, 'supplier_name')).sort((a, b) => b.sales - a.sales);
            const supplierPsMap = {};
            const supplierCreditMap = {};
            records.forEach(record => {
                const supplier = record.supplier_name || 'N/A';
                const posting = record.posting_date ? new Date(record.posting_date) : null;
                const purch = record.purch_posting_date ? new Date(record.purch_posting_date) : null;
                const due = record.due_date ? new Date(record.due_date) : null;
                if (posting && purch && !isNaN(posting) && !isNaN(purch)) {
                    const diff = (posting - purch) / (1000 * 60 * 60 * 24);
                    if (diff >= 0) {
                        if (!supplierPsMap[supplier]) supplierPsMap[supplier] = [];
                        supplierPsMap[supplier].push(diff);
                    }
                }
                if (due && purch && !isNaN(due) && !isNaN(purch)) {
                    const credit = (due - purch) / (1000 * 60 * 60 * 24);
                    if (credit >= 0) {
                        if (!supplierCreditMap[supplier]) supplierCreditMap[supplier] = [];
                        supplierCreditMap[supplier].push(credit);
                    }
                }
            });
            const supplierPsRows = Object.entries(supplierPsMap).map(([supplier, values]) => ({ supplier, avg: values.reduce((a, b) => a + b, 0) / values.length, max: Math.max(...values), count: values.length })).sort((a, b) => b.avg - a.avg);
            const supplierCreditRows = Object.entries(supplierCreditMap).map(([supplier, values]) => ({ supplier, avg: values.reduce((a, b) => a + b, 0) / values.length, max: Math.max(...values), count: values.length })).sort((a, b) => b.avg - a.avg);
            const lossItemMap = {};
            records.forEach(record => {
                const key = `${record.supplier_name || 'N/A'}|||${record.product_name || 'N/A'}`;
                if (!lossItemMap[key]) {
                    lossItemMap[key] = { supplier: record.supplier_name || 'N/A', product: record.product_name || 'N/A', sales: 0, gp: 0 };
                }
                lossItemMap[key].sales += Number(record.sales_jod || 0);
                lossItemMap[key].gp += Number(record.gp_jod || 0);
            });
            const lossItems = Object.values(lossItemMap).map(item => ({
                ...item,
                gp_percent: item.sales > 0 ? (item.gp / item.sales) * 100 : 0
            })).filter(item => item.gp < 0).sort((a, b) => a.gp - b.gp);

            document.querySelector('#vendorGpTable').innerHTML = suppliers.slice(0, 20).map(s => `<tr><td class="text-left">${(s.name || 'N/A').substring(0, 25)}</td><td class="text-right">${formatNumber(s.sales)}</td><td class="text-right">${formatNumber(s.gp)}</td><td class="text-right">${s.gm.toFixed(1)}%</td></tr>`).join('') || '<tr><td colspan="4" class="text-center">No data</td></tr>';
            document.querySelector('#supplierPsDaysTable').innerHTML = supplierPsRows.slice(0, 20).map(s => `<tr><td class="text-left">${s.supplier.substring(0, 20)}</td><td class="text-right">${s.avg.toFixed(1)} days</td><td class="text-right">${s.max.toFixed(0)} days</td><td class="text-right">${s.count}</td></tr>`).join('') || '<tr><td colspan="4" class="text-center">No data</td></tr>';
            document.querySelector('#lossMakingVendorItemsTable').innerHTML = lossItems.slice(0, 20).map(p => `<tr><td class="text-left">${p.supplier.substring(0, 15)}</td><td class="text-left">${p.product.substring(0, 20)}</td><td class="text-right">${formatNumber(p.sales)}</td><td class="text-right text-danger">${formatNumber(p.gp)}</td><td class="text-right text-danger">${p.gp_percent.toFixed(1)}%</td></tr>`).join('') || '<tr><td colspan="5" class="text-center">No loss-making items</td></tr>';
            document.querySelector('#supplierVolumeTable').innerHTML = suppliers.slice(0, 20).map(s => `<tr><td class="text-left">${(s.name || 'N/A').substring(0, 22)}</td><td class="text-right">${formatNumber(s.sales)}</td><td class="text-right">${formatNumber(s.qty)}</td><td class="text-right">${s.gm.toFixed(1)}%</td></tr>`).join('') || '<tr><td colspan="4" class="text-center">No data</td></tr>';
            document.querySelector('#vendorCreditDaysTable').innerHTML = supplierCreditRows.slice(0, 20).map(s => `<tr><td class="text-left">${s.supplier.substring(0, 20)}</td><td class="text-right">${s.avg.toFixed(1)} days</td><td class="text-right">${s.max.toFixed(0)} days</td><td class="text-right">${s.count}</td></tr>`).join('') || '<tr><td colspan="4" class="text-center">No data</td></tr>';
            document.querySelector('#topWorstSuppliersTable').innerHTML = buildTopWorstTableRows(suppliers, 'gp', formatNumber, (item) => `<td class="text-left">${(item.name || 'N/A').substring(0, 22)}</td><td class="text-right">${formatNumber(item.purchase || item.sales)}</td><td class="text-right">${formatNumber(item.gp)}</td><td class="text-right">${item.gm.toFixed(1)}%</td><td class="text-right">${item.gpShare.toFixed(1)}%</td>`) || '<tr><td colspan="6" class="text-center">No data</td></tr>';
            renderBubbleMatrix('supplierMatrixChart', 'supplierMatrixChart', suppliers.map(s => ({ ...s, purchaseMetric: s.purchase || s.sales })), 'purchaseMetric', 'gm', 'gp', 'name', 'rgba(99,102,241,0.65)', 'rgba(220,38,38,0.65)');
            const totalPurchase = suppliers.reduce((sum, s) => sum + (s.purchase || s.sales), 0);
            const riskRows = suppliers.map(s => ({
                name: s.name,
                purchase: s.purchase || s.sales,
                purchasePct: totalPurchase > 0 ? ((s.purchase || s.sales) / totalPurchase) * 100 : 0,
                avgCreditDays: s.avgCreditDays,
                avgPsDays: s.avgPsDays
            })).sort((a, b) => b.purchasePct - a.purchasePct);
            document.querySelector('#supplierRiskTable').innerHTML = riskRows.slice(0, 10).map(r => `<tr><td class="text-left">${(r.name || 'N/A').substring(0, 22)}</td><td class="text-right">${formatNumber(r.purchase)}</td><td class="text-right">${r.purchasePct.toFixed(1)}%</td><td class="text-right">${r.avgCreditDays.toFixed(1)} days</td><td class="text-right">${r.avgPsDays.toFixed(1)} days</td></tr>`).join('') || '<tr><td colspan="5" class="text-center">No data</td></tr>';
            renderInsightsList('supplierInsightsBody', [
                `Top supplier by sales is ${suppliers[0]?.name || 'N/A'} with ${formatNumber(suppliers[0]?.sales || 0)}.`,
                `${lossItems.length} supplier-product combinations are currently loss-making.`,
                `Longest supplier purchase-to-sales cycle belongs to ${supplierPsRows[0]?.supplier || 'N/A'} at ${(supplierPsRows[0]?.avg || 0).toFixed(1)} days.`,
                `Highest average credit term belongs to ${supplierCreditRows[0]?.supplier || 'N/A'} at ${(supplierCreditRows[0]?.avg || 0).toFixed(1)} days.`,
                `Supplier concentration reaches ${riskRows[0]?.purchasePct?.toFixed(1) || '0.0'}% with the largest supplier in the filtered view.`
            ]);
        }

        function getTrackerRecords(statusName) {
            const records = dashboardData.sop_records || [];
            return records.filter(record =>
                String(record.section || '').trim() === 'Trade' &&
                String(record.status || '').trim() === statusName
            );
        }

        function getTrackerFilterLabel(defaultLabel, selectedValues) {
            const safeSelectedValues = Array.isArray(selectedValues) ? selectedValues : [];
            if (!safeSelectedValues.length) return defaultLabel;
            if (safeSelectedValues.length === 1) return safeSelectedValues[0];
            return `${safeSelectedValues.length} selected`;
        }

        function toggleTrackerDropdown(field) {
            document.querySelectorAll('.tracker-filter-menu').forEach(menu => {
                if (menu.dataset.field !== field) menu.classList.remove('open');
            });
            const menu = document.querySelector(`.tracker-filter-menu[data-field="${field}"]`);
            const button = document.querySelector(`.tracker-filter-btn[data-field="${field}"]`);
            if (!menu || !button) return;
            const willOpen = !menu.classList.contains('open');
            menu.classList.toggle('open', willOpen);
            button.classList.toggle('active', willOpen);
        }

        function closeTrackerDropdowns() {
            document.querySelectorAll('.tracker-filter-menu').forEach(menu => menu.classList.remove('open'));
            document.querySelectorAll('.tracker-filter-btn').forEach(button => button.classList.remove('active'));
        }

        function toggleTrackerFilterValue(field, value, checked) {
            const currentValues = Array.isArray(trackerFilters[field]) ? [...trackerFilters[field]] : [];
            const nextValues = checked
                ? [...new Set([...currentValues, value])]
                : currentValues.filter(item => item !== value);
            trackerFilters[field] = nextValues;
            populateTrackerFilters();
            renderTrackerAnalysis();
        }

        function populateTrackerSelect(selectId, field, values, selectedValues, defaultLabel) {
            const container = document.getElementById(selectId);
            if (!container) return;
            const uniqueValues = [...new Set(values.map(value => normalizeTrackerLabel(value, '')).filter(Boolean))].sort((a, b) => a.localeCompare(b));
            const safeSelectedValues = Array.isArray(selectedValues) ? selectedValues : [];
            const optionsHtml = uniqueValues.length
                ? uniqueValues.map(value => `<label class="tracker-filter-option"><input type="checkbox" ${safeSelectedValues.includes(value) ? 'checked' : ''} onchange="toggleTrackerFilterValue('${field}', decodeURIComponent('${encodeURIComponent(value)}'), this.checked)"> <span>${escapeHtml(value)}</span></label>`).join('')
                : `<div class="tracker-filter-empty">No values</div>`;
            container.innerHTML = `
                <div class="tracker-filter-dropdown">
                    <button type="button" class="tracker-filter-btn ${safeSelectedValues.length ? 'active' : ''}" data-field="${field}" onclick="toggleTrackerDropdown('${field}')">
                        <span>${escapeHtml(getTrackerFilterLabel(defaultLabel, safeSelectedValues))}</span>
                        <i class="bi bi-chevron-down"></i>
                    </button>
                    <div class="tracker-filter-menu" data-field="${field}">
                        ${optionsHtml}
                    </div>
                </div>`;
        }

        function populateTrackerFilters() {
            const trackerRecords = (dashboardData.sop_records || []).filter(record => String(record.section || '').trim() === 'Trade');
            populateTrackerSelect('trackerFileFilter', 'file', trackerRecords.map(record => record.file), trackerFilters.file, 'All Files');
            populateTrackerSelect('trackerStatusFilter', 'status', trackerRecords.map(record => record.status), trackerFilters.status, 'All Statuses');
            populateTrackerSelect('trackerMonthFilter', 'month', trackerRecords.map(record => record.actual_month), trackerFilters.month, 'All Months');
            populateTrackerSelect('trackerSalesPersonFilter', 'sales_person', trackerRecords.map(record => record.sales_person), trackerFilters.sales_person, 'Sales Persons');
            populateTrackerSelect('trackerClientFilter', 'client', trackerRecords.map(record => record.client), trackerFilters.client, 'All Clients');
            populateTrackerSelect('trackerProductFilter', 'product', trackerRecords.map(record => record.product), trackerFilters.product, 'All Products');
        }

        function resetTrackerFilters() {
            trackerFilters = { file: [], status: [], month: [], sales_person: [], client: [], product: [] };
            populateTrackerFilters();
            renderTrackerAnalysis();
        }

        function applyTrackerFilters(records) {
            return records.filter(record => {
                if (trackerFilters.file.length && !trackerFilters.file.includes(normalizeTrackerLabel(record.file, ''))) return false;
                if (trackerFilters.status.length && !trackerFilters.status.includes(normalizeTrackerLabel(record.status, ''))) return false;
                if (trackerFilters.month.length && !trackerFilters.month.includes(normalizeTrackerLabel(record.actual_month, ''))) return false;
                if (trackerFilters.sales_person.length && !trackerFilters.sales_person.includes(normalizeTrackerLabel(record.sales_person, ''))) return false;
                if (trackerFilters.client.length && !trackerFilters.client.includes(normalizeTrackerLabel(record.client, ''))) return false;
                if (trackerFilters.product.length && !trackerFilters.product.includes(normalizeTrackerLabel(record.product, ''))) return false;
                return true;
            });
        }

        function normalizeTrackerLabel(value, fallback = 'N/A') {
            const text = value === undefined || value === null ? '' : String(value).trim();
            return text || fallback;
        }

        function trackerClientDisplayName(value) {
            const words = normalizeTrackerLabel(value, '').split(/\\s+/).filter(Boolean);
            if (words.length <= 2) return words.join(' ');
            return `${words[0]} ${words[1]}`;
        }

        function aggregateTracker(records, fieldName, valueField) {
            const grouped = {};
            records.forEach(record => {
                const key = normalizeTrackerLabel(record[fieldName]);
                if (!grouped[key]) {
                    grouped[key] = { name: key, count: 0, qty: 0, sales: 0, gp: 0 };
                }
                grouped[key].count += 1;
                grouped[key].qty += Number(record.qty_mt || 0);
                grouped[key].sales += Number(record.sales_jod || 0);
                grouped[key].gp += Number(record.gp_jod || 0);
            });
            return Object.values(grouped).sort((a, b) => Number(b[valueField] || 0) - Number(a[valueField] || 0));
        }

        function escapeCsv(value) {
            const text = value === undefined || value === null ? '' : String(value);
            return `"${text.replace(/"/g, '""')}"`;
        }

        function escapeHtml(value) {
            const text = value === undefined || value === null ? '' : String(value);
            return text
                .replace(/&/g, '&amp;')
                .replace(/</g, '&lt;')
                .replace(/>/g, '&gt;')
                .replace(/"/g, '&quot;')
                .replace(/'/g, '&#39;');
        }

        function updateTrackerEdit(excelRow, field, value) {
            const rowKey = String(excelRow || '').trim();
            if (!rowKey) return;
            const currentRecord = trackerRecordLookup[rowKey] || {};
            if (!trackerEdits[rowKey]) {
                trackerEdits[rowKey] = {
                    excel_row: rowKey,
                    file: currentRecord.file || '',
                    current_status: currentRecord.status || '',
                    current_month: currentRecord.actual_month || '',
                    client: currentRecord.client || '',
                    product: currentRecord.product || '',
                    sales_person: currentRecord.sales_person || '',
                    action: '',
                    new_status: '',
                    new_month: '',
                    new_reason: '',
                    new_notes: ''
                };
            }
            trackerEdits[rowKey][field] = value;
        }

        function buildTrackerRow(record) {
            const excelRow = String(record.excel_row || '').trim();
            if (excelRow) trackerRecordLookup[excelRow] = record;
            const rowEdit = trackerEdits[excelRow] || {};
            const recordStatus = String(record.status || '').trim();
            const badgeClass = recordStatus === 'Booked' ? 'tracker-status-booked' : 'tracker-status-not-booked';
            const currentAction = rowEdit.action || '';
            const currentStatus = rowEdit.new_status || '';
            const currentMonth = rowEdit.new_month || '';
            const currentReason = rowEdit.new_reason !== undefined ? rowEdit.new_reason : (record.reason || '');
            const currentNotes = rowEdit.new_notes !== undefined ? rowEdit.new_notes : (record.notes || '');
            const monthOptions = monthNames.map(month => `<option value="${month}" ${currentMonth === month ? 'selected' : ''}>${month}</option>`).join('');
            const statusOptions = ['Booked', 'Not Booked', 'Invoiced', 'Removed'].map(status => `<option value="${status}" ${currentStatus === status ? 'selected' : ''}>${status}</option>`).join('');
            const actionOptions = [
                { value: '', label: 'Keep' },
                { value: 'update', label: 'Update' },
                { value: 'shift', label: 'Shift' },
                { value: 'remove', label: 'Remove' }
            ].map(item => `<option value="${item.value}" ${currentAction === item.value ? 'selected' : ''}>${item.label}</option>`).join('');

            return `<tr>
                <td><span class="status-badge ${badgeClass}">${escapeHtml(recordStatus || '')}</span></td>
                <td><span class="status-badge ${badgeClass}">${escapeHtml(record.actual_month || '')}</span></td>
                <td>${escapeHtml(record.sales_person || '')}</td>
                <td>${escapeHtml(trackerClientDisplayName(record.client || ''))}</td>
                <td>${escapeHtml(record.product || '')}</td>
                <td class="text-right">${formatNumber(record.sales_jod || 0)}</td>
                <td class="text-right">${formatNumber(record.gp_jod || 0)}</td>
                <td>
                    <select onchange="updateTrackerEdit('${excelRow}', 'action', this.value)">
                        ${actionOptions}
                    </select>
                </td>
                <td>
                    <select onchange="updateTrackerEdit('${excelRow}', 'new_status', this.value)">
                        <option value="">Keep Current</option>
                        ${statusOptions}
                    </select>
                </td>
                <td>
                    <select onchange="updateTrackerEdit('${excelRow}', 'new_month', this.value)">
                        <option value="">Keep Current</option>
                        ${monthOptions}
                    </select>
                </td>
                <td><input type="text" value="${escapeHtml(currentReason)}" placeholder="Reason" oninput="updateTrackerEdit('${excelRow}', 'new_reason', this.value)"></td>
            </tr>`;
        }

        function renderTrackerSummary(tableBodyId, rows, emptyMessage) {
            const body = document.getElementById(tableBodyId);
            if (!body) return;
            body.innerHTML = rows.length ? rows.map(row => `<tr>
                <td>${escapeHtml(row.name)}</td>
                <td class="text-right">${formatNumber(row.count)}</td>
                <td class="text-right">${formatNumber(row.sales)}</td>
                <td class="text-right">${formatNumber(row.gp)}</td>
            </tr>`).join('') : `<tr><td colspan="5" class="text-center">${emptyMessage}</td></tr>`;
        }

        function renderTrackerAnalysis() {
            trackerRecordLookup = {};
            const notBookedRecords = applyTrackerFilters(getTrackerRecords('Not Booked')).sort((a, b) => Number(b.sales_jod || 0) - Number(a.sales_jod || 0));
            const bookedRecords = applyTrackerFilters(getTrackerRecords('Booked')).sort((a, b) => Number(b.sales_jod || 0) - Number(a.sales_jod || 0));
            const combinedRecords = [...notBookedRecords, ...bookedRecords].sort((a, b) => Number(b.sales_jod || 0) - Number(a.sales_jod || 0));

            renderTrackerSummary('notBookedMonthTable', aggregateTracker(notBookedRecords, 'actual_month', 'sales'), 'No not booked records');
            const trackerBody = document.getElementById('trackerCombinedBody');
            if (trackerBody) {
                trackerBody.innerHTML = combinedRecords.length
                    ? combinedRecords.map(record => buildTrackerRow(record)).join('')
                    : '<tr><td colspan="13" class="text-center">No tracker data</td></tr>';
            }
        }

        function exportTrackerUpdates() {
            const rows = Object.values(trackerEdits).filter(item =>
                item.action || item.new_status || item.new_month || item.new_reason || item.new_notes
            );
            if (!rows.length) {
                alert('No tracker changes to export.');
                return;
            }
            const headers = [
                'excel_row', 'action', 'new_status', 'new_month', 'new_reason', 'new_notes',
                'current_status', 'current_month', 'file', 'client', 'product', 'sales_person'
            ];
            const csvLines = [headers.join(',')];
            rows.forEach(row => {
                csvLines.push(headers.map(header => escapeCsv(row[header] || '')).join(','));
            });
            const blob = new Blob([csvLines.join('\\r\\n')], { type: 'text/csv;charset=utf-8;' });
            const url = URL.createObjectURL(blob);
            const link = document.createElement('a');
            link.href = url;
            link.download = 'tracker_updates.csv';
            document.body.appendChild(link);
            link.click();
            document.body.removeChild(link);
            URL.revokeObjectURL(url);
        }

        window.onload = initDashboard;
    </script>
</body>
</html>
"""

def load_excel_data():
    """Load ALL data from Excel file dynamically"""
    print("\n" + "="*70)
    print("LOADING EXCEL DATA")
    print("="*70)
    
    candidate_paths = [
        r'C:\Users\a.harayri\OneDrive - Al-Behar Group\Desktop\python\Olis.xlsx',
        r'C:\Users\a.harayri\OneDrive - Al-Behar Group\Desktop\python\oils.xlsx',
        r'C:\Users\a.harayri\OneDrive - Al-Behar Group\Desktop\python\Oils.xlsx',
        r'C:\Users\a.harayri\OneDrive - Al-Behar Group\Desktop\python\Demand.xlsx'
    ]
    excel_path = next((path for path in candidate_paths if os.path.exists(path)), candidate_paths[0])
    
    print(f"\nLooking for: {excel_path}")
    
    if not os.path.exists(excel_path):
        print("File not found. Checked:")
        for path in candidate_paths:
            print(f"   - {path}")
        return None, None, None, [], []
    
    print(f"File found! Size: {os.path.getsize(excel_path)} bytes")
    
    try:
        apply_tracker_updates(excel_path)
        xl = pd.ExcelFile(excel_path, engine='openpyxl')
        sheet_names = xl.sheet_names
        print(f"\nSheets found: {sheet_names}")
        
        df_sop = None
        if 'Data' in sheet_names:
            df_sop = pd.read_excel(excel_path, sheet_name='Data', engine='openpyxl')
            print(f"\n'Data' sheet loaded: {len(df_sop)} rows, {len(df_sop.columns)} columns")
            print(f"   Columns: {list(df_sop.columns)}")
        else:
            print("\n'Data' sheet not found")
        
        df_demand = None
        demand_sheet_name = None
        for sheet in sheet_names:
            if 'Demand Analysis' in sheet:
                demand_sheet_name = sheet
                break
        
        if demand_sheet_name:
            df_demand = pd.read_excel(excel_path, sheet_name=demand_sheet_name, engine='openpyxl')
            print(f"\n'{demand_sheet_name}' sheet loaded: {len(df_demand)} rows")
        else:
            print("\n'Demand Analysis' sheet not found")
        
        file_names = []
        if df_sop is not None and 'File' in df_sop.columns:
            df_sop['File'] = df_sop['File'].fillna('').astype(str).str.strip()
            file_names = [f for f in df_sop['File'].dropna().unique().tolist() if f]
            print(f"\nUnique File names: {len(file_names)}")
            for i, f in enumerate(file_names[:10]):
                print(f"   {i+1}. {f}")
        
        df_summary = None
        summary_columns = []
        if 'summary' in [s.lower() for s in sheet_names]:
            for sheet in sheet_names:
                if sheet.lower() == 'summary':
                    df_summary = pd.read_excel(excel_path, sheet_name=sheet, engine='openpyxl')
                    summary_columns = list(df_summary.columns)
                    print(f"\nFound 'Summary' sheet: {len(df_summary)} rows")
                    break
        
        return df_sop, df_demand, df_summary, file_names, summary_columns
        
    except Exception as e:
        print(f"Error loading file: {e}")
        import traceback
        traceback.print_exc()
        return None, None, None, [], []

def process_sop_data(df_sop, file_names):
    """Process S&OP data from Data sheet"""
    print("\n" + "="*70)
    print("PROCESSING S&OP DATA")
    print("="*70)
    
    sop_data = {
        'file_data': {},
        'filtered_data': {},
        'commission_data': {},
        'records': [],
        'file_names': file_names
    }
    
    if df_sop is None or len(df_sop) == 0:
        print("No S&OP data to process")
        return sop_data
    
    df_sop = df_sop.copy()
    df_sop['excel_row'] = df_sop.index + 2
    df_sop.columns = df_sop.columns.str.strip()
    print(f"\nProcessing columns: {list(df_sop.columns)}")
    
    text_columns = [
        'Section', 'Client', 'Country', 'Product', 'Category', 'Period', 'Quarter',
        'Year', 'File', 'Actual Month', 'Supplier', 'Sales Person', 'New Customer Name',
        'Status', 'Reason', 'Sector', 'Notes'
    ]
    for col in text_columns:
        if col in df_sop.columns:
            df_sop[col] = df_sop[col].fillna('').astype(str).str.strip()
    
    numeric_mapping = {
        'Qty (MT)': 'qty_mt',
        'Sales (JOD)': 'sales_jod',
        'GP (JOD)': 'gp_jod',
        'GM%': 'gm'
    }
    
    for old_col, new_col in numeric_mapping.items():
        if old_col in df_sop.columns:
            df_sop[new_col] = pd.to_numeric(df_sop[old_col], errors='coerce').fillna(0)
            print(f"   Converted: {old_col} -> {new_col}")
        else:
            df_sop[new_col] = 0
            print(f"   Column not found: {old_col}")
    
    record_columns = {
        'excel_row': 'excel_row',
        'File': 'file',
        'Section': 'section',
        'Status': 'status',
        'Actual Month': 'actual_month',
        'Client': 'client',
        'Category': 'category',
        'Country': 'country',
        'Supplier': 'supplier',
        'Sales Person': 'sales_person',
        'Product': 'product',
        'Notes': 'notes',
        'Reason': 'reason',
        'sales_jod': 'sales_jod',
        'gp_jod': 'gp_jod',
        'qty_mt': 'qty_mt',
        'gm': 'gm_percent'
    }
    available_record_columns = [col for col in record_columns if col in df_sop.columns]
    if available_record_columns:
        sop_records_df = df_sop[available_record_columns].copy().rename(columns=record_columns)
        sop_data['records'] = sop_records_df.to_dict(orient='records')
        print(f"   Detailed S&OP records prepared: {len(sop_data['records'])}")
    
    if 'File' in df_sop.columns:
        month_abbrs = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        for file_name in file_names:
            file_df = df_sop[df_sop['File'] == file_name]
            if len(file_df) > 0:
                clean_key = clean_filename_for_key(file_name)
                sop_data['file_data'][f'{clean_key}_sales'] = float(file_df['sales_jod'].sum())
                sop_data['file_data'][f'{clean_key}_qty'] = float(file_df['qty_mt'].sum())
                sop_data['file_data'][f'{clean_key}_gp'] = float(file_df['gp_jod'].sum())
                print(f"\n   {file_name}: Sales={sop_data['file_data'][f'{clean_key}_sales']:,.0f}, GP={sop_data['file_data'][f'{clean_key}_gp']:,.0f}")
                
                if 'Section' in df_sop.columns and 'Actual Month' in df_sop.columns and 'Status' in df_sop.columns:
                    for month_abbr in month_abbrs:
                        invoiced_df = file_df[(file_df['Section'] == 'Trade') & (file_df['Status'] == 'Invoiced') & (file_df['Actual Month'] == month_abbr)]
                        if len(invoiced_df) > 0:
                            filtered_key = f"{clean_key}_Trade_Invoiced_{month_abbr}"
                            sales_sum = float(invoiced_df['sales_jod'].sum())
                            gp_sum = float(invoiced_df['gp_jod'].sum())
                            sop_data['filtered_data'][filtered_key] = {
                                'sales': sales_sum,
                                'qty': float(invoiced_df['qty_mt'].sum()),
                                'gp': gp_sum,
                                'gm': (gp_sum / sales_sum * 100) if sales_sum > 0 else 0
                            }
                        
                        booked_df = file_df[(file_df['Section'] == 'Trade') & (file_df['Status'].isin(['Booked', 'Not Booked'])) & (file_df['Actual Month'] == month_abbr)]
                        if len(booked_df) > 0:
                            filtered_key = f"{clean_key}_Trade_Booked_Not Booked_{month_abbr}"
                            sales_sum = float(booked_df['sales_jod'].sum())
                            gp_sum = float(booked_df['gp_jod'].sum())
                            sop_data['filtered_data'][filtered_key] = {
                                'sales': sales_sum,
                                'qty': float(booked_df['qty_mt'].sum()),
                                'gp': gp_sum,
                                'gm': (gp_sum / sales_sum * 100) if sales_sum > 0 else 0
                            }
                        
                        prev_year_df = file_df[(file_df['Section'] == 'Trade') & (file_df['Actual Month'] == month_abbr)]
                        if len(prev_year_df) > 0:
                            filtered_key = f"{clean_key}_Trade_All_Status_{month_abbr}"
                            sales_sum = float(prev_year_df['sales_jod'].sum())
                            gp_sum = float(prev_year_df['gp_jod'].sum())
                            sop_data['filtered_data'][filtered_key] = {
                                'sales': sales_sum,
                                'qty': float(prev_year_df['qty_mt'].sum()),
                                'gp': gp_sum,
                                'gm': (gp_sum / sales_sum * 100) if sales_sum > 0 else 0
                            }
                    
                    for month_abbr in month_abbrs:
                        commission_df = file_df[(file_df['Section'] == 'Broker') & (file_df['Actual Month'] == month_abbr)]
                        if len(commission_df) > 0:
                            commission_key = f"{clean_key}_Broker_{month_abbr}"
                            sop_data['commission_data'][commission_key] = float(commission_df['gp_jod'].sum())
                            print(f"      Commission for {file_name} - {month_abbr}: {sop_data['commission_data'][commission_key]:,.2f}")
    
    print(f"\nS&OP DATA PROCESSING COMPLETE")
    print(f"   Filtered data combinations: {len(sop_data['filtered_data'])}")
    print(f"   Commission data entries: {len(sop_data['commission_data'])}")
    
    return sop_data

def process_demand_data(df_demand):
    """Process Demand Analysis data and create extended customer, product, and supplier analyses"""
    print("\n" + "="*70)
    print("PROCESSING DEMAND ANALYSIS DATA")
    print("="*70)
    
    demand_data = {
        'summary': {'sales_jod': 0, 'gp_jod': 0, 'gm': 0, 'qty_mt': 0, 'cogs': 0, 'total_clients': 0, 'total_suppliers': 0, 'total_categories': 0, 'total_rows': 0},
        'products': [], 'customers': [], 'suppliers': [], 'categories': [], 'sections': [], 'quarters': [], 'statuses': [], 'countries': [],
        'sales_trend': {'labels': [], 'values': []}, 'category_sales': {'labels': [], 'values': []}, 'records': []
    }
    
    if df_demand is None or len(df_demand) == 0:
        print("No Demand Analysis data")
        return demand_data, {}, {}, {}
    
    df_demand.columns = df_demand.columns.str.strip()
    print(f"\nProcessing columns: {list(df_demand.columns)}")
    
    column_mapping = {
        'Sales Quantity': 'qty_mt', 'Sales Amount': 'sales_jod', 'GP': 'gp_jod',
        'Customer Name': 'customer_name', 'Item Name': 'product_name',
        'Parent Category': 'category', 'Vendor Name': 'supplier_name',
        'Sales Invoice Status': 'status', 'Posting Date': 'posting_date',
        'Due Date': 'due_date', 'Purch Posting Date': 'purch_posting_date',
        'Other Charges-Cost': 'other_charges_cost', 'Other Charges Cost Brokerage': 'brokerage_cost',
        'Invoice No.': 'invoice_no', 'Country': 'country',
        'Purchase Amount': 'purchase_amount', 'Purchase Amount (LCY)': 'purchase_amount_lcy'
    }
    
    for old_col, new_col in column_mapping.items():
        if old_col in df_demand.columns:
            if new_col in ['qty_mt', 'sales_jod', 'gp_jod', 'other_charges_cost', 'brokerage_cost']:
                df_demand[new_col] = pd.to_numeric(df_demand[old_col], errors='coerce').fillna(0)
            else:
                df_demand[new_col] = df_demand[old_col]
            print(f"   Mapped: {old_col} -> {new_col}")
    
    if 'gp_jod' not in df_demand.columns:
        gp_like_columns = [col for col in df_demand.columns if str(col).strip().lower().startswith('gp')]
        if gp_like_columns:
            df_demand['gp_jod'] = pd.to_numeric(df_demand[gp_like_columns[0]], errors='coerce').fillna(0)
            print(f"   Using: {gp_like_columns[0]} -> gp_jod")
    
    for required_numeric_col in ['qty_mt', 'sales_jod', 'gp_jod', 'other_charges_cost', 'brokerage_cost', 'purchase_amount', 'purchase_amount_lcy']:
        if required_numeric_col not in df_demand.columns:
            df_demand[required_numeric_col] = 0
    
    for required_text_col in ['customer_name', 'product_name', 'category', 'supplier_name', 'status']:
        if required_text_col not in df_demand.columns:
            df_demand[required_text_col] = ''
    
    if 'Sales Amount (LCY)' in df_demand.columns:
        df_demand['sales_jod'] = pd.to_numeric(df_demand['Sales Amount (LCY)'], errors='coerce').fillna(0)
        print("   Using: Sales Amount (LCY) -> sales_jod")
    
    date_columns = ['posting_date', 'due_date', 'purch_posting_date']
    for col in date_columns:
        if col in df_demand.columns:
            df_demand[col] = pd.to_datetime(df_demand[col], errors='coerce')

    record_columns = [
        'posting_date', 'customer_name', 'product_name', 'category', 'supplier_name',
        'status', 'sales_jod', 'gp_jod', 'qty_mt', 'invoice_no', 'due_date', 'purch_posting_date',
        'purchase_amount', 'purchase_amount_lcy'
    ]
    available_record_columns = [col for col in record_columns if col in df_demand.columns]
    if available_record_columns:
        demand_data['records'] = df_demand[available_record_columns].to_dict(orient='records')
    
    df_demand['section'] = 'Trade'
    
    demand_data['summary']['sales_jod'] = float(df_demand['sales_jod'].sum())
    demand_data['summary']['gp_jod'] = float(df_demand['gp_jod'].sum())
    demand_data['summary']['qty_mt'] = float(df_demand['qty_mt'].sum())
    demand_data['summary']['gm'] = (demand_data['summary']['gp_jod'] / demand_data['summary']['sales_jod'] * 100) if demand_data['summary']['sales_jod'] > 0 else 0
    demand_data['summary']['cogs'] = demand_data['summary']['sales_jod'] - demand_data['summary']['gp_jod']
    demand_data['summary']['total_rows'] = len(df_demand)
    
    print(f"\nSummary: Sales={demand_data['summary']['sales_jod']:,.0f}, GP={demand_data['summary']['gp_jod']:,.0f}, GM={demand_data['summary']['gm']:.1f}%")
    
    # Customer Analysis
    customer_analysis = {
        'profitability': [],
        'loss_low_margin': [],
        'purchase_sales_days': [],
        'hidden_costs': []
    }
    
    # Product Analysis
    product_analysis = {
        'top_products': [],
        'most_profitable': [],
        'category_contribution': [],
        'invoice_count_per_product': [],
        'best_vendor_per_product': []
    }
    
    # Supplier Analysis
    supplier_analysis = {
        'vendor_gp': [],
        'supplier_ps_days': [],
        'loss_making_items': [],
        'vendor_hidden_costs': [],
        'vendor_credit_days': []
    }
    
    # Process Products
    if 'product_name' in df_demand.columns:
        for name, group in df_demand.groupby('product_name'):
            sales = float(group['sales_jod'].sum())
            gp = float(group['gp_jod'].sum())
            qty = float(group['qty_mt'].sum())
            gp_percent = (gp / sales * 100) if sales > 0 else 0
            invoice_count = group['invoice_no'].nunique() if 'invoice_no' in group.columns else len(group)
            avg_invoice_value = sales / invoice_count if invoice_count > 0 else 0
            
            demand_data['products'].append({'name': str(name), 'sales': sales, 'gp': gp, 'qty': qty, 'gm': gp_percent})
            
            product_analysis['top_products'].append({'product': str(name), 'sales': sales, 'qty': qty})
            product_analysis['most_profitable'].append({'product': str(name), 'gp': gp, 'gp_percent': gp_percent})
            product_analysis['invoice_count_per_product'].append({
                'product': str(name), 
                'invoice_count': invoice_count, 
                'avg_invoice_value': avg_invoice_value,
                'total_sales': sales
            })
            
            if 'supplier_name' in group.columns:
                vendor_stats = group.groupby('supplier_name').agg({
                    'purch_posting_date': 'count',
                    'gp_jod': 'sum',
                    'sales_jod': 'sum'
                }).rename(columns={'purch_posting_date': 'count', 'gp_jod': 'gp', 'sales_jod': 'sales'})
                
                for vendor, row in vendor_stats.iterrows():
                    if pd.notna(vendor):
                        vendor_gp_percent = (row['gp'] / row['sales'] * 100) if row['sales'] > 0 else 0
                        product_analysis['best_vendor_per_product'].append({
                            'product': str(name),
                            'best_vendor': str(vendor),
                            'total_purchases': row['count'],
                            'avg_gp_percent': vendor_gp_percent
                        })
                        break
        
        product_analysis['top_products'].sort(key=lambda x: x['sales'], reverse=True)
        product_analysis['most_profitable'].sort(key=lambda x: x['gp'], reverse=True)
        product_analysis['invoice_count_per_product'].sort(key=lambda x: x['invoice_count'], reverse=True)
        
        if 'category' in df_demand.columns:
            total_sales = demand_data['summary']['sales_jod']
            for name, group in df_demand.groupby('category'):
                sales = float(group['sales_jod'].sum())
                gp = float(group['gp_jod'].sum())
                gp_percent = (gp / sales * 100) if sales > 0 else 0
                percent_of_total = (sales / total_sales * 100) if total_sales > 0 else 0
                product_analysis['category_contribution'].append({
                    'category': str(name),
                    'sales': sales,
                    'percent_of_total': percent_of_total,
                    'gp': gp,
                    'gp_percent': gp_percent
                })
                demand_data['categories'].append({'name': str(name), 'sales': sales, 'gp': gp, 'gm': gp_percent, 'product_count': group['product_name'].nunique() if 'product_name' in group else 0})
            product_analysis['category_contribution'].sort(key=lambda x: x['sales'], reverse=True)
    
    # Process Customers
    if 'customer_name' in df_demand.columns:
        customer_groups = df_demand.groupby('customer_name')
        
        for customer_name, group in customer_groups:
            sales = float(group['sales_jod'].sum())
            gp = float(group['gp_jod'].sum())
            gp_percent = (gp / sales * 100) if sales > 0 else 0
            
            customer_analysis['profitability'].append({
                'customer': str(customer_name),
                'sales': sales,
                'gp': gp,
                'gp_percent': gp_percent
            })
            
            customer_transactions = []
            for idx, row in group.iterrows():
                trans_sales = float(row['sales_jod']) if pd.notna(row['sales_jod']) else 0
                trans_gp = float(row['gp_jod']) if pd.notna(row['gp_jod']) else 0
                trans_gp_percent = (trans_gp / trans_sales * 100) if trans_sales > 0 else 0
                customer_transactions.append(trans_gp_percent)
            
            avg_customer_gp = np.mean(customer_transactions) if customer_transactions else 0
            is_loss = gp < 0
            is_low_margin = (gp_percent < avg_customer_gp) and not is_loss
            
            customer_analysis['loss_low_margin'].append({
                'customer': str(customer_name),
                'sales': sales,
                'gp': gp,
                'gp_percent': gp_percent,
                'avg_gp_percent': avg_customer_gp,
                'is_loss': int(is_loss),
                'is_low_margin': int(is_low_margin)
            })
            
            ps_days_list = []
            for idx, row in group.iterrows():
                posting = row['posting_date'] if pd.notna(row['posting_date']) else None
                purch_posting = row['purch_posting_date'] if pd.notna(row['purch_posting_date']) else None
                if posting and purch_posting and isinstance(posting, pd.Timestamp) and isinstance(purch_posting, pd.Timestamp):
                    ps_days = (posting - purch_posting).days
                    if ps_days > 0:
                        ps_days_list.append(ps_days)
            
            avg_ps_days = np.mean(ps_days_list) if ps_days_list else 0
            max_ps_days = max(ps_days_list) if ps_days_list else 0
            
            customer_analysis['purchase_sales_days'].append({
                'customer': str(customer_name),
                'avg_ps_days': avg_ps_days,
                'max_ps_days': max_ps_days,
                'transaction_count': len(group)
            })
            
            other_charges = float(group['other_charges_cost'].sum()) if 'other_charges_cost' in group.columns else 0
            brokerage = float(group['brokerage_cost'].sum()) if 'brokerage_cost' in group.columns else 0
            
            customer_analysis['hidden_costs'].append({
                'customer': str(customer_name),
                'other_charges_cost': other_charges,
                'brokerage_cost': brokerage,
                'total_hidden_cost': other_charges + brokerage
            })
            
            demand_data['customers'].append({'name': str(customer_name), 'sales': sales, 'gp': gp, 'qty': float(group['qty_mt'].sum()), 'gm': gp_percent})
        
        customer_analysis['profitability'].sort(key=lambda x: x['sales'], reverse=True)
        customer_analysis['loss_low_margin'].sort(key=lambda x: (x['is_loss'], x['gp_percent']))
        customer_analysis['purchase_sales_days'].sort(key=lambda x: x['avg_ps_days'], reverse=True)
        customer_analysis['hidden_costs'].sort(key=lambda x: x['total_hidden_cost'], reverse=True)
        
        demand_data['summary']['total_clients'] = len(demand_data['customers'])
        print(f"   Customers: {len(demand_data['customers'])}")
    
    # Process Suppliers
    if 'supplier_name' in df_demand.columns:
        supplier_groups = df_demand.groupby('supplier_name')
        
        for supplier_name, group in supplier_groups:
            sales = float(group['sales_jod'].sum())
            gp = float(group['gp_jod'].sum())
            gp_percent = (gp / sales * 100) if sales > 0 else 0
            
            supplier_analysis['vendor_gp'].append({
                'supplier': str(supplier_name),
                'sales': sales,
                'gp': gp,
                'gp_percent': gp_percent
            })
            
            ps_days_list = []
            for idx, row in group.iterrows():
                posting = row['posting_date'] if pd.notna(row['posting_date']) else None
                purch_posting = row['purch_posting_date'] if pd.notna(row['purch_posting_date']) else None
                if posting and purch_posting and isinstance(posting, pd.Timestamp) and isinstance(purch_posting, pd.Timestamp):
                    ps_days = (posting - purch_posting).days
                    if ps_days > 0:
                        ps_days_list.append(ps_days)
            
            avg_ps_days = np.mean(ps_days_list) if ps_days_list else 0
            max_ps_days = max(ps_days_list) if ps_days_list else 0
            
            supplier_analysis['supplier_ps_days'].append({
                'supplier': str(supplier_name),
                'avg_ps_days': avg_ps_days,
                'max_ps_days': max_ps_days,
                'transaction_count': len(group)
            })
            
            for product_name, product_group in group.groupby('product_name'):
                product_sales = float(product_group['sales_jod'].sum())
                product_gp = float(product_group['gp_jod'].sum())
                product_gp_percent = (product_gp / product_sales * 100) if product_sales > 0 else 0
                if product_gp < 0:
                    supplier_analysis['loss_making_items'].append({
                        'supplier': str(supplier_name),
                        'product': str(product_name),
                        'sales': product_sales,
                        'gp': product_gp,
                        'gp_percent': product_gp_percent
                    })
            
            other_charges = float(group['other_charges_cost'].sum()) if 'other_charges_cost' in group.columns else 0
            brokerage = float(group['brokerage_cost'].sum()) if 'brokerage_cost' in group.columns else 0
            
            supplier_analysis['vendor_hidden_costs'].append({
                'supplier': str(supplier_name),
                'other_charges': other_charges,
                'brokerage_cost': brokerage,
                'total_hidden_cost': other_charges + brokerage
            })
            
            credit_days_list = []
            for idx, row in group.iterrows():
                purch_posting = row['purch_posting_date'] if pd.notna(row['purch_posting_date']) else None
                due_date = row['due_date'] if pd.notna(row['due_date']) else None
                if purch_posting and due_date and isinstance(purch_posting, pd.Timestamp) and isinstance(due_date, pd.Timestamp):
                    credit_days = (due_date - purch_posting).days
                    if credit_days > 0:
                        credit_days_list.append(credit_days)
            
            avg_credit_days = np.mean(credit_days_list) if credit_days_list else 0
            max_credit_days = max(credit_days_list) if credit_days_list else 0
            
            supplier_analysis['vendor_credit_days'].append({
                'supplier': str(supplier_name),
                'avg_credit_days': avg_credit_days,
                'max_credit_days': max_credit_days,
                'transaction_count': len(group)
            })
            
            demand_data['suppliers'].append({'name': str(supplier_name), 'sales': sales, 'gp': gp, 'cogs': sales - gp, 'qty': float(group['qty_mt'].sum()), 'gm': gp_percent})
        
        supplier_analysis['vendor_gp'].sort(key=lambda x: x['gp'], reverse=True)
        supplier_analysis['supplier_ps_days'].sort(key=lambda x: x['avg_ps_days'], reverse=True)
        supplier_analysis['loss_making_items'].sort(key=lambda x: x['gp'])
        supplier_analysis['vendor_hidden_costs'].sort(key=lambda x: x['total_hidden_cost'], reverse=True)
        supplier_analysis['vendor_credit_days'].sort(key=lambda x: x['avg_credit_days'], reverse=True)
        
        demand_data['summary']['total_suppliers'] = len(demand_data['suppliers'])
        print(f"   Suppliers: {len(demand_data['suppliers'])}")
        print(f"      Loss-making items: {len(supplier_analysis['loss_making_items'])}")
    
    # Process Countries
    if 'country' in df_demand.columns:
        for name, group in df_demand.groupby('country'):
            sales = float(group['sales_jod'].sum())
            gp = float(group['gp_jod'].sum())
            gp_percent = (gp / sales * 100) if sales > 0 else 0
            demand_data['countries'].append({
                'name': str(name), 
                'sales': sales, 
                'gm': gp, 
                'gp_percent': gp_percent,
                'prev_year_sales': 0,
                'forecast': 0,
                'target': 0,
                'prev_year_gm': 0,
                'forecast_gm': 0,
                'target_gm': 0
            })
        print(f"   Countries: {len(demand_data['countries'])}")
    
    # Categories
    if 'category' in df_demand.columns and not demand_data['categories']:
        for name, group in df_demand.groupby('category'):
            sales = float(group['sales_jod'].sum())
            gp = float(group['gp_jod'].sum())
            demand_data['categories'].append({'name': str(name), 'sales': sales, 'gp': gp, 'gm': (gp / sales * 100) if sales > 0 else 0, 'product_count': group['product_name'].nunique() if 'product_name' in group else 0})
        demand_data['summary']['total_categories'] = len(demand_data['categories'])
        print(f"   Categories: {len(demand_data['categories'])}")
    
    # Sections
    if 'section' in df_demand.columns:
        for name, group in df_demand.groupby('section'):
            sales = float(group['sales_jod'].sum())
            gp = float(group['gp_jod'].sum())
            demand_data['sections'].append({'name': str(name), 'sales': sales, 'gp': gp, 'cogs': sales - gp})
        print(f"   Sections: {len(demand_data['sections'])}")
    
    # Statuses
    if 'status' in df_demand.columns:
        status_counts = df_demand['status'].value_counts().to_dict()
        for name, group in df_demand.groupby('status'):
            sales = float(group['sales_jod'].sum())
            gp = float(group['gp_jod'].sum())
            demand_data['statuses'].append({'name': str(name), 'sales': sales, 'gp': gp, 'count': status_counts.get(name, 0)})
        print(f"   Statuses: {len(demand_data['statuses'])}")
    
    # Sales Trend
    if 'posting_date' in df_demand.columns:
        date_sales = df_demand.groupby(df_demand['posting_date'].dt.date)['sales_jod'].sum().sort_index()
        demand_data['sales_trend']['labels'] = [str(d) for d in date_sales.index]
        demand_data['sales_trend']['values'] = [float(v) for v in date_sales.values]
        print(f"   Sales Trend: {len(demand_data['sales_trend']['labels'])} points")
    
    # Category Sales
    if 'category' in df_demand.columns and not demand_data['category_sales']['labels']:
        cat_sales = df_demand.groupby('category')['sales_jod'].sum().sort_values(ascending=False)
        demand_data['category_sales']['labels'] = [str(c) for c in cat_sales.index[:10]]
        demand_data['category_sales']['values'] = [float(v) for v in cat_sales.values[:10]]
        print(f"   Category Sales: {len(demand_data['category_sales']['labels'])} categories")
    
    # Quarters
    if 'posting_date' in df_demand.columns:
        df_demand['quarter'] = df_demand['posting_date'].dt.to_period('Q')
        for name, group in df_demand.groupby('quarter'):
            sales = float(group['sales_jod'].sum())
            gp = float(group['gp_jod'].sum())
            demand_data['quarters'].append({'name': str(name), 'sales': sales, 'gp': gp, 'gm': (gp / sales * 100) if sales > 0 else 0})
        print(f"   Quarters: {len(demand_data['quarters'])}")
    
    return demand_data, customer_analysis, product_analysis, supplier_analysis

def main():
    print("=" * 70)
    print("Al-Behar Trading Company - Complete S&OP Dashboard")
    print("=" * 70)
    print("Achievement Bars from Summary sheet")
    print("S&OP Cycle tables from Data sheet")
    print("Analysis tabs from Demand Analysis sheet")
    print("=" * 70)
    print("IMPORTANT:")
    print("   - Gross Margin (GM): Calculated from Section = 'Trade'")
    print("   - Commission: Calculated from Section = 'Broker'")
    print("   - Gross Profit (GP) = GM + Commission")
    print("   - COGS = Sales - GM")
    print("   - GM% = (GM / Sales) x 100")
    print("   - GP% = (GP / Sales) x 100")
    print("=" * 70)
    
    df_sop, df_demand, df_summary, file_names, summary_columns = load_excel_data()
    
    if df_sop is None and df_demand is None:
        print("\nERROR: Could not load any Excel data!")
        return
    
    sop_data = process_sop_data(df_sop, file_names)
    demand_data, customer_analysis, product_analysis, supplier_analysis = process_demand_data(df_demand)
    
    achievement_data = []
    if df_summary is not None and len(df_summary) > 0:
        print("\nProcessing Summary sheet...")
        for idx, row in df_summary.iterrows():
            achievement_row = {}
            for col in df_summary.columns:
                achievement_row[str(col).strip()] = row[col]
            achievement_data.append(achievement_row)
        print(f"Loaded {len(achievement_data)} achievement records")
    
    combined_data = {
        'summary': demand_data['summary'],
        'products': demand_data['products'],
        'customers': demand_data['customers'],
        'suppliers': demand_data['suppliers'],
        'categories': demand_data['categories'],
        'sections': demand_data['sections'],
        'quarters': demand_data['quarters'],
        'statuses': demand_data['statuses'],
        'countries': demand_data.get('countries', []),
        'file_data': sop_data['file_data'],
        'filtered_data': sop_data['filtered_data'],
        'commission_data': sop_data['commission_data'],
        'sop_records': sop_data['records'],
        'file_names': file_names,
        'sales_trend': demand_data['sales_trend'],
        'category_sales': demand_data['category_sales'],
        'data_info': {
            'total_rows': (len(df_sop) if df_sop is not None else 0) + (len(df_demand) if df_demand is not None else 0),
            'date_range': 'N/A',
            'source': 'Excel',
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        },
        'achievement': achievement_data,
        'demand_analysis': demand_data,
        'demand_records': demand_data.get('records', []),
        'customer_analysis_data': customer_analysis,
        'product_analysis_data': product_analysis,
        'supplier_analysis_data': supplier_analysis
    }
    
    logo_base64 = get_logo_base64()
    
    print("\nConverting to JSON...")
    chart_data_json = json.dumps(combined_data, cls=NumpyEncoder, ensure_ascii=False)
    
    print("Generating HTML dashboard...")
    html_content = HTML_TEMPLATE_START.replace('{{ chart_data }}', chart_data_json)
    
    if logo_base64:
        html_content = html_content.replace('{{ logo_base64 }}', logo_base64)
        print("Logo embedded")
    else:
        html_content = html_content.replace('{{ logo_base64 }}', '')
    
    output_file = 'al_behar_dashboard_final.html'
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    full_path = os.path.abspath(output_file)
    print(f"\nDashboard created: {full_path}")
    
    try:
        webbrowser.open('file://' + full_path)
        print("   Dashboard opened in browser")
    except:
        print("   Please open the file manually")
    
    print("\n" + "="*70)
    print("DASHBOARD READY")
    print(f"   S&OP Dataset Files: {len(file_names)}")
    print(f"   Filtered Combinations: {len(sop_data['filtered_data'])}")
    print(f"   Commission Data Entries: {len(sop_data['commission_data'])}")
    print(f"   Demand Analysis Records: {combined_data['summary']['total_rows']}")
    print(f"   Customer Analysis: {len(customer_analysis['profitability'])} customers")
    print(f"   Product Analysis: {len(product_analysis['top_products'])} products")
    print(f"   Supplier Analysis: {len(supplier_analysis['vendor_gp'])} suppliers")
    print(f"   Countries: {len(demand_data.get('countries', []))} countries")
    print("="*70)

if __name__ == "__main__":
    main()
